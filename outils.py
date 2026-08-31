import datetime
import functools
import os

import openpyxl


def lister_exports(dossier):
    exports = []
    for nom_fichier in os.listdir(dossier):
        if nom_fichier.startswith("export_") and nom_fichier.endswith(".xlsx"):
            texte_date = nom_fichier[7:17]  # ex: "2025-09-01"
            annee = int(texte_date[0:4])
            mois = int(texte_date[5:7])
            jour = int(texte_date[8:10])
            date_export = datetime.date(annee, mois, jour)
            chemin_complet = os.path.join(dossier, nom_fichier)
            exports.append((date_export, chemin_complet))

    return sorted(exports)


def fichiers_dans_plage(exports_disponibles, date_debut, date_fin):
    fichiers = []
    for date_export, chemin in exports_disponibles:
        if date_debut <= date_export <= date_fin:
            fichiers.append(chemin)
    return fichiers


def charger_periode(fichiers):
    tous_les_tickets = []
    for chemin in fichiers:
        tous_les_tickets.extend(charger_tickets(chemin))
    return tous_les_tickets


def texte_horaires_jour(plages):
    if len(plages) == 0:
        return "-"

    premier_debut, premiere_fin = plages[0]
    texte = str(premier_debut) + "h-" + str(premiere_fin) + "h"

    for i in range(1, len(plages)):
        debut, fin = plages[i]
        texte = texte + " / " + str(debut) + "h-" + str(fin) + "h"

    return texte


def charger_commandes(chemin):
    classeur = openpyxl.load_workbook(chemin, data_only=True)
    feuille = classeur["ORDERS"]

    entetes = []
    for cellule in feuille[1]:
        entetes.append(cellule.value)

    commandes = {}
    for ligne in feuille.iter_rows(min_row=2, values_only=True):
        commande = {}
        for i in range(len(entetes)):
            commande[entetes[i]] = ligne[i]
        commandes[commande["order_id"]] = commande

    return commandes


def charger_couts_produits(chemin):
    classeur = openpyxl.load_workbook(chemin, data_only=True)
    feuille = classeur["COSTS"]

    entetes = []
    for cellule in feuille[1]:
        entetes.append(cellule.value)

    couts = {}
    for ligne in feuille.iter_rows(min_row=2, values_only=True):
        cout = {}
        for i in range(len(entetes)):
            cout[entetes[i]] = ligne[i]
        cle = (cout["product_category"], cout["product_name"])
        couts[cle] = cout

    return couts


def montant_ticket(ticket, commandes):
    order_id = ticket["order_id"]
    if order_id is None or order_id not in commandes:
        return None
    return commandes[order_id]["montant_total"]


def obtenir_commande(ticket, commandes):
    order_id = ticket["order_id"]
    if order_id is None or order_id not in commandes:
        return None
    return commandes[order_id]


# Le montant payé par le client (remboursement intégral) est un coût réel, pas une estimation.
# Un geste commercial cash n'a pas de montant réellement accordé dans les données disponibles
# (aucun champ ticket ne le capture) — reste une fraction estimée du prix de vente, marquée
# comme telle dans l'UI plutôt que mélangée silencieusement aux coûts réels ci-dessous.
FRACTION_GESTE_COMMERCIAL = 0.15


# Coût réel d'un remplacement : coût de revient du produit (au même ratio que le catalogue de
# coûts, appliqué au montant réellement payé pour rester cohérent avec d'éventuelles variantes de
# prix) + coût logistique d'expédition du remplacement + coût de retour si le client renvoie
# l'article défectueux (pas pertinent pour un petit accessoire).
def montant_cout_remplacement(ticket, commandes, couts_produits, avec_retour):
    commande = obtenir_commande(ticket, commandes)
    if commande is None:
        return None

    cle_produit = (commande["product_category"], commande["product_name"])
    if cle_produit not in couts_produits:
        return None

    cout_produit = couts_produits[cle_produit]
    prix_reference = cout_produit["prix_vente_ttc"]
    if prix_reference is None or prix_reference == 0:
        return None

    ratio_cout = cout_produit["cout_revient_produit"] / prix_reference
    montant_reel = commande["montant_total"]

    cout = montant_reel * ratio_cout + cout_produit["cout_logistique_remplacement"]
    if avec_retour:
        cout = cout + cout_produit["cout_retour"]
    return cout


def montant_perte_estime(ticket, commandes, type_perte, couts_produits):
    if type_perte == "Remboursement":
        return montant_ticket(ticket, commandes)
    if type_perte == "Remplacement produit":
        return montant_cout_remplacement(ticket, commandes, couts_produits, True)
    if type_perte == "Remplacement accessoire":
        return montant_cout_remplacement(ticket, commandes, couts_produits, False)
    if type_perte == "Geste commercial":
        montant_commande = montant_ticket(ticket, commandes)
        if montant_commande is None:
            return None
        return montant_commande * FRACTION_GESTE_COMMERCIAL
    return montant_ticket(ticket, commandes)


def montant_cout_garantie(ticket, commandes, couts_produits):
    return montant_cout_remplacement(ticket, commandes, couts_produits, True)


def formater_nombre_espace(nombre_entier):
    texte = str(abs(nombre_entier))

    groupes = []
    while len(texte) > 3:
        groupes.append(texte[-3:])
        texte = texte[:-3]
    groupes.append(texte)
    groupes.reverse()

    resultat = " ".join(groupes)
    if nombre_entier < 0:
        return "-" + resultat
    return resultat


def formater_montant(valeur):
    return formater_nombre_espace(round(valeur)) + " €"


def commandes_par_email(commandes):
    par_email = {}
    for order_id in commandes:
        commande = commandes[order_id]
        email = commande["email_client"]
        if email in par_email:
            par_email[email].append(commande)
        else:
            par_email[email] = [commande]
    return par_email


def premiere_commande_apres(ticket, index_par_email, fenetre_jours):
    email = ticket["requester_email"]
    commandes_client = index_par_email.get(email, [])

    date_debut = ticket["created_at"]
    date_fin = date_debut + datetime.timedelta(days=fenetre_jours)

    candidates = []
    for commande in commandes_client:
        if date_debut < commande["order_date"] <= date_fin:
            candidates.append(commande)

    if len(candidates) == 0:
        return None

    def obtenir_date(commande):
        return commande["order_date"]

    candidates_triees = sorted(candidates, key=obtenir_date)
    return candidates_triees[0]


def tickets_par_email(tickets):
    par_email = {}
    for ticket in tickets:
        email = ticket["requester_email"]
        if email in par_email:
            par_email[email].append(ticket)
        else:
            par_email[email] = [ticket]
    return par_email


# Miroir de premiere_commande_apres, mais en sens inverse : le dernier ticket du même client
# avant une réponse NPS, dans une fenêtre glissante — sert à rapprocher un score de satisfaction
# d'une expérience de contact plausible, sans prétendre à un lien démontré (aucun ID ticket n'est
# stocké dans les réponses NPS elles-mêmes).
def dernier_ticket_avant(reponse, index_tickets_email, fenetre_jours):
    email = reponse["email_client"]
    tickets_client = index_tickets_email.get(email, [])

    date_limite = reponse["date_reponse"]
    date_debut_fenetre = date_limite - datetime.timedelta(days=fenetre_jours)

    candidates = []
    for ticket in tickets_client:
        if date_debut_fenetre <= ticket["created_at"] <= date_limite:
            candidates.append(ticket)

    if len(candidates) == 0:
        return None

    def obtenir_date_ticket(ticket):
        return ticket["created_at"]

    candidates_triees = sorted(candidates, key=obtenir_date_ticket, reverse=True)
    return candidates_triees[0]


def charger_nps(chemin):
    classeur = openpyxl.load_workbook(chemin, data_only=True)
    feuille = classeur["NPS"]

    entetes = []
    for cellule in feuille[1]:
        entetes.append(cellule.value)

    reponses = []
    for ligne in feuille.iter_rows(min_row=2, values_only=True):
        reponse = {}
        for i in range(len(entetes)):
            reponse[entetes[i]] = ligne[i]
        reponses.append(reponse)

    return reponses


def calculer_nps(reponses):
    if len(reponses) == 0:
        return None

    promoteurs = 0
    detracteurs = 0

    for reponse in reponses:
        score = reponse["score"]
        if score >= 9:
            promoteurs = promoteurs + 1
        elif score <= 6:
            detracteurs = detracteurs + 1

    return (promoteurs - detracteurs) / len(reponses) * 100


def extraire_code_macro(texte):
    if texte is None:
        return None

    debut = texte.find("MAC-")
    if debut == -1:
        return None

    fin = debut + 4
    while fin < len(texte) and texte[fin].isdigit():
        fin = fin + 1

    return texte[debut:fin]


def charger_texte_macro(code, dossier_macros):
    if code is None:
        return None

    chemin = os.path.join(dossier_macros, code + ".md")
    if not os.path.exists(chemin):
        return None

    with open(chemin, "r", encoding="utf-8") as fichier:
        return fichier.read()


def extraire_nom_fichier_faq(texte_macro):
    if texte_macro is None:
        return None

    marqueur = "knowledge_base/faq/"
    debut = texte_macro.find(marqueur)
    if debut == -1:
        return None

    debut = debut + len(marqueur)
    fin = debut
    while fin < len(texte_macro) and texte_macro[fin] not in (" ", "\n", "\r"):
        fin = fin + 1

    return texte_macro[debut:fin]


def charger_texte_faq(nom_fichier, dossier_faq):
    if nom_fichier is None:
        return None

    chemin = os.path.join(dossier_faq, nom_fichier)
    if not os.path.exists(chemin):
        return None

    with open(chemin, "r", encoding="utf-8") as fichier:
        return fichier.read()


# ---------------------------------------------------------------------------
# Referentiel evenements / contexte periode (calendrier_evenements.xlsx)
# ---------------------------------------------------------------------------

TYPE_COMMERCIAL = "Commercial"
TYPE_PRODUIT = "Produit"
TYPE_STAFFING = "Staffing"
TYPE_OPERATIONNEL = "Opérationnel"


def _date_depuis_cellule(valeur):
    if isinstance(valeur, datetime.datetime):
        return valeur.date()
    return valeur


# Fusionne les deux feuilles du calendrier (EVENEMENTS : Commercial/Produit, EVENEMENTS_RH :
# Staffing/Operationnel) en une liste unique et normalisee -- meme forme de dict quelle que soit
# la feuille d'origine, pour que contexte_periode() n'ait pas a connaitre la structure des
# feuilles. Une ligne avec une date de debut ou de fin manquante est ignoree (donnee invalide,
# pas une periode ouverte).
def charger_evenements_calendrier(chemin):
    classeur = openpyxl.load_workbook(chemin, data_only=True)
    evenements = []

    if "EVENEMENTS" in classeur.sheetnames:
        feuille = classeur["EVENEMENTS"]
        entetes = [cellule.value for cellule in feuille[1]]
        idx = {nom: i for i, nom in enumerate(entetes)}
        for ligne in feuille.iter_rows(min_row=2, values_only=True):
            date_debut = _date_depuis_cellule(ligne[idx["date_debut"]])
            date_fin = _date_depuis_cellule(ligne[idx["date_fin"]])
            if date_debut is None or date_fin is None:
                continue
            evenements.append({
                "date_debut": date_debut,
                "date_fin": date_fin,
                "type": ligne[idx["type"]],
                "nature": ligne[idx["nature"]] if "nature" in idx else None,
                "nom_evenement": ligne[idx["nom_evenement"]],
                "description": ligne[idx["notes"]] if "notes" in idx else None,
                "perimetre": None,
            })

    if "EVENEMENTS_RH" in classeur.sheetnames:
        feuille = classeur["EVENEMENTS_RH"]
        entetes = [cellule.value for cellule in feuille[1]]
        idx = {nom: i for i, nom in enumerate(entetes)}
        for ligne in feuille.iter_rows(min_row=2, values_only=True):
            date_debut = _date_depuis_cellule(ligne[idx["date_debut"]])
            date_fin = _date_depuis_cellule(ligne[idx["date_fin"]])
            if date_debut is None or date_fin is None:
                continue
            evenements.append({
                "date_debut": date_debut,
                "date_fin": date_fin,
                "type": ligne[idx["type"]],
                "nature": ligne[idx["nature"]] if "nature" in idx else None,
                "nom_evenement": ligne[idx["nom_evenement"]],
                "description": ligne[idx["description"]] if "description" in idx else None,
                "perimetre": ligne[idx["perimetre"]] if "perimetre" in idx else None,
            })

    return evenements


# Contexte structure, jamais une interpretation : renvoie les evenements (Commercial, Produit,
# Staffing, Operationnel confondus) dont la periode chevauche [date_debut, date_fin], tries par
# date de debut puis type puis nom pour un ordre stable. Le chevauchement couvre les 5 cas
# standard (avant->pendant, pendant->apres, pendant->pendant, couvre toute la periode, un seul
# jour) via un simple test d'intersection d'intervalles. Ne calcule ni ne suggere aucun lien de
# cause a effet avec les donnees observees -- ca reste au jugement de qui consomme le resultat.
def contexte_periode(evenements, date_debut, date_fin):
    actifs = []
    for evenement in evenements:
        if evenement["date_debut"] is None or evenement["date_fin"] is None:
            continue
        if evenement["date_fin"] >= date_debut and evenement["date_debut"] <= date_fin:
            actifs.append(evenement)

    def cle_tri(evenement):
        return (evenement["date_debut"], evenement["type"], evenement["nom_evenement"])

    return sorted(actifs, key=cle_tri)


def charger_suivi_suggestions(chemin):
    if not os.path.exists(chemin):
        return {}

    classeur = openpyxl.load_workbook(chemin, data_only=True)
    if "SUIVI" not in classeur.sheetnames:
        return {}

    feuille = classeur["SUIVI"]
    entetes = []
    for cellule in feuille[1]:
        entetes.append(cellule.value)

    if "sujet" not in entetes or "statut" not in entetes or "date_action" not in entetes or "notes" not in entetes:
        return {}

    idx_sujet = entetes.index("sujet")
    idx_statut = entetes.index("statut")
    idx_date = entetes.index("date_action")
    idx_notes = entetes.index("notes")

    suivi = {}
    for ligne in feuille.iter_rows(min_row=2, values_only=True):
        sujet = ligne[idx_sujet]
        if sujet is None:
            continue

        date_action = ligne[idx_date]
        if isinstance(date_action, datetime.datetime):
            date_action = date_action.date()

        suivi[sujet] = {
            "statut": ligne[idx_statut],
            "date_action": date_action,
            "notes": ligne[idx_notes],
        }

    return suivi


def impact_avant_apres(tickets_sujet, date_action):
    avant = []
    apres = []
    for ticket in tickets_sujet:
        date_ticket = ticket["created_at"]
        if isinstance(date_ticket, datetime.datetime):
            date_ticket = date_ticket.date()

        if date_ticket < date_action:
            avant.append(ticket)
        else:
            apres.append(ticket)

    return {
        "volume_avant": len(avant),
        "volume_apres": len(apres),
        "csat_avant": moyenne(avant, "csat"),
        "csat_apres": moyenne(apres, "csat"),
        "macro_avant": taux_rempli(avant, "macro_applied"),
        "macro_apres": taux_rempli(apres, "macro_applied"),
    }


# Mise en cache (Étape 6A, audit performance) : chaque export xlsx est relu jusqu'à 9-10 fois par
# rechargement de page (chaque onglet reconstruit indépendamment son propre historique) -- source de
# coût numéro un mesurée (~67s sur ~102s au chargement complet). Le fichier ne change jamais pendant
# une session (chemin -> contenu strictement stable), et aucun appelant ne mute la liste/les dicts
# retournés (vérifié -- toujours des lectures ou des listes/dicts neufs construits à partir du
# résultat) : la mise en cache ne change donc aucune sortie, seulement le temps de calcul.
@functools.lru_cache(maxsize=None)
def charger_tickets(chemin):
    classeur = openpyxl.load_workbook(chemin, data_only=True)
    feuille = classeur["RAW_TICKETS"]

    entetes = []
    for cellule in feuille[1]:
        entetes.append(cellule.value)

    tickets = []
    for ligne in feuille.iter_rows(min_row=2, values_only=True):
        ticket = {}
        for i in range(len(entetes)):
            ticket[entetes[i]] = ligne[i]
        tickets.append(ticket)

    return tickets


def delai_jours(date_debut, date_fin):
    if date_debut is None or date_fin is None:
        return None
    difference = date_fin - date_debut
    return difference.days


def niveau_anciennete_defaut(jours):
    if jours < 30:
        return "Défaut précoce (< 30j)"
    elif jours < 180:
        return "Défaut intermédiaire (30-180j)"
    else:
        return "Usure normale probable (> 180j)"


def evolution_pourcentage(ancienne_valeur, nouvelle_valeur):
    difference = nouvelle_valeur - ancienne_valeur
    return difference / ancienne_valeur * 100


def moyenne(tickets, champ):
    valeurs = []
    for ticket in tickets:
        if ticket[champ] is not None:
            valeurs.append(ticket[champ])

    if len(valeurs) == 0:
        return None

    return sum(valeurs) / len(valeurs)


def taux_rempli(tickets, champ):
    if len(tickets) == 0:
        return None

    nombre_rempli = 0
    for ticket in tickets:
        if ticket[champ] is not None:
            nombre_rempli = nombre_rempli + 1
    return nombre_rempli / len(tickets) * 100


def evenements_periode(tickets):
    evenements = []
    for ticket in tickets:
        valeur = ticket["evenement_semaine"]
        if valeur is not None and valeur not in evenements:
            evenements.append(valeur)

    if len(evenements) == 0:
        return "Non renseigné"

    texte = evenements[0]
    for i in range(1, len(evenements)):
        texte = texte + " | " + evenements[i]
    return texte


SEUIL_DEBORDEMENT_MIN = 480  # au-delà, on suppose que la réponse a débordé hors créneau

NOM_AGENT_DEFAUT = "DEFAUT"  # horaires utilisés pour un agent absent du planning

# Horaires de secours si le fichier n'a pas d'onglet PLANNING (ex : anciens exports).
# Base 35h/semaine pour un temps plein (9h-17h, 1h de pause dejeuner = 7h/jour x 5 jours).
HORAIRES_PAR_DEFAUT = {
    0: [(9, 12), (13, 17)],
    1: [(9, 12), (13, 17)],
    2: [(9, 12), (13, 17)],
    3: [(9, 12), (13, 17)],
    4: [(9, 12), (13, 17)],
    5: [],
    6: [],
}

JOURS_SEMAINE = {
    "Lundi": 0,
    "Mardi": 1,
    "Mercredi": 2,
    "Jeudi": 3,
    "Vendredi": 4,
    "Samedi": 5,
    "Dimanche": 6,
}


# Même mise en cache que charger_tickets ci-dessus, même justification (fichier stable, jamais muté
# par un appelant).
@functools.lru_cache(maxsize=None)
def charger_planning(chemin):
    classeur = openpyxl.load_workbook(chemin, data_only=True)

    if "PLANNING" not in classeur.sheetnames:
        return {NOM_AGENT_DEFAUT: HORAIRES_PAR_DEFAUT}

    feuille = classeur["PLANNING"]
    entetes = []
    for cellule in feuille[1]:
        entetes.append(cellule.value)
    idx_agent = entetes.index("agent")
    idx_jour = entetes.index("jour")
    idx_debut = entetes.index("heure_debut")
    idx_fin = entetes.index("heure_fin")

    planning = {}

    for ligne in feuille.iter_rows(min_row=2, values_only=True):
        agent = ligne[idx_agent]
        jour_texte = ligne[idx_jour]
        heure_debut = ligne[idx_debut]
        heure_fin = ligne[idx_fin]
        jour_numero = JOURS_SEMAINE.get(jour_texte)
        if jour_numero is None:
            continue

        if agent not in planning:
            planning[agent] = {}
        if jour_numero not in planning[agent]:
            planning[agent][jour_numero] = []

        planning[agent][jour_numero].append((heure_debut, heure_fin))

    return planning


def charger_roles_planning(chemin):
    classeur = openpyxl.load_workbook(chemin, data_only=True)

    if "PLANNING" not in classeur.sheetnames:
        return {}

    feuille = classeur["PLANNING"]
    entetes = []
    for cellule in feuille[1]:
        entetes.append(cellule.value)

    if "role" not in entetes:
        return {}

    idx_agent = entetes.index("agent")
    idx_role = entetes.index("role")

    roles = {}
    for ligne in feuille.iter_rows(min_row=2, values_only=True):
        agent = ligne[idx_agent]
        role = ligne[idx_role]
        if role is not None and agent not in roles:
            roles[agent] = role

    return roles


def detecter_changements_planning(agents_s1, agents_s2, planning_s1, planning_s2):
    changements = []

    for agent in agents_s2:
        if agent not in agents_s1:
            changements.append("Nouvel agent : " + agent)

    for agent in agents_s1:
        if agent not in agents_s2:
            changements.append("Agent absent cette période : " + agent)

    for agent in agents_s2:
        if agent in agents_s1:
            horaires_avant = horaires_agent(planning_s1, agent)
            horaires_apres = horaires_agent(planning_s2, agent)
            if horaires_avant != horaires_apres:
                changements.append("Horaires modifiés : " + agent)

    return changements


def horaires_agent(planning, agent):
    if agent in planning:
        return planning[agent]
    return planning.get(NOM_AGENT_DEFAUT, HORAIRES_PAR_DEFAUT)


# Capacité prévue à un créneau précis : agents dont le PLANNING couvre ce (jour, heure).
# Volontairement pas horaires_agent() ici : sa bascule vers l'horaire DEFAUT pour un agent
# absent du planning gonflerait artificiellement les effectifs des semaines dont le PLANNING
# est incomplet — un agent sans ligne cette semaine-là doit compter pour 0 heure, pas suivre le
# créneau standard.
def agents_en_poste(planning, agents_grille, jour, heure):
    presents = []
    for agent in agents_grille:
        if agent not in planning:
            continue

        plages = planning[agent].get(jour, [])
        for debut, fin in plages:
            if debut <= heure < fin:
                presents.append(agent)
                break
    return presents


# Activité observée = agents ayant réellement créé/traité un ticket dans ce créneau précis,
# indépendamment de ce que dit le planning. Le découpage horaire (heure exacte du ticket) évite
# d'étendre la présence d'un agent au-delà du créneau où son activité a réellement été observée :
# un ticket à 14h03 ne compte que pour le créneau 14h-15h, jamais pour 15h-16h ou toute la
# journée. heure_debut/heure_fin bornent la plage suivie (ex. 7h-21h pour la heatmap Couverture).
def construire_activite_par_jour_heure(tickets, heure_debut, heure_fin):
    activite = {}
    for jour in range(7):
        activite[jour] = {}
        for heure in range(heure_debut, heure_fin):
            activite[jour][heure] = set()

    for ticket in tickets:
        moment = ticket["created_at"]
        jour_ticket = moment.weekday()
        heure_ticket = moment.hour
        if jour_ticket in activite and heure_ticket in activite[jour_ticket]:
            if ticket["assignee"] is not None:
                activite[jour_ticket][heure_ticket].add(ticket["assignee"])

    return activite


def activite_observee(activite_par_jour_heure, jour, heure):
    return activite_par_jour_heure.get(jour, {}).get(heure, set())


# Jamais mélangé au dénominateur de charge (agents_en_poste reste seul dans ce calcul) — sert
# uniquement à détecter un renfort ponctuel non planifié : un agent actif sur ce créneau précis
# mais absent de la capacité prévue. Fonction générique et pure, appelable pour n'importe quel
# nombre d'agents planifiés (0, 1, 2, 3...) — c'est à l'appelant de décider sur quels créneaux
# ce signal a un sens opérationnel (ex. seulement les créneaux "Couverture requise").
def renfort_non_planifie(agents_prevus, agents_actifs):
    return sorted(set(agents_actifs) - set(agents_prevus))


def construire_plannings_periode(fichiers, exports_disponibles):
    dates_par_fichier = {}
    for date_export, chemin in exports_disponibles:
        dates_par_fichier[chemin] = date_export

    plannings_periode = []
    for chemin in fichiers:
        date_debut_fichier = dates_par_fichier[chemin]
        date_fin_fichier = date_debut_fichier + datetime.timedelta(days=6)
        planning_fichier = charger_planning(chemin)
        plannings_periode.append((date_debut_fichier, date_fin_fichier, planning_fichier))

    return plannings_periode


def planning_pour_date(plannings_periode, date_cible):
    for date_debut, date_fin, planning in plannings_periode:
        if date_debut <= date_cible <= date_fin:
            return planning

    if len(plannings_periode) > 0:
        return plannings_periode[-1][2]

    return {NOM_AGENT_DEFAUT: HORAIRES_PAR_DEFAUT}


def horaires_agent_periode(plannings_periode, agent, date_cible):
    planning_actif = planning_pour_date(plannings_periode, date_cible)
    return horaires_agent(planning_actif, agent)


def type_creneau(moment, agent, plannings_periode):
    plages_du_jour = horaires_agent_periode(plannings_periode, agent, moment.date()).get(moment.weekday(), [])

    for debut, fin in plages_du_jour:
        if debut <= moment.hour < fin:
            return "En créneau"

    if len(plages_du_jour) > 0:
        premiere_plage_debut = plages_du_jour[0][0]
        derniere_plage_fin = plages_du_jour[-1][1]
        if premiere_plage_debut <= moment.hour < derniere_plage_fin:
            return "Pause déjeuner"

    return "Hors créneau (soir/nuit/week-end)"


def type_hors_creneau_detaille(moment, agent, plannings_periode):
    jour = moment.weekday()

    if jour == 5:
        return "Samedi"
    elif jour == 6:
        return "Dimanche"

    plages_du_jour = horaires_agent_periode(plannings_periode, agent, moment.date()).get(jour, [])

    if len(plages_du_jour) == 0:
        return "Jour sans couverture"

    premiere_plage_debut = plages_du_jour[0][0]
    derniere_plage_fin = plages_du_jour[-1][1]

    if moment.hour < premiere_plage_debut:
        return "Avant l'ouverture"
    elif moment.hour >= derniere_plage_fin:
        return "Après la fermeture"
    else:
        return "Pause déjeuner"


def separer_creneau(tickets, plannings_periode):
    en_creneau = []
    pause_dejeuner = []
    hors_creneau = []

    for ticket in tickets:
        type_du_ticket = type_creneau(ticket["created_at"], ticket["assignee"], plannings_periode)
        if type_du_ticket == "En créneau":
            en_creneau.append(ticket)
        elif type_du_ticket == "Pause déjeuner":
            pause_dejeuner.append(ticket)
        else:
            hors_creneau.append(ticket)

    return en_creneau, pause_dejeuner, hors_creneau


SEUIL_SLA_EN_CRENEAU_MIN = 60  # règle : 1re réponse en 1h max quand le ticket arrive en créneau


def echeance_sla(date_creation, agent, plannings_periode):
    jour_candidat = date_creation.date()
    heure_reference = date_creation.hour

    for _ in range(14):  # sécurité : on ne cherche pas plus de 2 semaines devant
        horaires = horaires_agent_periode(plannings_periode, agent, jour_candidat)
        plages = horaires.get(jour_candidat.weekday(), [])

        for debut, fin in plages:
            deja_passee = jour_candidat == date_creation.date() and heure_reference >= fin
            if not deja_passee:
                return datetime.datetime.combine(jour_candidat, datetime.time(fin, 0))

        jour_candidat = jour_candidat + datetime.timedelta(days=1)
        heure_reference = 0

    return None


def sla_respecte(ticket, plannings_periode):
    minutes = ticket["first_reply_time_min"]
    if minutes is None:
        return None

    created_at = ticket["created_at"]
    agent = ticket["assignee"]

    if type_creneau(created_at, agent, plannings_periode) == "En créneau":
        return minutes <= SEUIL_SLA_EN_CRENEAU_MIN

    echeance = echeance_sla(created_at, agent, plannings_periode)
    if echeance is None:
        return None

    moment_reponse = created_at + datetime.timedelta(minutes=minutes)
    return moment_reponse <= echeance


def taux_sla(tickets, plannings_periode):
    respectes = 0
    total = 0

    for ticket in tickets:
        resultat = sla_respecte(ticket, plannings_periode)
        if resultat is not None:
            total = total + 1
            if resultat:
                respectes = respectes + 1

    if total == 0:
        return None

    return respectes / total * 100


def niveau_reponse_ouvree(minutes):
    if minutes > SEUIL_DEBORDEMENT_MIN:
        return "DEBORDEMENT"
    elif minutes <= SEUIL_SLA_EN_CRENEAU_MIN:
        return "OK"
    elif minutes < 120:
        return "A SURVEILLER"
    else:
        return "CRITIQUE"


def niveau_macro(valeur):
    if valeur is None:
        return ""
    if valeur < 50:
        return "CRITIQUE"
    elif valeur < 70:
        return "A SURVEILLER"
    else:
        return "OK"


# niveau_charge_agent / niveau_charge_creneau (seuils absolus 15/30 demandes-par-agent, jamais
# atteints en pratique -- max réel observé 19 sur 14 exports, voir audit 5E) supprimées Étape
# 5E.1 : remplacées par la taxonomie pression/tension relative à l'historique disponible, voir
# section "Composition Couverture -- pression / tension" plus bas (niveau_pression_couverture,
# creneau_est_tension_couverture). Un "créneau hors couverture" (fermé par conception) reste géré
# séparément, jamais mélangé à la pression -- même principe qu'avant.


def couleur_niveau(valeur):
    if valeur in ("OK", "CORRECT", "Correct", "EXCELLENT", "Excellent", "En créneau", "Fort potentiel"):
        return "background-color: #D9EDDD"
    elif valeur in ("A SURVEILLER", "À surveiller", "Potentiel moyen"):
        return "background-color: #F7E2B8"
    elif valeur in ("CRITIQUE", "Critique", "DEBORDEMENT", "Débordement", "Risque de perte du prospect"):
        return "background-color: #F3D2CB"
    elif valeur in ("NOUVEAU", "Nouveau"):
        return "background-color: #D3E1F0"
    elif valeur in ("DISPARU", "Disparu"):
        return "background-color: #E4E1DB"
    else:
        return ""


LIBELLES_NIVEAUX = {
    "CORRECT": "Correct",
    "EXCELLENT": "Excellent",
    "A SURVEILLER": "À surveiller",
    "CRITIQUE": "Critique",
    "DEBORDEMENT": "Débordement",
    "NOUVEAU": "Nouveau",
    "DISPARU": "Disparu",
}


def libelle_niveau(valeur):
    return LIBELLES_NIVEAUX.get(valeur, valeur)


def grouper_par(tickets, champ):
    groupes = {}
    for ticket in tickets:
        cle = ticket[champ]
        if cle in groupes:
            groupes[cle].append(ticket)
        else:
            groupes[cle] = [ticket]
    return groupes


def cles_combinees(dict_actuel, dict_precedent):
    cles = []
    for cle in dict_actuel:
        cles.append(cle)
    for cle in dict_precedent:
        if cle not in cles:
            cles.append(cle)
    return cles


def grouper_par_categorie(tickets):
    par_categorie = {}
    for ticket in tickets:
        categorie = categoriser(ticket)
        if categorie in par_categorie:
            par_categorie[categorie].append(ticket)
        else:
            par_categorie[categorie] = [ticket]
    return par_categorie


CATEGORIE_SAV_PRODUIT = "SAV produit (défaut)"


# NB : la colonne "is_sav" du fichier source n'est pas utilisée ici — elle est
# redondante avec ticket_reason == "SAV" et cette fonction affine encore la
# distinction SAV usage / SAV produit à partir de resolution_type.
def categoriser(ticket):
    raison = ticket["ticket_reason"]

    if raison in ("Livraison", "Suivi commande"):
        return "Livraison"
    elif raison in ("Abonnement / paiement", "Retour / remboursement"):
        return "Après-vente commande/admin"
    elif raison == "Conseil programme / produit":
        return "Avant-vente / conseil"
    elif raison == "Utilisation / routine":
        return "SAV usage (besoin d'aide)"
    elif raison == "SAV":
        if ticket["resolution_type"] == "Information / résolution à distance":
            return "SAV usage (besoin d'aide)"
        else:
            return CATEGORIE_SAV_PRODUIT
    else:
        return "Autre"


def niveau_csat(valeur):
    if valeur is None:
        return ""
    if valeur <= 3:
        return "CRITIQUE"
    elif valeur < 4:
        return "A SURVEILLER"
    elif valeur < 5:
        return "CORRECT"
    else:
        return "EXCELLENT"


# Le CSAT n'a pas besoin d'une colonne "Niveau" séparée pour être lisible : la couleur
# directement sur le chiffre suffit (contrairement aux autres niveaux — macro, réponse —
# qui restent en fond de cellule via couleur_niveau, cf. afficher_tableau_colore).
COULEUR_TEXTE_CSAT = {
    "CRITIQUE": "#B23A2E",
    "A SURVEILLER": "#9A6B00",
    "CORRECT": "#1E7A42",
    "EXCELLENT": "#1E7A42",
}


def couleur_texte_csat(texte_csat):
    if texte_csat is None or texte_csat == "N/A":
        return ""

    # Certaines colonnes "CSAT" affichent une comparaison ("4.15 → 3.97") plutôt qu'un
    # chiffre unique — pas de couleur dans ce cas, seul un CSAT unique se colore.
    try:
        valeur = float(texte_csat)
    except ValueError:
        return ""

    niveau = niveau_csat(valeur)
    couleur = COULEUR_TEXTE_CSAT.get(niveau)
    if couleur is None:
        return ""
    return "color: " + couleur + "; font-weight: 600"




def type_perte_financiere(ticket):
    resolution = ticket["resolution_type"]

    if resolution == "Remboursement":
        return "Remboursement"
    elif resolution == "Remplacement produit":
        return "Remplacement produit"
    elif resolution == "Remplacement appareil / accessoire":
        return "Remplacement accessoire"
    elif resolution == "Geste commercial":
        return "Geste commercial"
    else:
        return None


MARQUEUR_HORS_CATALOGUE = "(hors catalogue)"


def detecter_opportunites_hors_catalogue(tickets, seuil):
    par_sujet = {}
    for ticket in tickets:
        sujet = ticket["subject_cluster"]
        if sujet is not None and MARQUEUR_HORS_CATALOGUE in sujet:
            if sujet in par_sujet:
                par_sujet[sujet].append(ticket)
            else:
                par_sujet[sujet] = [ticket]

    opportunites = []
    for sujet, tickets_sujet in par_sujet.items():
        if len(tickets_sujet) >= seuil:
            opportunites.append((sujet, tickets_sujet))

    return opportunites


def formater_pourcentage(valeur):
    if valeur is None:
        return "N/A"
    return str(round(valeur)) + " %"


def formater_csat(valeur):
    if valeur is None:
        return "N/A"
    texte_deux_decimales = "{:.2f}".format(valeur)
    return texte_deux_decimales.replace(".", ",")


def formater_duree(minutes):
    if minutes is None:
        return "N/A"

    minutes = round(minutes)
    jours = minutes // 1440
    heures = (minutes % 1440) // 60
    minutes_restantes = minutes % 60

    if jours > 0:
        return str(jours) + "j " + str(heures) + "h"
    elif heures > 0:
        return str(heures) + "h " + str(minutes_restantes) + "min"
    else:
        return str(minutes_restantes) + "min"


# ---------------------------------------------------------------------------
# Moteur de priorisation Produit (Étape 4A / 4A.1)
#
# Deux voies indépendantes, jamais fusionnées dans un score unique affiché :
#   - Voie A : analyse multicritère sur des candidats agrégés (composant, produit,
#     ou produit x nature du problème), avec seuil minimum d'évaluation et une
#     règle d'éligibilité multicritère -- aucune dimension isolée (volume, part,
#     écart temporel, concentration, CSAT ponctuel, coût isolé) ne suffit seule à
#     faire d'un candidat un signal à remonter. CANDIDAT ANALYSABLE != SIGNAL.
#   - Voie B : repérage heuristique de dossiers individuels graves, sans seuil
#     de volume, nécessitant une combinaison de facteurs (jamais un seul champ).
# Le score interne (voie A) ne sert qu'au classement interne des signaux déjà
# éligibles ; l'UI ne montre que des niveaux qualitatifs (dérivés du nombre de
# familles de preuve convergentes, jamais d'un seuil de score arbitraire) et des
# observations explicables.
# ---------------------------------------------------------------------------

SEUIL_MINIMUM_EVALUATION_PRODUIT = 5  # même valeur que SEUIL_MINIMUM_SUJET (app.py), concept partagé

GRAIN_COMPOSANT = "composant"
GRAIN_PRODUIT_COMPOSANT = "produit_composant"
GRAIN_PRODUIT_ISSUE = "produit_issue"

SEUIL_ECART_RELATIF_TEMPORALITE = 0.20
SEUIL_CONCENTRATION_NOTABLE = 0.55
SEUIL_CSAT_ECART_MARQUE_ABS = 0.4  # points d'écart sur l'échelle 1-5
SEUIL_EFFORT_ECART_MARQUE_REL = 0.30  # écart relatif sur replies / résolution / réouvertures

# Seuils d'éligibilité par famille de preuve (CANDIDAT ANALYSABLE != SIGNAL A REMONTER).
SEUIL_VOLUME_ABSOLU_NOTABLE = 15  # tickets -- ampleur opérationnelle réelle
SEUIL_VOLUME_PART_NOTABLE = 0.15  # part du SAV produit de la période
SEUIL_CONFIANCE_TEMPORELLE_MINIMALE = 0.5  # sous ce seuil, l'écart temporel ne compte dans aucune famille
SEUIL_PART_ELEVEES_PERSISTANCE = 0.5
NB_OBSERVATIONS_MIN_PERSISTANCE = 2

SEUIL_VOIE_B_CSAT_MAX = 2
SEUIL_VOIE_B_REOPENS_MIN = 2
SEUIL_VOIE_B_REPLIES_MULTIPLICATEUR = 2.5
SEUIL_VOIE_B_RESOLUTION_MULTIPLICATEUR = 3
RESOLUTIONS_GRAVES_VOIE_B = ("Remboursement", "Remplacement produit")

POIDS_SCORE_TEMPORALITE = 0.15
POIDS_SCORE_PERSISTANCE = 0.10
POIDS_SCORE_CSAT = 0.20
POIDS_SCORE_EFFORT = 0.15
POIDS_SCORE_COUT = 0.15
POIDS_SCORE_VOLUME_RELATIF = 0.10
POIDS_SCORE_VOLUME_ABSOLU = 0.10
POIDS_SCORE_CONCENTRATION = 0.05

PLAFOND_SCORE_TEMPORALITE = 1.0
PLAFOND_SCORE_CSAT = 2.0
PLAFOND_SCORE_EFFORT = 1.0
PLAFOND_SCORE_VOLUME_RELATIF = 0.30
PLAFOND_SCORE_VOLUME_ABSOLU = 40.0  # tickets -- au-delà, le volume absolu ne fait plus grimper le score


def _moyenne_liste(valeurs):
    if len(valeurs) == 0:
        return None
    somme = 0.0
    for valeur in valeurs:
        somme = somme + valeur
    return somme / len(valeurs)


def _compte_valeurs_non_nulles(tickets, champ):
    compte = 0
    for ticket in tickets:
        if ticket[champ] is not None:
            compte = compte + 1
    return compte


def _normaliser(valeur, plafond):
    if valeur is None or valeur <= 0:
        return 0.0
    ratio = valeur / plafond
    if ratio > 1.0:
        return 1.0
    return ratio


# Facteur de confiance selon le nombre d'observations historiques disponibles : un historique
# pauvre (0-1 export) ne peut jamais donner un avantage temporel maximal -- il faut au moins 2
# observations pour que l'écart temporel compte dans une famille de preuve (voir
# SEUIL_CONFIANCE_TEMPORELLE_MINIMALE), et la confiance ne devient pleine qu'à partir de 4
# observations disponibles.
def confiance_historique(nb_observations):
    if nb_observations <= 0:
        return 0.0
    if nb_observations == 1:
        return 0.3
    if nb_observations == 2:
        return 0.55
    if nb_observations == 3:
        return 0.75
    return 1.0


def calculer_part(sous_ensemble, univers):
    if univers == 0:
        return None
    return sous_ensemble / univers


# Coût réutilisant exactement les primitives et le motif de dé-duplication par order_id déjà
# établis dans l'onglet Impact & confiance (app.py) -- aucune nouvelle logique de coût ici.
# Retourne (montant_total, nombre_de_dossiers_costes) -- le nombre de dossiers permet à l'appelant
# de calculer un coût MOYEN par dossier, seule mesure réellement indépendante du volume.
def calculer_cout_candidat(tickets, commandes, couts_produits):
    total = 0.0
    commandes_comptees = set()

    for ticket in tickets:
        type_perte = type_perte_financiere(ticket)
        if type_perte is None:
            continue
        order_id = ticket["order_id"]
        if order_id in commandes_comptees:
            continue
        montant = montant_perte_estime(ticket, commandes, type_perte, couts_produits)
        if montant is None:
            continue
        total = total + montant
        commandes_comptees.add(order_id)

    if len(commandes_comptees) == 0:
        return None, 0
    return total, len(commandes_comptees)


# Référence de coût relative à l'entreprise observée elle-même (coût moyen par dossier costé sur
# l'ensemble du SAV produit de la période), jamais un seuil absolu codé en dur -- le moteur reste
# ainsi portable vers une autre échelle de prix sans recalibration manuelle.
def calculer_reference_cout_moyen(tickets_univers_periode, commandes, couts_produits):
    total, nb_dossiers = calculer_cout_candidat(tickets_univers_periode, commandes, couts_produits)
    if nb_dossiers == 0:
        return None
    return total / nb_dossiers


# Détection de concentration générique (aucun nom de produit ni de composant codé en dur) :
# quelle part des tickets d'un candidat porte sur une seule valeur du champ donné. Sert à la fois
# à détecter une concentration produit dans un signal composant, et à retrouver le composant
# dominant d'un candidat produit lors de la consolidation (aucune table codée en dur).
def calculer_concentration(tickets, champ):
    compte = {}
    for ticket in tickets:
        valeur = ticket[champ]
        if valeur is None:
            continue
        if valeur in compte:
            compte[valeur] = compte[valeur] + 1
        else:
            compte[valeur] = 1

    if len(compte) == 0:
        return None, 0.0, 0

    total = 0
    for nombre in compte.values():
        total = total + nombre

    dominant = None
    maximum = -1
    for cle in compte:
        if compte[cle] > maximum:
            maximum = compte[cle]
            dominant = cle

    return dominant, maximum / total, total


# Écart relatif du niveau actuel (une part 0-1, jamais un volume brut -- comparable même si les
# exports n'ont pas le même total) par rapport à la moyenne des observations historiques
# disponibles. Chaque export est une observation discontinue parmi d'autres, jamais une
# comparaison avant/après à deux points -- c'est l'esprit "baseline propre" demandé.
def ecart_relatif_temporel(niveaux_historiques, niveau_actuel):
    moyenne_hist = _moyenne_liste(niveaux_historiques)
    if moyenne_hist is None or moyenne_hist <= 0:
        return None
    return (niveau_actuel - moyenne_hist) / moyenne_hist


# Sépare le calcul de persistance (utile à la fois au texte de temporalité et à la famille de
# preuve "E. Récurrence/persistance") : un niveau actuel élevé ne compte comme "persistant" que
# si une majorité des observations historiques disponibles étaient déjà élevées elles aussi, et
# seulement à partir de NB_OBSERVATIONS_MIN_PERSISTANCE observations -- jamais sur un historique
# à 1 point.
def evaluer_persistance_temporelle(niveaux_historiques, niveau_actuel):
    ecart = ecart_relatif_temporel(niveaux_historiques, niveau_actuel)
    nb_observations = len(niveaux_historiques)

    if ecart is None or ecart < SEUIL_ECART_RELATIF_TEMPORALITE:
        return False, 0.0, nb_observations

    moyenne_hist = _moyenne_liste(niveaux_historiques)
    nb_elevees = 0
    for valeur in niveaux_historiques:
        ecart_valeur = (valeur - moyenne_hist) / moyenne_hist
        if ecart_valeur >= SEUIL_ECART_RELATIF_TEMPORALITE:
            nb_elevees = nb_elevees + 1

    part_elevees = 0.0
    if nb_observations > 0:
        part_elevees = nb_elevees / nb_observations

    return True, part_elevees, nb_observations


# Vocabulaire volontairement prudent : jamais "significatif"/"structurel" comme verdict
# automatique, toujours "seuil minimum d'évaluation"/nombre d'observations explicite, et une
# répétition apparente reste "à confirmer" plutôt qu'établie.
def evaluer_temporalite(niveaux_historiques, niveau_actuel):
    if len(niveaux_historiques) == 0:
        return "historique insuffisant (aucune observation antérieure disponible pour ce sujet)"

    ecart = ecart_relatif_temporel(niveaux_historiques, niveau_actuel)
    if ecart is None:
        return "niveau habituel, référence historique trop faible pour comparer utilement"

    if abs(ecart) < SEUIL_ECART_RELATIF_TEMPORALITE:
        return (
            "niveau habituel, cohérent avec les " + str(len(niveaux_historiques))
            + " observation(s) disponible(s)"
        )

    if ecart < 0:
        return "niveau plus bas que d'habitude sur cette observation — possible retour vers un niveau habituel"

    _eleve, part_elevees, nb_observations = evaluer_persistance_temporelle(niveaux_historiques, niveau_actuel)
    nb_elevees = round(part_elevees * nb_observations)

    if part_elevees >= SEUIL_PART_ELEVEES_PERSISTANCE and nb_observations >= NB_OBSERVATIONS_MIN_PERSISTANCE:
        return (
            "niveau inhabituel sur cette observation ; un niveau comparable a déjà été observé sur "
            + str(nb_elevees) + " des " + str(nb_observations) + " période(s) disponible(s) "
            "— répétition à confirmer, pas encore établie comme durable"
        )

    return (
        "niveau inhabituel sur cette observation, pas observé aussi haut sur les "
        + str(nb_observations) + " période(s) disponible(s) — signal ponctuel pour l'instant"
    )


VOCABULAIRE_ECART_INDISPONIBLE = "référence indisponible sur cette période"
VOCABULAIRE_ECART_PROCHE = "proche ou meilleur que la référence observée"
VOCABULAIRE_ECART_BAS = "plus bas que la référence observée"
VOCABULAIRE_ECART_HAUT = "plus élevé que la référence observée"
VOCABULAIRE_ECART_MARQUE = "écart marqué par rapport à la référence observée"


def lire_ecart_csat(csat_candidat, csat_reference):
    if csat_candidat is None or csat_reference is None:
        return VOCABULAIRE_ECART_INDISPONIBLE

    ecart = csat_reference - csat_candidat  # positif si le candidat est moins bien noté
    if abs(ecart) < 0.05:
        return VOCABULAIRE_ECART_PROCHE
    if ecart >= SEUIL_CSAT_ECART_MARQUE_ABS:
        return VOCABULAIRE_ECART_MARQUE
    if ecart > 0:
        return VOCABULAIRE_ECART_BAS
    return VOCABULAIRE_ECART_PROCHE


def lire_ecart_effort(valeur_candidat, valeur_reference):
    if valeur_candidat is None or valeur_reference is None or valeur_reference == 0:
        return VOCABULAIRE_ECART_INDISPONIBLE

    ecart_relatif = (valeur_candidat - valeur_reference) / valeur_reference
    if ecart_relatif <= 0.05:
        return VOCABULAIRE_ECART_PROCHE
    if ecart_relatif >= SEUIL_EFFORT_ECART_MARQUE_REL:
        return VOCABULAIRE_ECART_MARQUE
    return VOCABULAIRE_ECART_HAUT


def texte_sujet_candidat(cle, grain):
    if grain == GRAIN_PRODUIT_ISSUE:
        return cle[0] + " — " + cle[1]
    if grain == GRAIN_PRODUIT_COMPOSANT:
        return cle[0]
    return cle


def tickets_correspondant_candidat(tickets, cle, grain):
    resultat = []
    for ticket in tickets:
        if grain == GRAIN_COMPOSANT:
            if ticket["component"] == cle:
                resultat.append(ticket)
        elif grain == GRAIN_PRODUIT_COMPOSANT:
            if ticket["product_name"] == cle[0] and ticket["component"] == cle[1]:
                resultat.append(ticket)
        else:
            if ticket["product_name"] == cle[0] and ticket["issue_type"] == cle[1]:
                resultat.append(ticket)
    return resultat


# Trois grains de candidats générés en parallèle, jamais un seul :
#   - composant : capte les signaux transverses (ex. un composant sur plusieurs produits) ;
#   - produit x composant : capte un problème propre à UN produit ET une seule famille
#     fonctionnelle -- le champ "component" RÉEL de chaque ticket (jamais dérivé/deviné, jamais un
#     nom de produit ou de composant testé en dur). C'est ce grain qui consolide plusieurs
#     issue_types appartenant à la même histoire produit (même produit, même component), et qui
#     empêche à l'inverse de fusionner deux problèmes du même produit portant sur des composants
#     différents (voir "identité vs preuve", Étape 4A.3) ;
#   - produit x nature du problème : grain le plus fin, utilisé en repli quand le grain
#     produit x composant n'est lui-même pas éligible (deux problèmes distincts sur un même
#     produit ET un même composant, dont la combinaison ne forme pas, elle, un signal cohérent).
def generer_candidats_composant(tickets_sav_produit):
    groupes = grouper_par(tickets_sav_produit, "component")
    candidats = []
    for composant, tickets_candidat in groupes.items():
        if composant is None:
            continue
        candidats.append((composant, GRAIN_COMPOSANT, tickets_candidat))
    return candidats


def generer_candidats_produit_composant(tickets_sav_produit):
    groupes = {}
    for ticket in tickets_sav_produit:
        produit = ticket["product_name"]
        composant = ticket["component"]
        if produit is None or composant is None:
            continue
        cle = (produit, composant)
        if cle in groupes:
            groupes[cle].append(ticket)
        else:
            groupes[cle] = [ticket]

    candidats = []
    for cle, tickets_candidat in groupes.items():
        candidats.append((cle, GRAIN_PRODUIT_COMPOSANT, tickets_candidat))
    return candidats


def generer_candidats_produit_issue(tickets_sav_produit):
    groupes = {}
    for ticket in tickets_sav_produit:
        produit = ticket["product_name"]
        issue = ticket["issue_type"]
        if produit is None or issue is None:
            continue
        cle = (produit, issue)
        if cle in groupes:
            groupes[cle].append(ticket)
        else:
            groupes[cle] = [ticket]

    candidats = []
    for cle, tickets_candidat in groupes.items():
        candidats.append((cle, GRAIN_PRODUIT_ISSUE, tickets_candidat))
    return candidats


# Chaque export historique est traité comme une observation isolée (part du candidat dans le
# SAV produit de ce fichier), jamais combiné en une série continue -- cohérent avec le fait que
# les exports sont des semaines représentatives espacées, pas un historique hebdomadaire complet.
def construire_niveaux_historiques(historique_sav_produit_par_fichier, cle, grain):
    niveaux = []
    for tickets_fichier in historique_sav_produit_par_fichier:
        if len(tickets_fichier) == 0:
            continue
        tickets_candidat_fichier = tickets_correspondant_candidat(tickets_fichier, cle, grain)
        part = calculer_part(len(tickets_candidat_fichier), len(tickets_fichier))
        if part is not None:
            niveaux.append(part)
    return niveaux


def _phrase_elements(elements):
    if len(elements) == 0:
        return "Aucun écart marqué détecté sur cette période."

    texte = elements[0]
    for i in range(1, len(elements)):
        texte = texte + ", " + elements[i]

    return texte[0].upper() + texte[1:] + "."


NOMS_FAMILLES_PREUVE = {
    "A": "Demande / volume",
    "B": "Expérience client",
    "C": "Effort Care",
    "D": "Impact financier",
    "E": "Récurrence / persistance",
    "F": "Concentration",
}


# Évaluateur central d'un candidat Voie A, quel que soit son grain. Retourne None si le seuil
# minimum d'évaluation n'est pas atteint (jamais de carte sur un échantillon trop faible).
#
# "eligible" décide si le candidat mérite ne serait-ce qu'une mention -- et AUCUNE famille de
# preuve active seule ne suffit JAMAIS, même très marquée (volume énorme, écart temporel énorme,
# concentration forte...). Mais atteindre 2 familles ne suffit plus non plus à en faire une
# PRIORITÉ : Demande/volume (A) et Impact financier (D) sont des familles "de contexte", pas des
# preuves de conséquence client -- une vraie priorité exige toujours qu'Expérience (B) ou Effort
# (C) soit l'une des familles actives. Sans B ni C : le candidat reste "à surveiller" (ou, pour la
# combinaison structurelle volume + persistance + concentration, explicitement plafonné là) --
# jamais une carte principale sur la seule base "il y en a beaucoup et ça coûte cher".
def construire_signal_produit_voie_a(cle, grain, tickets_candidat, tickets_univers_periode,
                                      niveaux_historiques, commandes, couts_produits, reference_cout_moyen):
    n = len(tickets_candidat)
    if n < SEUIL_MINIMUM_EVALUATION_PRODUIT:
        return None

    total_univers = len(tickets_univers_periode)
    part_candidat = calculer_part(n, total_univers)
    ecart_temporel = ecart_relatif_temporel(niveaux_historiques, part_candidat)
    texte_temporalite = evaluer_temporalite(niveaux_historiques, part_candidat)
    confiance = confiance_historique(len(niveaux_historiques))
    _persistance_brute, part_elevees, nb_observations = evaluer_persistance_temporelle(
        niveaux_historiques, part_candidat
    )

    temporel_compte = (
        ecart_temporel is not None
        and ecart_temporel >= SEUIL_ECART_RELATIF_TEMPORALITE
        and confiance >= SEUIL_CONFIANCE_TEMPORELLE_MINIMALE
    )

    csat_candidat = moyenne(tickets_candidat, "csat")
    n_csat_candidat = _compte_valeurs_non_nulles(tickets_candidat, "csat")
    csat_reference = moyenne(tickets_univers_periode, "csat")
    n_csat_reference = _compte_valeurs_non_nulles(tickets_univers_periode, "csat")
    lecture_csat = lire_ecart_csat(csat_candidat, csat_reference)
    ecart_csat = None
    if csat_candidat is not None and csat_reference is not None:
        ecart_csat = csat_reference - csat_candidat

    replies_candidat = moyenne(tickets_candidat, "replies")
    replies_reference = moyenne(tickets_univers_periode, "replies")
    lecture_replies = lire_ecart_effort(replies_candidat, replies_reference)

    resolution_candidat = moyenne(tickets_candidat, "full_resolution_time_hours")
    resolution_reference = moyenne(tickets_univers_periode, "full_resolution_time_hours")
    lecture_resolution = lire_ecart_effort(resolution_candidat, resolution_reference)

    reopens_candidat = moyenne(tickets_candidat, "reopens")
    reopens_reference = moyenne(tickets_univers_periode, "reopens")
    lecture_reopens = lire_ecart_effort(reopens_candidat, reopens_reference)

    ecart_effort_max = 0.0
    for couple_effort in (
        (replies_candidat, replies_reference),
        (resolution_candidat, resolution_reference),
        (reopens_candidat, reopens_reference),
    ):
        valeur_candidat_effort, valeur_reference_effort = couple_effort
        if valeur_candidat_effort is not None and valeur_reference_effort is not None and valeur_reference_effort > 0:
            ecart = (valeur_candidat_effort - valeur_reference_effort) / valeur_reference_effort
            if ecart > ecart_effort_max:
                ecart_effort_max = ecart

    cout_montant, nb_dossiers_costes = calculer_cout_candidat(tickets_candidat, commandes, couts_produits)
    cout_info = None
    cout_moyen_dossier = None
    lecture_cout = VOCABULAIRE_ECART_INDISPONIBLE
    if cout_montant is not None:
        cout_moyen_dossier = cout_montant / nb_dossiers_costes
        lecture_cout = lire_ecart_effort(cout_moyen_dossier, reference_cout_moyen)
        cout_info = {
            "montant": cout_montant, "moyen_par_dossier": cout_moyen_dossier, "n": nb_dossiers_costes,
            "reference_moyen_par_dossier": reference_cout_moyen, "lecture": lecture_cout,
            "methode": "Coût de revient réel / estimation (voir Impact & confiance)",
        }

    concentration_info = None
    if grain == GRAIN_COMPOSANT:
        dominant, part_dominant, total_avec_produit = calculer_concentration(tickets_candidat, "product_name")
        if dominant is not None and total_avec_produit >= SEUIL_MINIMUM_EVALUATION_PRODUIT:
            if part_dominant >= SEUIL_CONCENTRATION_NOTABLE:
                concentration_info = {"produit_dominant": dominant, "part": part_dominant, "n": total_avec_produit}

    # ---- Familles de preuve (A-F), indépendantes -- voir NOMS_FAMILLES_PREUVE ----
    a_actif = (
        n >= SEUIL_VOLUME_ABSOLU_NOTABLE
        or (part_candidat is not None and part_candidat >= SEUIL_VOLUME_PART_NOTABLE)
        or temporel_compte
    )
    b_actif = lecture_csat == VOCABULAIRE_ECART_MARQUE
    c_actif = (
        lecture_replies == VOCABULAIRE_ECART_MARQUE
        or lecture_resolution == VOCABULAIRE_ECART_MARQUE
        or lecture_reopens == VOCABULAIRE_ECART_MARQUE
    )
    d_actif = lecture_cout == VOCABULAIRE_ECART_MARQUE
    e_actif = (
        temporel_compte
        and part_elevees >= SEUIL_PART_ELEVEES_PERSISTANCE
        and nb_observations >= NB_OBSERVATIONS_MIN_PERSISTANCE
    )
    f_actif = concentration_info is not None

    familles_actives = []
    if a_actif:
        familles_actives.append("A")
    if b_actif:
        familles_actives.append("B")
    if c_actif:
        familles_actives.append("C")
    if d_actif:
        familles_actives.append("D")
    if e_actif:
        familles_actives.append("E")
    if f_actif:
        familles_actives.append("F")

    # Famille "de conséquence client/opérationnelle" : seule Expérience ou Effort en fait foi.
    # Demande/volume (A) et Impact financier (D) restent des familles de CONTEXTE -- utiles pour
    # étoffer une priorité déjà fondée sur B ou C, mais jamais suffisantes à eux deux pour en
    # créer une (voir docstring de la fonction).
    famille_consequence_active = b_actif or c_actif
    nb_familles_actives = len(familles_actives)

    tier = None
    regle_eligibilite = None
    if nb_familles_actives >= 2 and famille_consequence_active:
        tier = "priorite"
        if nb_familles_actives >= 3:
            regle_eligibilite = "convergence d'au moins 3 familles de preuve, dont expérience client ou effort de traitement"
        else:
            regle_eligibilite = "convergence de 2 familles de preuve incluant expérience client ou effort de traitement"
    elif a_actif and e_actif and f_actif:
        tier = "a_surveiller"
        regle_eligibilite = "combinaison volume + récurrence/persistance + concentration, sans preuve d'impact client ou coût direct"
    elif nb_familles_actives >= 2:
        tier = "a_surveiller"
        regle_eligibilite = "convergence de plusieurs familles de preuve, sans expérience client ni effort de traitement démontré"

    eligible = tier is not None

    if tier == "priorite":
        if nb_familles_actives >= 3:
            niveau_priorite = "Priorité principale"
        else:
            niveau_priorite = "Priorité secondaire"
    elif tier == "a_surveiller":
        niveau_priorite = "À surveiller"
    else:
        niveau_priorite = None

    elements_contributifs = []
    if a_actif:
        elements_contributifs.append("volume/demande au-dessus du niveau habituel")
    if b_actif:
        elements_contributifs.append("satisfaction nettement sous la référence SAV produit")
    if c_actif:
        elements_contributifs.append("effort de traitement nettement au-dessus de la référence")
    if d_actif:
        elements_contributifs.append("coût moyen par dossier supérieur à la référence SAV produit observée")
    if e_actif:
        elements_contributifs.append("récurrence observée sur plusieurs périodes")
    if f_actif:
        elements_contributifs.append("concentré sur un même produit")

    niveau_historique_moyen_pct = None
    moyenne_hist = _moyenne_liste(niveaux_historiques)
    if moyenne_hist is not None:
        niveau_historique_moyen_pct = moyenne_hist * 100

    part_univers_pct = None
    if part_candidat is not None:
        part_univers_pct = part_candidat * 100

    ecart_pct = None
    if ecart_temporel is not None:
        ecart_pct = ecart_temporel * 100

    produit_cle = None
    issue_cle = None
    composant_reel = None
    if grain == GRAIN_PRODUIT_COMPOSANT:
        produit_cle = cle[0]
        composant_reel = cle[1]  # exact -- fait partie de la clé du candidat, jamais deviné
    elif grain == GRAIN_PRODUIT_ISSUE:
        produit_cle = cle[0]
        issue_cle = cle[1]
        # issue_type ne garantit pas structurellement un component unique -- valeur observée
        # (dominante) sur les tickets du candidat, jamais une table codée en dur.
        composant_reel, _part_cr, _total_cr = calculer_concentration(tickets_candidat, "component")

    return {
        "sujet": texte_sujet_candidat(cle, grain),
        "grain": grain,
        "eligible": eligible,
        "tier": tier,
        "regle_eligibilite": regle_eligibilite,
        "niveau_priorite": niveau_priorite,
        "familles_actives": familles_actives,
        "observation_principale": _phrase_elements(elements_contributifs),
        "elements_contributifs": elements_contributifs,
        "elements_consolides": [],
        "regroupement_produit": None,  # renseigné après coup par consolider_signaux_voie_a si pertinent
        "volume": {"n": n, "part_univers_pct": part_univers_pct, "univers": total_univers},
        "reference": {
            "niveau_historique_moyen_pct": niveau_historique_moyen_pct,
            "nb_observations": nb_observations,
            "ecart_pct": ecart_pct,
            "confiance_historique": confiance,
        },
        "experience": {
            "csat": csat_candidat, "n_csat": n_csat_candidat,
            "csat_reference": csat_reference, "n_csat_reference": n_csat_reference,
            "lecture": lecture_csat,
        },
        "effort": {
            "replies_moyen": replies_candidat, "replies_reference": replies_reference,
            "lecture_replies": lecture_replies,
            "resolution_h_moyenne": resolution_candidat, "resolution_h_reference": resolution_reference,
            "lecture_resolution": lecture_resolution,
            "reopens_moyen": reopens_candidat, "reopens_reference": reopens_reference,
            "lecture_reopens": lecture_reopens,
        },
        "cout": cout_info,
        "temporalite": texte_temporalite,
        "concentration": concentration_info,
        "prudence": "Association observée sur les données disponibles, pas une cause démontrée.",
        "score_interne": None,  # usage interne uniquement (classement) -- jamais affiché en UI
        "_ecart_temporel": ecart_temporel,
        "_ecart_csat": ecart_csat,
        "_ecart_effort_max": ecart_effort_max,
        "_confiance_historique": confiance,
        "_produit": produit_cle,
        "_issue_type": issue_cle,
        "_composant": composant_reel,
    }


def calculer_score_interne(signal, cout_maximum):
    composante_temporalite = 0.0
    if signal["_ecart_temporel"] is not None and signal["_ecart_temporel"] > 0:
        composante_temporalite = (
            _normaliser(signal["_ecart_temporel"], PLAFOND_SCORE_TEMPORALITE) * signal["_confiance_historique"]
        )

    composante_persistance = 0.0
    if "E" in signal["familles_actives"]:
        composante_persistance = 1.0

    composante_csat = 0.0
    if signal["_ecart_csat"] is not None and signal["_ecart_csat"] > 0:
        composante_csat = _normaliser(signal["_ecart_csat"], PLAFOND_SCORE_CSAT)

    composante_effort = _normaliser(signal["_ecart_effort_max"], PLAFOND_SCORE_EFFORT)

    composante_cout = 0.0
    if cout_maximum > 0 and signal["cout"] is not None:
        composante_cout = signal["cout"]["montant"] / cout_maximum

    composante_volume_relatif = 0.0
    if signal["volume"]["part_univers_pct"] is not None:
        composante_volume_relatif = _normaliser(
            signal["volume"]["part_univers_pct"] / 100, PLAFOND_SCORE_VOLUME_RELATIF
        )

    composante_volume_absolu = _normaliser(signal["volume"]["n"], PLAFOND_SCORE_VOLUME_ABSOLU)

    composante_concentration = 0.0
    if signal["concentration"] is not None:
        composante_concentration = 1.0

    return (
        composante_temporalite * POIDS_SCORE_TEMPORALITE
        + composante_persistance * POIDS_SCORE_PERSISTANCE
        + composante_csat * POIDS_SCORE_CSAT
        + composante_effort * POIDS_SCORE_EFFORT
        + composante_cout * POIDS_SCORE_COUT
        + composante_volume_relatif * POIDS_SCORE_VOLUME_RELATIF
        + composante_volume_absolu * POIDS_SCORE_VOLUME_ABSOLU
        + composante_concentration * POIDS_SCORE_CONCENTRATION
    )


def obtenir_score_interne(signal):
    return signal["score_interne"]


def _evaluer_grain(cle, grain, tickets_candidat, tickets_sav_produit_periode,
                    historique_sav_produit_par_fichier, commandes, couts_produits, reference_cout_moyen):
    niveaux_historiques = construire_niveaux_historiques(historique_sav_produit_par_fichier, cle, grain)
    return construire_signal_produit_voie_a(
        cle, grain, tickets_candidat, tickets_sav_produit_periode, niveaux_historiques,
        commandes, couts_produits, reference_cout_moyen,
    )


def _cle_identite(signal):
    return (signal["_produit"], signal["_composant"])


# Repère, après consolidation, les cas où un même produit porte plusieurs histoires distinctes
# (des composants différents -- jamais fusionnées entre elles, voir consolider_signaux_voie_a) --
# un regroupement purement VISUEL pour la lecture manager, jamais une fusion de données : chaque
# signal garde son propre volume/preuve/niveau. "Clarté — Charge" et "Clarté — Capsule" peuvent
# ainsi apparaître comme deux cartes liées plutôt que deux cartes dispersées sans rapport visible.
def _marquer_regroupements_produit(signaux):
    sujets_par_produit = {}
    for signal in signaux:
        produit = signal["_produit"]
        if produit is None:
            continue
        if produit in sujets_par_produit:
            sujets_par_produit[produit].append(signal)
        else:
            sujets_par_produit[produit] = [signal]

    for produit, groupe in sujets_par_produit.items():
        if len(groupe) < 2:
            continue
        for signal in groupe:
            autres_sujets = []
            for autre in groupe:
                if autre is not signal:
                    autres_sujets.append(autre["sujet"])
            signal["regroupement_produit"] = {"produit": produit, "autres_sujets": autres_sujets}


# Consolidation par IDENTITÉ STRUCTURELLE réelle, jamais par preuve partagée ni par comparaison de
# chaînes de caractères. Principe verrouillé (Étape 4A.3) : les familles de preuve (B/C/D/E/F)
# disent SEULEMENT "ce sujet mérite de l'attention" -- elles ne servent JAMAIS à décider si deux
# sujets sont le même problème. Seule la structure des tickets (product_name, component réel de
# chaque ticket) définit l'identité.
#
# Le grain produit x composant regroupe déjà, par construction, tous les issue_types d'un même
# produit qui partagent le même component réel -- absorber ses enfants produit x issue est donc
# toujours structurellement sûr (même produit, même composant = même famille fonctionnelle), sans
# avoir besoin de vérifier quoi que ce soit sur leurs preuves. À l'inverse, deux issue_types du
# même produit mais de components différents ne se rencontrent JAMAIS dans le même candidat
# produit x composant -- ils ne peuvent donc pas être fusionnés à tort, même s'ils partagent un
# CSAT bas, un effort élevé ou une persistance : ce ne sont pas des preuves d'identité.
#
# Le composant reste le signal principal quand il est diffus sur plusieurs produits (aucun ne
# domine) ; le produit x composant devient le signal principal quand un seul produit concentre le
# problème. Si plusieurs produit x composant distincts et éligibles subsistent pour le même
# produit (plusieurs vrais problèmes indépendants), ils restent des cartes séparées, simplement
# regroupées visuellement (_marquer_regroupements_produit) -- jamais fusionnés en une seule.
def consolider_signaux_voie_a(signaux_composant, signaux_produit_composant, signaux_produit_issue):
    signaux_issue_par_identite = {}
    for signal in signaux_produit_issue:
        cle = _cle_identite(signal)
        if cle in signaux_issue_par_identite:
            signaux_issue_par_identite[cle].append(signal)
        else:
            signaux_issue_par_identite[cle] = [signal]

    signaux_produit_composant_par_identite = {}
    for signal_pc in signaux_produit_composant:
        signaux_produit_composant_par_identite[_cle_identite(signal_pc)] = signal_pc

    resultat = []
    identites_traitees = set()

    for signal_comp in signaux_composant:
        if not signal_comp["eligible"]:
            continue

        composant = signal_comp["sujet"]
        absorbe_par_produit = False

        if signal_comp["concentration"] is not None:
            produit_dominant = signal_comp["concentration"]["produit_dominant"]
            cle_dominante = (produit_dominant, composant)
            signal_pc_dominant = signaux_produit_composant_par_identite.get(cle_dominante)
            if signal_pc_dominant is not None and signal_pc_dominant["eligible"]:
                for signal_i in signaux_issue_par_identite.get(cle_dominante, []):
                    signal_pc_dominant["elements_consolides"].append(
                        str(signal_i["_issue_type"]) + " (" + str(signal_i["volume"]["n"]) + " tickets)"
                    )
                signal_pc_dominant["elements_consolides"].append(
                    "contexte composant : " + str(signal_comp["volume"]["n"]) + " tickets au total sur " + composant
                )
                resultat.append(signal_pc_dominant)
                identites_traitees.add(cle_dominante)
                absorbe_par_produit = True

        if not absorbe_par_produit:
            for cle, signal_pc in signaux_produit_composant_par_identite.items():
                if cle[1] == composant and signal_pc["eligible"] and cle not in identites_traitees:
                    signal_comp["elements_consolides"].append(
                        cle[0] + " (" + str(signal_pc["volume"]["n"]) + " tickets)"
                    )
                    identites_traitees.add(cle)
            resultat.append(signal_comp)

    for cle, signal_pc in signaux_produit_composant_par_identite.items():
        if cle in identites_traitees or not signal_pc["eligible"]:
            continue
        for signal_i in signaux_issue_par_identite.get(cle, []):
            signal_pc["elements_consolides"].append(
                str(signal_i["_issue_type"]) + " (" + str(signal_i["volume"]["n"]) + " tickets)"
            )
        resultat.append(signal_pc)
        identites_traitees.add(cle)

    for cle, groupe_issue in signaux_issue_par_identite.items():
        if cle in identites_traitees:
            continue
        for signal_i in groupe_issue:
            if signal_i["eligible"]:
                resultat.append(signal_i)

    _marquer_regroupements_produit(resultat)

    return resultat


# Orchestration Voie A : génère les trois grains de candidats, évalue chacun contre une référence
# de coût propre à la période (jamais un seuil absolu codé en dur), consolide les doublons par
# chevauchement réel de preuve, sépare le résultat en deux collections (priorités / à surveiller)
# selon le "tier" gagné par chaque signal consolidé, puis plafonne CHAQUE collection au nombre
# maximum demandé -- une pure limite d'affichage, jamais un filtre métier : le nombre de
# priorités peut légitimement être 0, 1, 2... comme 5.
def moteur_produit_voie_a(tickets_sav_produit_periode, historique_sav_produit_par_fichier,
                           commandes, couts_produits, nombre_max_signaux):
    reference_cout_moyen = calculer_reference_cout_moyen(tickets_sav_produit_periode, commandes, couts_produits)

    signaux_composant = []
    for cle, grain, tickets_candidat in generer_candidats_composant(tickets_sav_produit_periode):
        signal = _evaluer_grain(
            cle, grain, tickets_candidat, tickets_sav_produit_periode,
            historique_sav_produit_par_fichier, commandes, couts_produits, reference_cout_moyen,
        )
        if signal is not None:
            signaux_composant.append(signal)

    signaux_produit_composant = []
    for cle, grain, tickets_candidat in generer_candidats_produit_composant(tickets_sav_produit_periode):
        signal = _evaluer_grain(
            cle, grain, tickets_candidat, tickets_sav_produit_periode,
            historique_sav_produit_par_fichier, commandes, couts_produits, reference_cout_moyen,
        )
        if signal is not None:
            signaux_produit_composant.append(signal)

    signaux_produit_issue = []
    for cle, grain, tickets_candidat in generer_candidats_produit_issue(tickets_sav_produit_periode):
        signal = _evaluer_grain(
            cle, grain, tickets_candidat, tickets_sav_produit_periode,
            historique_sav_produit_par_fichier, commandes, couts_produits, reference_cout_moyen,
        )
        if signal is not None:
            signaux_produit_issue.append(signal)

    signaux_consolides = consolider_signaux_voie_a(signaux_composant, signaux_produit_composant, signaux_produit_issue)

    cout_maximum = 0.0
    for signal in signaux_consolides:
        if signal["cout"] is not None and signal["cout"]["montant"] > cout_maximum:
            cout_maximum = signal["cout"]["montant"]

    for signal in signaux_consolides:
        signal["score_interne"] = calculer_score_interne(signal, cout_maximum)

    signaux_prioritaires_bruts = []
    signaux_a_surveiller_bruts = []
    for signal in signaux_consolides:
        if signal["tier"] == "priorite":
            signaux_prioritaires_bruts.append(signal)
        elif signal["tier"] == "a_surveiller":
            signaux_a_surveiller_bruts.append(signal)

    signaux_prioritaires_tries = sorted(signaux_prioritaires_bruts, key=obtenir_score_interne, reverse=True)
    signaux_a_surveiller_tries = sorted(signaux_a_surveiller_bruts, key=obtenir_score_interne, reverse=True)

    prioritaires = []
    for i in range(min(nombre_max_signaux, len(signaux_prioritaires_tries))):
        prioritaires.append(signaux_prioritaires_tries[i])

    a_surveiller = []
    for i in range(min(nombre_max_signaux, len(signaux_a_surveiller_tries))):
        a_surveiller.append(signaux_a_surveiller_tries[i])

    return {
        "prioritaires": prioritaires,
        "a_surveiller": a_surveiller,
        # Histoires métier réellement distinctes (après consolidation, avant plafond d'affichage)
        # -- permet de vérifier que le plafond ne masque pas une histoire, sans jamais recalibrer
        # les critères pour la faire "tenir" dans le nombre affiché.
        "nb_prioritaires_avant_plafond": len(signaux_prioritaires_tries),
        "nb_a_surveiller_avant_plafond": len(signaux_a_surveiller_tries),
    }


# Voie B : jamais un seul champ (ex. remboursement + CSAT faible seuls) ne déclenche un signal
# -- il faut en plus au moins un facteur de gravité parmi réouvertures / échanges / délai de
# résolution, chacun comparé à la référence de la période plutôt qu'à un seuil absolu arbitraire.
def evaluer_gravite_ticket_voie_b(ticket, reference_replies, reference_resolution_h):
    resolution = ticket["resolution_type"]
    csat = ticket["csat"]

    if resolution not in RESOLUTIONS_GRAVES_VOIE_B:
        return None
    if csat is None or csat > SEUIL_VOIE_B_CSAT_MAX:
        return None

    reopens = ticket["reopens"]
    replies = ticket["replies"]
    resolution_h = ticket["full_resolution_time_hours"]

    raisons = []
    if reopens is not None and reopens >= SEUIL_VOIE_B_REOPENS_MIN:
        raisons.append(str(reopens) + " réouvertures")
    if replies is not None and reference_replies is not None and reference_replies > 0:
        if replies >= reference_replies * SEUIL_VOIE_B_REPLIES_MULTIPLICATEUR:
            raisons.append("nombre d'échanges nettement au-dessus de la référence de la période")
    if resolution_h is not None and reference_resolution_h is not None and reference_resolution_h > 0:
        if resolution_h >= reference_resolution_h * SEUIL_VOIE_B_RESOLUTION_MULTIPLICATEUR:
            raisons.append("délai de résolution nettement au-dessus de la référence de la période")

    if len(raisons) == 0:
        return None

    texte_raison = raisons[0]
    for i in range(1, len(raisons)):
        texte_raison = texte_raison + " ; " + raisons[i]

    return {
        "ticket_id": ticket["ticket_id"],
        "sujet": ticket["subject_cluster"],
        "produit": ticket["product_name"],
        "composant": ticket["component"],
        "resolution_type": resolution,
        "csat": csat,
        "reopens": reopens,
        "raison": texte_raison,
        "avertissement": "Heuristique de gravité sur un dossier individuel, pas une détection de risque sécurité certifiée.",
    }


def moteur_produit_voie_b(tickets_sav_produit_periode):
    reference_replies = moyenne(tickets_sav_produit_periode, "replies")
    reference_resolution_h = moyenne(tickets_sav_produit_periode, "full_resolution_time_hours")

    signaux = []
    for ticket in tickets_sav_produit_periode:
        signal = evaluer_gravite_ticket_voie_b(ticket, reference_replies, reference_resolution_h)
        if signal is not None:
            signaux.append(signal)
    return signaux


# ---------------------------------------------------------------------------
# Moteur Tendances / synthèse temporelle (Étape 4B)
#
# Chaque export est une observation isolée, jamais un point d'une série continue : les exports
# disponibles sont espacés de façon irrégulière (parfois 1 semaine, parfois 6-7) sur une année
# complète. Aucune fonction ci-dessous ne calcule de pente hebdomadaire ni ne suppose une
# continuité entre deux exports voisins dans la liste -- les comparaisons se font contre
# l'ensemble des observations disponibles (ou un sous-ensemble comparable), et la comparaison à
# l'observation immédiatement précédente est toujours présentée comme telle (jamais comme une
# évolution semaine par semaine).
# ---------------------------------------------------------------------------

SEUIL_ECART_VOLUME_NOTABLE = 0.30
# Points d'écart CSAT -- volontairement plus strict que le seuil Produit (0.4) : une moyenne
# hebdomadaire agrégée varie naturellement beaucoup moins qu'un sous-ensemble de tickets, un même
# seuil absolu manquerait presque tous les écarts réels à cette échelle.
SEUIL_CSAT_ECART_MARQUE_TENDANCES = 0.20
SEUIL_DIFFICULTE_ECART_NOTABLE = 0.25
SEUIL_CAPACITE_COMPARABLE_PCT = 0.10
# L'observation "plus difficile" doit avoir un volume au moins 15 % plus bas pour que le contraste
# volume/difficulté ait un sens à signaler (sinon ce n'est qu'un pic de volume classique).
SEUIL_VOLUME_INFERIEUR_POUR_CONTRASTE = 0.85
SEUIL_PART_CATEGORIE_ECART_NOTABLE = 0.25
NB_OBSERVATIONS_MIN_SAISON = 2
NB_SAISONS_MIN_PROFIL = 2

SAISON_PAR_MOIS = {
    9: "automne", 10: "automne", 11: "automne",
    12: "hiver", 1: "hiver", 2: "hiver",
    3: "printemps", 4: "printemps", 5: "printemps",
    6: "été", 7: "été", 8: "été",
}

NOMS_MOIS = {
    1: "janvier", 2: "février", 3: "mars", 4: "avril", 5: "mai", 6: "juin",
    7: "juillet", 8: "août", 9: "septembre", 10: "octobre", 11: "novembre", 12: "décembre",
}

# Position relative dans la distribution (rang), pas un seuil d'écart : robuste aux valeurs
# extrêmes qui déformeraient une moyenne (un ou deux pics de fin d'année ne doivent pas rendre
# "normale" une observation qui reste pourtant parmi les plus hautes de l'année -- voir Étape
# 4B.1). Les seuils "avec_contexte" sont volontairement moins stricts que les seuils "strict" :
# une position seulement modérément écartée peut mériter d'être racontée si un événement métier
# réel lui donne un sens, mais jamais sans ce lien.
SEUIL_RANG_HAUT_STRICT = 0.75
SEUIL_RANG_HAUT_AVEC_CONTEXTE = 0.60
SEUIL_RANG_BAS_STRICT = 0.25
SEUIL_RANG_BAS_AVEC_CONTEXTE = 0.40
SEUIL_RANG_EXCEPTIONNEL_HAUT = 0.90
SEUIL_RANG_EXCEPTIONNEL_BAS = 0.10

# Trois modes de lecture pour l'onglet Tendances (Étape 4B.3), déduits uniquement du nombre
# d'observations dans la période sélectionnée par l'utilisatrice vs le nombre total d'observations
# disponibles jusqu'à cette date -- jamais d'un mois ou d'une date codée en dur.
MODE_OBSERVATION_UNIQUE = "observation_unique"
MODE_PERIODE_ETENDUE = "periode_etendue"
MODE_HISTORIQUE_COMPLET = "historique_complet"


def determiner_mode_tendances(nb_observations_periode, nb_observations_historique):
    if nb_observations_periode <= 1:
        return MODE_OBSERVATION_UNIQUE
    if nb_observations_periode >= nb_observations_historique:
        return MODE_HISTORIQUE_COMPLET
    return MODE_PERIODE_ETENDUE


def capacite_totale_heures(planning):
    total = 0
    for agent, jours in planning.items():
        if agent == NOM_AGENT_DEFAUT:
            continue
        for jour, plages in jours.items():
            for debut, fin in plages:
                total = total + (fin - debut)
    return total


# Une observation = tout ce que le moteur sait d'un export (ou d'une période) pris isolément :
# volume, mix des demandes, expérience, effort Care, capacité prévue, contexte métier propre à
# CETTE période uniquement -- contexte_periode() est bornée à [date_debut, date_fin] de cette seule
# observation, jamais étendue pour aller chercher un événement futur ou passé qui ne la chevauche
# pas (ex. la fin d'alternance de Sofia le 30/09 ne doit jamais apparaître dans le contexte d'un
# export du 7-13/09).
def construire_profil_observation(tickets, planning, evenements, date_debut, date_fin):
    plannings_periode = [(date_debut, date_fin, planning)]

    mix_categories = {}
    for categorie, tickets_categorie in grouper_par_categorie(tickets).items():
        mix_categories[categorie] = len(tickets_categorie)

    return {
        "date_debut": date_debut,
        "date_fin": date_fin,
        "volume": len(tickets),
        "mix_categories": mix_categories,
        "csat": moyenne(tickets, "csat"),
        "n_csat": _compte_valeurs_non_nulles(tickets, "csat"),
        "frt": moyenne(tickets, "first_reply_time_min"),
        "taux_sla": taux_sla(tickets, plannings_periode),
        "resolution_h": moyenne(tickets, "full_resolution_time_hours"),
        "reopens": moyenne(tickets, "reopens"),
        "replies": moyenne(tickets, "replies"),
        "macro_pct": taux_rempli(tickets, "macro_applied"),
        "capacite_heures": capacite_totale_heures(planning),
        "contexte": contexte_periode(evenements, date_debut, date_fin),
    }


# Écart relatif d'une observation par rapport à la moyenne de TOUTES LES AUTRES observations
# disponibles (leave-one-out) -- jamais contre une moyenne mobile ou un point voisin dans le temps.
def ecart_relatif_vs_reste(valeurs, index):
    valeur_cible = valeurs[index]
    reste = []
    for i in range(len(valeurs)):
        if i != index and valeurs[i] is not None:
            reste.append(valeurs[i])
    if valeur_cible is None or len(reste) == 0:
        return None
    moyenne_reste = sum(reste) / len(reste)
    if moyenne_reste == 0:
        return None
    return (valeur_cible - moyenne_reste) / moyenne_reste


def moyenne_du_reste(valeurs, index):
    reste = []
    for i in range(len(valeurs)):
        if i != index and valeurs[i] is not None:
            reste.append(valeurs[i])
    if len(reste) == 0:
        return None
    return sum(reste) / len(reste)


def mediane(valeurs):
    valeurs_valides = []
    for valeur in valeurs:
        if valeur is not None:
            valeurs_valides.append(valeur)
    if len(valeurs_valides) == 0:
        return None
    valeurs_triees = sorted(valeurs_valides)
    n = len(valeurs_triees)
    milieu = n // 2
    if n % 2 == 1:
        return valeurs_triees[milieu]
    return (valeurs_triees[milieu - 1] + valeurs_triees[milieu]) / 2.0


def mediane_du_reste(valeurs, index):
    reste = []
    for i in range(len(valeurs)):
        if i != index and valeurs[i] is not None:
            reste.append(valeurs[i])
    return mediane(reste)


# Position relative de l'observation DANS LA DISTRIBUTION COMPLÈTE des observations disponibles
# (elle-même incluse) -- 0 = la plus basse, 1 = la plus haute. Une lecture de position (rang),
# jamais une pente ni un pourcentage d'écart à une moyenne fragile : un ou deux pics extrêmes
# (Black Friday, Noël) ne peuvent pas, par construction, rendre "normale" une observation qui
# reste pourtant parmi les plus hautes de l'année (ex. janvier) -- ils ne font que prendre leur
# propre place dans le classement, sans déformer la position relative des autres.
def rang_relatif(valeurs, index):
    valeur_cible = valeurs[index]
    if valeur_cible is None:
        return None
    valeurs_valides = []
    for valeur in valeurs:
        if valeur is not None:
            valeurs_valides.append(valeur)
    if len(valeurs_valides) <= 1:
        return None
    nb_inferieurs = 0
    for valeur in valeurs_valides:
        if valeur < valeur_cible:
            nb_inferieurs = nb_inferieurs + 1
    return nb_inferieurs / (len(valeurs_valides) - 1)


def texte_position_relative(rang):
    if rang is None:
        return None
    if rang >= SEUIL_RANG_EXCEPTIONNEL_HAUT:
        return "figure parmi les niveaux les plus hauts observés"
    if rang >= SEUIL_RANG_HAUT_AVEC_CONTEXTE:
        return "reste élevé à l'échelle des observations disponibles"
    if rang <= SEUIL_RANG_EXCEPTIONNEL_BAS:
        return "figure parmi les niveaux les plus bas observés"
    if rang <= SEUIL_RANG_BAS_AVEC_CONTEXTE:
        return "reste bas à l'échelle des observations disponibles"
    return "se situe dans une zone médiane des observations disponibles"


# Compose une phrase française unique à partir de la position dans la distribution (déjà
# déterminée par rang_relatif/texte_position_relative, non recalculée ici) et de la tendance vs la
# dernière observation disponible -- jamais deux fragments juxtaposés mécaniquement ("Volume X ;
# Y."). Le connecteur ("et"/"mais") est choisi selon que la tendance va dans le même sens que la
# position ou la contredit -- purement rédactionnel, ne décide d'aucun seuil ni classification.
def texte_volume_naturel(rang, direction_precedente):
    texte_position = texte_position_relative(rang)
    if texte_position is None and direction_precedente is None:
        return None
    if direction_precedente is None:
        return "Le volume " + texte_position + "."
    if direction_precedente == "hausse":
        clause_tendance = "progresse par rapport à la dernière observation disponible"
        contraste = rang is not None and rang <= SEUIL_RANG_BAS_AVEC_CONTEXTE
    else:
        clause_tendance = "recule par rapport à la dernière observation disponible"
        contraste = rang is not None and rang >= SEUIL_RANG_HAUT_AVEC_CONTEXTE
    if texte_position is None:
        return "Le volume " + clause_tendance + "."
    connecteur = "mais"
    if not contraste:
        connecteur = "et"
    return "Le volume " + clause_tendance + ", " + connecteur + " " + texte_position + "."


# Nom de catégorie tel qu'affiché dans un texte exécutif -- retire le suffixe technique entre
# parenthèses (ex. "SAV produit (défaut)" -> "SAV produit") sans toucher à la donnée sous-jacente,
# utilisée telle quelle partout ailleurs (regroupement, comptage, filtres).
def simplifier_nom_categorie(categorie):
    if categorie is None:
        return None
    index_parenthese = categorie.find(" (")
    if index_parenthese == -1:
        return categorie
    return categorie[:index_parenthese]


def nom_mois(numero_mois):
    return NOMS_MOIS[numero_mois]


def lire_ecart_csat_tendances(valeur, reference):
    if valeur is None or reference is None:
        return VOCABULAIRE_ECART_INDISPONIBLE
    ecart = reference - valeur  # positif si la période est moins bien notée que le reste
    if abs(ecart) < 0.02:
        return VOCABULAIRE_ECART_PROCHE
    if ecart >= SEUIL_CSAT_ECART_MARQUE_TENDANCES:
        return VOCABULAIRE_ECART_MARQUE
    if ecart > 0:
        return VOCABULAIRE_ECART_BAS
    return VOCABULAIRE_ECART_PROCHE


def texte_contexte(evenements_contexte):
    if evenements_contexte is None or len(evenements_contexte) == 0:
        return None
    noms = []
    for evenement in evenements_contexte:
        noms.append(evenement["nom_evenement"])
    return "Chevauche : " + " · ".join(noms)


# Chaque insight suit la même forme, quel que soit son type (fait marquant, contraste,
# saisonnalité, comparaison à l'observation précédente) : une observation factuelle, pourquoi elle
# mérite l'attention, un contexte juxtaposé le cas échéant (jamais présenté comme une cause), et
# une prudence explicite quand elle s'impose.
def construire_insight(type_insight, observation, pourquoi, contexte=None, prudence=None):
    return {
        "type": type_insight,
        "observation": observation,
        "pourquoi": pourquoi,
        "contexte": contexte,
        "prudence": prudence,
    }


# Détecte, pour une observation donnée, si une catégorie de demande concentre une part nettement
# plus élevée que sur les autres observations disponibles -- generique sur toutes les catégories
# réellement présentes dans les données, jamais une catégorie testée en dur.
def categorie_ecart_notable(profils, index):
    toutes_categories = set()
    for profil in profils:
        for categorie in profil["mix_categories"]:
            toutes_categories.add(categorie)

    meilleure_categorie = None
    meilleur_ecart = 0.0

    for categorie in toutes_categories:
        parts = []
        for profil in profils:
            if profil["volume"] == 0:
                parts.append(None)
                continue
            n_categorie = profil["mix_categories"].get(categorie, 0)
            parts.append(n_categorie / profil["volume"])

        ecart = ecart_relatif_vs_reste(parts, index)
        if ecart is not None and ecart >= SEUIL_PART_CATEGORIE_ECART_NOTABLE and ecart > meilleur_ecart:
            meilleur_ecart = ecart
            meilleure_categorie = categorie

    if meilleure_categorie is None:
        return None
    return meilleure_categorie, meilleur_ecart


# Indice de difficulté opérationnelle relatif (jamais affiché tel quel -- usage interne uniquement
# pour classer/contraster des observations) : moyenne des écarts relatifs (vs le reste des
# observations disponibles) sur résolution, réouvertures et échanges (plus haut = plus dur), et
# l'écart CSAT inversé (plus bas que le reste = plus dur). Générique -- ne connaît aucune date ni
# aucun nom de mois.
def calculer_indice_difficulte(profils, index):
    resolutions = []
    reopens_liste = []
    replies_liste = []
    csats = []
    for profil in profils:
        resolutions.append(profil["resolution_h"])
        reopens_liste.append(profil["reopens"])
        replies_liste.append(profil["replies"])
        csats.append(profil["csat"])

    composantes = []

    ecart_resolution = ecart_relatif_vs_reste(resolutions, index)
    if ecart_resolution is not None:
        composantes.append(ecart_resolution)

    ecart_reopens = ecart_relatif_vs_reste(reopens_liste, index)
    if ecart_reopens is not None:
        composantes.append(ecart_reopens)

    ecart_replies = ecart_relatif_vs_reste(replies_liste, index)
    if ecart_replies is not None:
        composantes.append(ecart_replies)

    ecart_csat = ecart_relatif_vs_reste(csats, index)
    if ecart_csat is not None:
        composantes.append(-ecart_csat)  # csat plus bas que le reste = plus difficile

    if len(composantes) == 0:
        return None
    return sum(composantes) / len(composantes)


# Regroupe les observations dont la capacité prévue est proche (tolérance relative), pour comparer
# des périodes réellement comparables en charge de travail plutôt que deux instantanés
# arbitraires. Regroupement glissant simple (chaque membre comparé à la référence du groupe, pas
# au précédent) -- suffisant pour la douzaine d'observations disponibles.
def grouper_par_capacite_comparable(profils, tolerance_pct):
    paires = []
    for i in range(len(profils)):
        capacite = profils[i]["capacite_heures"]
        if capacite is not None and capacite > 0:
            paires.append((capacite, i))
    paires_triees = sorted(paires)

    groupes = []
    groupe_courant = []
    capacite_reference_groupe = None

    for capacite, indice in paires_triees:
        if len(groupe_courant) > 0 and abs(capacite - capacite_reference_groupe) / capacite_reference_groupe <= tolerance_pct:
            groupe_courant.append(indice)
        else:
            if len(groupe_courant) >= 2:
                groupes.append(groupe_courant)
            groupe_courant = [indice]
            capacite_reference_groupe = capacite

    if len(groupe_courant) >= 2:
        groupes.append(groupe_courant)

    return groupes


# "Décembre et janvier ont une capacité comparable" (exemple type, jamais codé en dur) : au sein
# de chaque groupe à capacité comparable, compare le membre le plus facile et le plus difficile
# (indice de difficulté) -- ne remonte que si l'écart est réellement notable.
def detecter_contrastes_capacite(profils, tolerance_pct=SEUIL_CAPACITE_COMPARABLE_PCT):
    groupes = grouper_par_capacite_comparable(profils, tolerance_pct)
    contrastes = []

    for groupe in groupes:
        indice_facile = None
        indice_difficile = None
        difficulte_min = None
        difficulte_max = None

        for indice in groupe:
            difficulte = calculer_indice_difficulte(profils, indice)
            if difficulte is None:
                continue
            if difficulte_min is None or difficulte < difficulte_min:
                difficulte_min = difficulte
                indice_facile = indice
            if difficulte_max is None or difficulte > difficulte_max:
                difficulte_max = difficulte
                indice_difficile = indice

        if indice_facile is None or indice_difficile is None or indice_facile == indice_difficile:
            continue
        if difficulte_max - difficulte_min < SEUIL_DIFFICULTE_ECART_NOTABLE:
            continue

        profil_facile = profils[indice_facile]
        profil_difficile = profils[indice_difficile]

        texte_volume = ""
        if profil_difficile["volume"] is not None and profil_facile["volume"] is not None:
            if profil_difficile["volume"] <= profil_facile["volume"] * SEUIL_VOLUME_INFERIEUR_POUR_CONTRASTE:
                texte_volume = " malgré un volume inférieur"

        observation = (
            "Deux observations à capacité prévue comparable (" + str(round(profil_facile["capacite_heures"]))
            + "h vs " + str(round(profil_difficile["capacite_heures"])) + "h) montrent un niveau de tension "
            "opérationnelle très différent" + texte_volume + " : la période du " + str(profil_difficile["date_debut"])
            + " ressort comme nettement plus tendue que celle du " + str(profil_facile["date_debut"]) + "."
        )
        contraste_insight = construire_insight(
            "contraste", observation,
            "Une capacité prévue comparable ne garantit pas une charge réelle comparable -- la complexité des "
            "demandes traitées compte aussi.",
            prudence="Association observée sur les données disponibles, pas une cause démontrée.",
        )
        # Dates des deux observations impliquées -- métadonnée additive, utilisée par le scope de
        # période (Étape 4B.3) pour ne retenir que les contrastes pertinents pour la période
        # analysée. N'affecte ni le texte ni la détection elle-même.
        contraste_insight["date_debut_a"] = profil_facile["date_debut"]
        contraste_insight["date_debut_b"] = profil_difficile["date_debut"]
        contrastes.append(contraste_insight)

    return contrastes


# Évaluateur central d'une observation prise isolément. Distingue explicitement deux registres,
# jamais confondus (Étape 4B.1) :
#   - VIGILANCE : expérience et/ou effort réellement dégradés -- prioritaire, l'observation reste
#     "vigilance" même si son volume est par ailleurs élevé (ex. janvier : volume qui reste haut
#     dans le classement, MAIS le sujet à raconter est la dégradation, pas le volume).
#   - JALON MÉTIER : aucune dégradation, mais l'observation mérite d'être racontée -- position
#     extrême dans le classement des observations disponibles (rang_relatif, jamais une moyenne
#     fragile), éventuellement épaulée par un contexte métier réel qui abaisse le seuil requis.
# Le contexte seul, sans aucune manifestation dans les métriques, ne déclenche jamais un jalon --
# il ne fait qu'abaisser le seuil de position déjà nécessaire (voir SEUIL_RANG_*_AVEC_CONTEXTE).
# Retourne None pour une période sans rien à raconter -- une période calme est une information,
# pas une absence de résultat à combler.
def detecter_fait_marquant(profils, index):
    profil = profils[index]

    volumes = []
    for p in profils:
        volumes.append(p["volume"])
    rang_volume = rang_relatif(volumes, index)
    texte_position = texte_position_relative(rang_volume)

    direction_precedente = None
    if index > 0:
        volume_precedent = profils[index - 1]["volume"]
        if volume_precedent is not None and volume_precedent > 0 and profil["volume"] is not None:
            ecart_precedent = (profil["volume"] - volume_precedent) / volume_precedent
            if ecart_precedent >= SEUIL_ECART_VOLUME_NOTABLE:
                direction_precedente = "hausse"
            elif ecart_precedent <= -SEUIL_ECART_VOLUME_NOTABLE:
                direction_precedente = "baisse"

    texte_volume_complet = texte_volume_naturel(rang_volume, direction_precedente)

    csats = []
    for p in profils:
        csats.append(p["csat"])
    lecture_csat = lire_ecart_csat_tendances(profil["csat"], moyenne_du_reste(csats, index))

    replies_liste = []
    resolutions = []
    reopens_liste = []
    for p in profils:
        replies_liste.append(p["replies"])
        resolutions.append(p["resolution_h"])
        reopens_liste.append(p["reopens"])

    lecture_replies = lire_ecart_effort(profil["replies"], moyenne_du_reste(replies_liste, index))
    lecture_resolution = lire_ecart_effort(profil["resolution_h"], moyenne_du_reste(resolutions, index))
    lecture_reopens = lire_ecart_effort(profil["reopens"], moyenne_du_reste(reopens_liste, index))

    experience_degradee = lecture_csat == VOCABULAIRE_ECART_MARQUE
    effort_degrade = (
        lecture_replies == VOCABULAIRE_ECART_MARQUE
        or lecture_resolution == VOCABULAIRE_ECART_MARQUE
        or lecture_reopens == VOCABULAIRE_ECART_MARQUE
    )

    categorie_info = categorie_ecart_notable(profils, index)
    mix_inhabituel = categorie_info is not None
    contexte_texte = texte_contexte(profil["contexte"])
    a_contexte = contexte_texte is not None

    # ---- VIGILANCE : prioritaire, jamais fondue avec un jalon sur la même observation ----
    if experience_degradee or effort_degrade:
        elements = []
        if experience_degradee:
            elements.append("une satisfaction client en retrait")
        if effort_degrade:
            elements.append("un effort de traitement plus soutenu (échanges, délai de résolution ou réouvertures)")
        if mix_inhabituel:
            categorie, _ecart_categorie = categorie_info
            elements.append("une part plus marquée de " + simplifier_nom_categorie(categorie))

        observation = "Cette période se distingue par " + _joindre_liste_fr(elements) + "."
        if texte_volume_complet is not None:
            observation = observation + " " + texte_volume_complet

        insight = construire_insight(
            "fait_marquant", observation,
            "La dégradation touche l'expérience client et/ou l'effort de traitement -- à regarder de plus près, "
            "indépendamment du niveau de volume.",
            contexte=contexte_texte,
            prudence="Association observée sur les données disponibles, pas une cause démontrée.",
        )
        insight["categorie"] = "vigilance"
        insight["registre"] = "tension opérationnelle"
        insight["date_debut"] = profil["date_debut"]
        insight["date_fin"] = profil["date_fin"]
        return insight

    # ---- JALON MÉTIER : position extrême (seule ou épaulée par un contexte réel) ----
    volume_haut_strict = rang_volume is not None and rang_volume >= SEUIL_RANG_HAUT_STRICT
    volume_haut_avec_contexte = (
        rang_volume is not None and a_contexte
        and SEUIL_RANG_HAUT_AVEC_CONTEXTE <= rang_volume < SEUIL_RANG_HAUT_STRICT
    )
    volume_bas_strict = rang_volume is not None and rang_volume <= SEUIL_RANG_BAS_STRICT
    volume_bas_avec_contexte = (
        rang_volume is not None and a_contexte
        and SEUIL_RANG_BAS_AVEC_CONTEXTE >= rang_volume > SEUIL_RANG_BAS_STRICT
    )

    jalon_volume_haut = volume_haut_strict or volume_haut_avec_contexte
    jalon_volume_bas = volume_bas_strict or volume_bas_avec_contexte
    jalon_mix_seul = mix_inhabituel and a_contexte and not jalon_volume_haut and not jalon_volume_bas

    if not jalon_volume_haut and not jalon_volume_bas and not jalon_mix_seul:
        return None

    if jalon_volume_haut:
        if rang_volume >= SEUIL_RANG_EXCEPTIONNEL_HAUT:
            registre = "pic absorbé"
        else:
            registre = "activité soutenue maîtrisée"
        observation = "Le volume " + texte_position + ", sans dégradation d'expérience ni d'effort associée."
    elif jalon_volume_bas:
        if rang_volume <= SEUIL_RANG_EXCEPTIONNEL_BAS:
            registre = "creux sans tension"
        else:
            registre = "activité calme"
        observation = "Le volume " + texte_position + ", sans tension particulière."
    else:
        registre = "mix de demandes inhabituel"
        categorie, _ecart_categorie = categorie_info
        observation = "Les demandes se concentrent nettement sur " + simplifier_nom_categorie(categorie) + ", sans dégradation associée."

    if mix_inhabituel and (jalon_volume_haut or jalon_volume_bas):
        categorie, _ecart_categorie = categorie_info
        observation = observation + " Porté notamment par " + simplifier_nom_categorie(categorie) + "."

    if direction_precedente is not None:
        if direction_precedente == "hausse":
            observation = observation + " Le volume progresse par rapport à la dernière observation disponible."
        else:
            observation = observation + " Le volume recule par rapport à la dernière observation disponible."

    insight = construire_insight(
        "fait_marquant", observation,
        "Ce moment aide à comprendre l'activité de la période -- pas un problème, un jalon pour la lecture "
        "d'ensemble.",
        contexte=contexte_texte,
        prudence="Association observée sur les données disponibles, pas une cause démontrée.",
    )
    insight["categorie"] = "jalon"
    insight["registre"] = registre
    insight["date_debut"] = profil["date_debut"]
    insight["date_fin"] = profil["date_fin"]
    return insight


def detecter_pics_et_creux(profils):
    faits = []
    for i in range(len(profils)):
        fait = detecter_fait_marquant(profils, i)
        if fait is not None:
            faits.append(fait)
    return faits


def saison_du_mois(mois):
    return SAISON_PAR_MOIS[mois]


# Saisonnalité APPARENTE, jamais démontrée : exige au moins NB_SAISONS_MIN_PROFIL saisons
# représentées par au moins NB_OBSERVATIONS_MIN_SAISON observations chacune avant de proposer une
# lecture -- une seule observation dans une saison ne suffit jamais (voir tests G/H).
def detecter_saisonnalite_apparente(profils):
    volumes_par_saison = {}
    for profil in profils:
        saison = saison_du_mois(profil["date_debut"].month)
        if saison in volumes_par_saison:
            volumes_par_saison[saison].append(profil["volume"])
        else:
            volumes_par_saison[saison] = [profil["volume"]]

    saisons_valables = []
    for saison, volumes in volumes_par_saison.items():
        if len(volumes) >= NB_OBSERVATIONS_MIN_SAISON:
            saisons_valables.append(saison)

    if len(saisons_valables) < NB_SAISONS_MIN_PROFIL:
        return None

    moyennes = []
    somme_generale = 0.0
    for saison in saisons_valables:
        volumes = volumes_par_saison[saison]
        moyenne_saison = sum(volumes) / len(volumes)
        moyennes.append((moyenne_saison, saison))
        somme_generale = somme_generale + moyenne_saison
    moyenne_generale = somme_generale / len(moyennes)

    moyennes_triees = sorted(moyennes, reverse=True)

    saisons_hautes = []
    saisons_basses = []
    for moyenne_saison, saison in moyennes_triees:
        if moyenne_saison >= moyenne_generale * 1.10:
            saisons_hautes.append(saison)
        elif moyenne_saison <= moyenne_generale * 0.90:
            saisons_basses.append(saison)

    if len(saisons_hautes) == 0 and len(saisons_basses) == 0:
        observation = (
            "Les observations disponibles ne dessinent pas de profil saisonnier apparent net : le volume reste "
            "globalement comparable d'une saison à l'autre sur les périodes couvertes."
        )
    else:
        parties = []
        if len(saisons_hautes) > 0:
            parties.append("activité plus soutenue en " + " et ".join(saisons_hautes))
        if len(saisons_basses) > 0:
            parties.append("plus calme en " + " et ".join(saisons_basses))
        observation = "Les observations disponibles dessinent un profil saisonnier apparent : " + ", ".join(parties) + "."

    if len(saisons_valables) >= 4:
        confiance = "modérée (les 4 saisons sont représentées par au moins " + str(NB_OBSERVATIONS_MIN_SAISON) + " observations)"
    else:
        confiance = "limitée (seulement " + str(len(saisons_valables)) + " saison(s) sur 4 avec assez d'observations)"

    return construire_insight(
        "saisonnalite", observation,
        "Une lecture saisonnière aide à distinguer un mouvement récurrent d'un aléa ponctuel.",
        prudence=(
            str(len(profils)) + " observation(s) disponible(s) ne constituent pas une preuve formelle de "
            "saisonnalité -- lecture indicative, confiance " + confiance + "."
        ),
    )


# Analyse d'UNE observation précise (utilisé pour "que dit le moteur sur telle période", en plus
# de la synthèse longue sur l'ensemble). "comparaison_precedente" reste toujours nommée comme une
# comparaison à l'observation disponible juste avant, jamais comme une évolution hebdomadaire.
def analyser_observation(profils, index):
    profil = profils[index]
    fait_marquant = detecter_fait_marquant(profils, index)

    comparaison_precedente = None
    if index > 0:
        profil_precedent = profils[index - 1]
        volume_precedent = profil_precedent["volume"]
        if volume_precedent is not None and volume_precedent > 0:
            ecart = (profil["volume"] - volume_precedent) / volume_precedent
            if abs(ecart) >= SEUIL_ECART_VOLUME_NOTABLE:
                if ecart > 0:
                    texte = (
                        "Reprise observée par rapport à la dernière observation disponible ("
                        + str(profil_precedent["date_debut"]) + ")."
                    )
                else:
                    texte = (
                        "Recul observé par rapport à la dernière observation disponible ("
                        + str(profil_precedent["date_debut"]) + ")."
                    )
                comparaison_precedente = construire_insight(
                    "comparaison_precedente", texte,
                    "Comparaison au dernier point de données disponible, pas une évolution semaine par semaine "
                    "réelle -- l'écart de calendrier entre les deux peut atteindre plusieurs semaines.",
                )

    return {
        "date_debut": profil["date_debut"],
        "date_fin": profil["date_fin"],
        "volume": profil["volume"],
        "fait_marquant": fait_marquant,
        "comparaison_precedente": comparaison_precedente,
        "contexte": profil["contexte"],
    }


def construire_comparaisons_locales(profils):
    comparaisons = []
    for i in range(1, len(profils)):
        resultat = analyser_observation(profils, i)
        if resultat["comparaison_precedente"] is not None:
            comparaison = dict(resultat["comparaison_precedente"])
            comparaison["date_debut"] = profils[i]["date_debut"]
            comparaisons.append(comparaison)
    return comparaisons


# Intensité générique d'un jalon, déduite de son registre déjà détecté (Étape 4B.1, non
# recalculée ici) -- ne connaît aucun nom de mois ni de composant Emyria.
def _intensite_jalon(jalon):
    if jalon["registre"] == "pic absorbé":
        return "exceptionnel_haut"
    if jalon["registre"] == "activité soutenue maîtrisée":
        return "modere_haut"
    if jalon["registre"] == "creux sans tension":
        return "exceptionnel_bas"
    if jalon["registre"] == "activité calme":
        return "modere_bas"
    return "autre"


def _joindre_liste_fr(elements):
    if len(elements) == 0:
        return ""
    if len(elements) == 1:
        return elements[0]
    return ", ".join(elements[:-1]) + " et " + elements[-1]


PREFIXE_OBSERVATION_VIGILANCE = "Cette période se distingue par "


def _capitaliser(texte):
    if len(texte) == 0:
        return texte
    return texte[0].upper() + texte[1:]


# Retire le préambule déjà déduit du "X rompt avec cette dynamique :" qui précède, et ne garde que
# la première phrase (le détail complet reste disponible dans la carte Vigilance) -- pure
# recomposition de texte déjà détecté, aucune donnée ni lecture n'est recalculée.
def _texte_vigilance_condense(observation):
    reste = observation
    if observation.startswith(PREFIXE_OBSERVATION_VIGILANCE):
        reste = observation[len(PREFIXE_OBSERVATION_VIGILANCE):]
    if len(reste) == 0:
        return reste
    index_fin_phrase = reste.find(". ")
    if index_fin_phrase != -1:
        reste = reste[:index_fin_phrase] + "."
    return reste[0].lower() + reste[1:]


# Choisit, parmi un pool de jalons modérés (ni pic ni creux exceptionnel), celui qui apporte le
# plus de valeur narrative pour illustrer une phase de normalisation : priorité à un jalon avec un
# contexte métier réel ET une activité soutenue (plus parlant qu'une simple période calme), sinon
# n'importe quel jalon avec un contexte réel, sinon aucun exemple. Générique -- ne privilégie aucun
# mois en tant que tel, seul le contenu du jalon détermine le choix (Étape 4B.3).
def _choisir_jalon_illustratif(pool):
    candidat_avec_activite = None
    candidat_avec_contexte = None
    for insight in pool:
        if insight["contexte"] is None:
            continue
        if candidat_avec_contexte is None:
            candidat_avec_contexte = insight
        if _intensite_jalon(insight) == "modere_haut" and candidat_avec_activite is None:
            candidat_avec_activite = insight
    if candidat_avec_activite is not None:
        return candidat_avec_activite
    return candidat_avec_contexte


# Synthèse éditoriale condensée (maximum 5 phrases), composée par SÉQUENCES génériques et non par
# récitation chronologique de chaque jalon (Étape 4B.2, affinée en 4B.3) -- la détection elle-même
# (jalons, vigilances, registres, contexte) n'est ni recalculée ni modifiée ici, seulement
# recomposée en texte. Structure fixe mais son contenu émerge entièrement des insights déjà
# détectés :
#   1. temps forts exceptionnels (registre "pic absorbé"), regroupés en une seule phrase ;
#   2. les vigilances -- individuellement si une seule, groupées sinon pour garder le total sous 5 ;
#   3. le reste des jalons modérés, condensés en une phrase de "normalisation" ;
#   4. les creux exceptionnels (registre "creux sans tension"), regroupés en une seule phrase ;
#   5. le jalon le plus récent, toujours cité en dernier -- "où on en est".
# La saisonnalité n'est pas répétée ici : elle est déjà affichée séparément par l'appelant.
def construire_synthese_generale(jalons_metier, vigilances, saisonnalite):
    moments = []
    for jalon in jalons_metier:
        moments.append((jalon["date_debut"], "jalon", jalon))
    for vigilance in vigilances:
        moments.append((vigilance["date_debut"], "vigilance", vigilance))
    moments_tries = sorted(moments)

    if len(moments_tries) == 0:
        return "Les observations disponibles restent relativement homogènes, sans rupture opérationnelle marquée."

    # Le moment le plus récent (jalon OU vigilance, jamais restreint aux seuls jalons) ferme
    # naturellement le récit. S'il s'agit d'une vigilance, sa propre phrase EST déjà "où on en
    # est" : aucune phrase de clôture distincte n'est ajoutée, et aucun jalon n'est exclu des
    # groupes ci-dessous pour cette seule raison.
    _date_derniere, type_dernier_moment, insight_dernier_moment = moments_tries[-1]
    dernier_jalon = None
    if type_dernier_moment == "jalon":
        dernier_jalon = insight_dernier_moment

    date_derniere_vigilance = None
    for date_debut, type_moment, insight in moments_tries:
        if type_moment == "vigilance":
            date_derniere_vigilance = date_debut

    pics = []
    creux = []
    normalisation = []
    for date_debut, type_moment, insight in moments_tries:
        if type_moment != "jalon" or insight is dernier_jalon:
            continue
        intensite = _intensite_jalon(insight)
        if intensite == "exceptionnel_haut":
            pics.append(insight)
        elif intensite == "exceptionnel_bas":
            creux.append(insight)
        else:
            # "Revenir vers un fonctionnement plus maîtrisé" n'a de sens qu'APRÈS une rupture --
            # un jalon modéré antérieur à la dernière vigilance n'illustre pas un retour au calme,
            # il reste silencieusement dans le détail (comme un jalon sans exemple narratif dédié).
            if date_derniere_vigilance is not None and date_debut < date_derniere_vigilance:
                continue
            normalisation.append(insight)

    phrases = []

    if len(pics) > 0:
        mois_pics = []
        for insight in pics:
            mois_pics.append(nom_mois(insight["date_debut"].month))
        phrases.append(
            "L'activité est fortement rythmée par les temps forts de " + _joindre_liste_fr(mois_pics)
            + ", concentrant les niveaux les plus élevés des observations disponibles, globalement bien absorbés."
        )

    vigilance_moments = []
    for date_debut, type_moment, insight in moments_tries:
        if type_moment == "vigilance":
            vigilance_moments.append(insight)

    if len(vigilance_moments) == 1:
        insight = vigilance_moments[0]
        mois = nom_mois(insight["date_debut"].month)
        phrases.append(
            _capitaliser(mois) + " rompt avec cette dynamique : " + _texte_vigilance_condense(insight["observation"])
        )
    elif len(vigilance_moments) > 1:
        mois_vigilances = []
        for insight in vigilance_moments:
            mois_vigilances.append(nom_mois(insight["date_debut"].month))
        phrases.append(
            _capitaliser(_joindre_liste_fr(mois_vigilances)) + " rompent avec cette dynamique, avec une "
            "expérience et/ou un effort de traitement dégradés."
        )

    if len(normalisation) > 0:
        exemple = _choisir_jalon_illustratif(normalisation)
        if exemple is not None:
            mois_exemple = nom_mois(exemple["date_debut"].month)
            if _intensite_jalon(exemple) == "modere_haut":
                phrases.append(
                    "Le fonctionnement redevient plus maîtrisé sur les périodes suivantes, à l'image de "
                    + mois_exemple + ", où l'activité reste soutenue et bien tenue."
                )
            else:
                phrases.append(
                    "Le fonctionnement redevient plus maîtrisé sur les périodes suivantes, à l'image de "
                    + mois_exemple + ", plus calme et sans tension particulière."
                )
        else:
            phrases.append("Le fonctionnement redevient plus maîtrisé sur les périodes suivantes.")

    if len(creux) > 0:
        mois_creux = []
        for insight in creux:
            mois_creux.append(nom_mois(insight["date_debut"].month))
        phrases.append("Un creux se dessine sur " + _joindre_liste_fr(mois_creux) + ", sans tension particulière.")

    if dernier_jalon is not None:
        mois_dernier = nom_mois(dernier_jalon["date_debut"].month)
        if _intensite_jalon(dernier_jalon) in ("exceptionnel_haut", "modere_haut"):
            phrases.append(
                _capitaliser(mois_dernier) + " montre une reprise de l'activité, avec une expérience qui reste "
                "correctement tenue."
            )
        else:
            phrases.append(_capitaliser(mois_dernier) + " referme la période sans tension particulière.")

    return " ".join(phrases)


# Orchestration de la vue longue : synthèse générale condensée + jalons métier + vigilances +
# contrastes (périodes comparables mais inégalement difficiles) + comparaisons locales (vs
# observation précédente) + saisonnalité apparente + niveau de confiance -- jamais un nombre
# imposé d'insights, jamais une formulation de série continue.
def construire_synthese_longue(profils):
    if len(profils) == 0:
        return {
            "synthese_generale": "Aucune observation disponible sur cette période.",
            "jalons_metier": [],
            "vigilances": [],
            "contrastes": [],
            "comparaisons_locales": [],
            "saisonnalite": None,
            "niveau_confiance": "aucune donnée",
            "nb_observations": 0,
        }

    faits = detecter_pics_et_creux(profils)
    jalons_metier = []
    vigilances = []
    for fait in faits:
        if fait["categorie"] == "jalon":
            jalons_metier.append(fait)
        else:
            vigilances.append(fait)

    contrastes = detecter_contrastes_capacite(profils)
    saisonnalite = detecter_saisonnalite_apparente(profils)
    comparaisons_locales = construire_comparaisons_locales(profils)
    synthese_generale = construire_synthese_generale(jalons_metier, vigilances, saisonnalite)

    if len(profils) >= 10:
        niveau_confiance = "correct pour une lecture d'ensemble (" + str(len(profils)) + " observations sur la période couverte)"
    elif len(profils) >= 4:
        niveau_confiance = "modéré (" + str(len(profils)) + " observations disponibles)"
    else:
        niveau_confiance = "limité (" + str(len(profils)) + " observation(s) seulement)"

    return {
        "synthese_generale": synthese_generale,
        "jalons_metier": jalons_metier,
        "vigilances": vigilances,
        "contrastes": contrastes,
        "comparaisons_locales": comparaisons_locales,
        "saisonnalite": saisonnalite,
        "niveau_confiance": niveau_confiance,
        "nb_observations": len(profils),
    }


# ------------------------------------------------------------------------------------------------
# Étape 4B.3 -- scope de période : PÉRIODE ANALYSÉE (ce que l'utilisatrice a demandé à regarder) vs
# HISTORIQUE DE RÉFÉRENCE (observations antérieures disponibles pour contextualiser). La logique
# analytique elle-même (rang_relatif, détection jalons/vigilances/contrastes, saisonnalité,
# seuils...) n'est ni recalculée ni modifiée ici -- ce bloc décide uniquement QUEL périmètre lui est
# soumis et QUELLE PARTIE du résultat est présentée, selon le mode déduit du scope sélectionné.
# ------------------------------------------------------------------------------------------------


def _texte_titre_situation_semaine(fait):
    if fait is None:
        return "Une période sans signal particulier"
    if fait["categorie"] == "vigilance":
        return "Une période qui appelle une vigilance"
    if fait["registre"] == "pic absorbé":
        return "Un pic d'activité bien absorbé"
    if fait["registre"] == "activité soutenue maîtrisée":
        return "Une activité soutenue et maîtrisée"
    if fait["registre"] == "creux sans tension":
        return "Un creux d'activité sans tension"
    if fait["registre"] == "activité calme":
        return "Une période calme"
    return "Un mix de demandes inhabituel"


# MODE 1 : une seule observation sélectionnée -- profils_historique contient uniquement des
# observations dont la date est antérieure ou égale à la fin de cette semaine (jamais de fuite du
# futur, par construction de l'appelant, voir app.py). Le sujet est CETTE semaine, jamais un récit
# annuel : ni liste de Jalons historiques, ni saisonnalité globale, ni graphique cadré sur toute
# l'année sans distinction visuelle.
def _lecture_mode_observation_unique(profils_historique):
    index_semaine = len(profils_historique) - 1
    profil = profils_historique[index_semaine]
    contexte_texte = texte_contexte(profil["contexte"])

    if index_semaine == 0:
        return {
            "mode": MODE_OBSERVATION_UNIQUE,
            "titre": "Comment se situe cette période ?",
            "synthese": "Première observation disponible : aucun historique antérieur suffisant pour situer cette période.",
            "reperes": [str(profil["volume"]) + " demandes sur cette période."],
            "contexte": contexte_texte,
            "jalons_metier": [],
            "vigilances": [],
            "contrastes": [],
            "saisonnalite": None,
            "niveau_confiance": "Aucun historique antérieur disponible pour situer cette période.",
            "nb_observations_periode": 1,
            "nb_observations_historique": 1,
            "premiere_observation_sans_historique": True,
            "profils_periode": [profil],
        }

    resultat = analyser_observation(profils_historique, index_semaine)
    fait = resultat["fait_marquant"]
    comparaison_precedente = resultat["comparaison_precedente"]

    phrases = [_texte_titre_situation_semaine(fait) + "."]
    if fait is not None:
        phrases.append(fait["observation"])
    else:
        phrases.append(
            str(profil["volume"]) + " demandes sur cette période, sans écart notable par rapport à "
            "l'historique disponible."
        )
        if comparaison_precedente is not None:
            phrases.append(comparaison_precedente["observation"])

    vigilances_mode1 = []
    if fait is not None and fait["categorie"] == "vigilance":
        vigilances_mode1 = [fait]

    contrastes_historique = detecter_contrastes_capacite(profils_historique)
    contrastes_mode1 = []
    for contraste in contrastes_historique:
        if contraste["date_debut_a"] == profil["date_debut"] or contraste["date_debut_b"] == profil["date_debut"]:
            contrastes_mode1.append(contraste)

    niveau_confiance = (
        "Cette période est replacée parmi " + str(index_semaine) + " observation(s) antérieure(s) disponible(s)."
    )

    return {
        "mode": MODE_OBSERVATION_UNIQUE,
        "titre": "Comment se situe cette période ?",
        "synthese": " ".join(phrases),
        "reperes": [str(profil["volume"]) + " demandes sur cette période."],
        "contexte": contexte_texte,
        "jalons_metier": [],  # jamais de liste Jalons en mode 1 -- repris dans la synthèse si pertinent
        "vigilances": vigilances_mode1,
        "contrastes": contrastes_mode1,
        "saisonnalite": None,
        "niveau_confiance": niveau_confiance,
        "nb_observations_periode": 1,
        "nb_observations_historique": len(profils_historique),
        "premiere_observation_sans_historique": False,
        "profils_periode": [profil],
    }


# MODE 2 : plusieurs observations sélectionnées, mais pas tout l'historique disponible jusqu'à
# cette date. Jalons/vigilances/contrastes sont détectés sur profils_historique (rang_relatif a
# besoin du contexte complet pour rester correct) puis FILTRÉS pour ne garder que ceux dont la date
# appartient à la fenêtre sélectionnée -- une observation antérieure à la fenêtre peut nourrir un
# contraste ou un rang, mais ne devient jamais un Jalon "de la période".
def _lecture_mode_periode_etendue(profils_historique, indice_debut_periode):
    faits = detecter_pics_et_creux(profils_historique)
    jalons_fenetre = []
    vigilances_fenetre = []
    for fait in faits:
        if fait["date_debut"] < profils_historique[indice_debut_periode]["date_debut"]:
            continue
        if fait["categorie"] == "jalon":
            jalons_fenetre.append(fait)
        else:
            vigilances_fenetre.append(fait)

    contrastes_historique = detecter_contrastes_capacite(profils_historique)
    date_debut_fenetre = profils_historique[indice_debut_periode]["date_debut"]
    contrastes_fenetre = []
    for contraste in contrastes_historique:
        if contraste["date_debut_a"] >= date_debut_fenetre or contraste["date_debut_b"] >= date_debut_fenetre:
            contrastes_fenetre.append(contraste)

    profils_fenetre = profils_historique[indice_debut_periode:]
    saisonnalite_fenetre = detecter_saisonnalite_apparente(profils_fenetre)

    synthese = construire_synthese_generale(jalons_fenetre, vigilances_fenetre, saisonnalite_fenetre)

    niveau_confiance = (
        "Lecture sur les " + str(len(profils_fenetre)) + " observation(s) de cette période (" +
        str(len(profils_historique)) + " observations disponibles au total jusqu'à cette date)."
    )

    return {
        "mode": MODE_PERIODE_ETENDUE,
        "titre": "Ce que raconte la période sélectionnée",
        "synthese": synthese,
        "reperes": [],
        "contexte": None,
        "jalons_metier": jalons_fenetre,
        "vigilances": vigilances_fenetre,
        "contrastes": contrastes_fenetre,
        "saisonnalite": saisonnalite_fenetre,
        "niveau_confiance": niveau_confiance,
        "nb_observations_periode": len(profils_fenetre),
        "nb_observations_historique": len(profils_historique),
        "premiere_observation_sans_historique": False,
        "profils_periode": profils_fenetre,
    }


def _lecture_mode_historique_complet(profils_historique):
    synthese_longue = construire_synthese_longue(profils_historique)
    return {
        "mode": MODE_HISTORIQUE_COMPLET,
        "titre": "Ce que racontent les observations disponibles",
        "synthese": synthese_longue["synthese_generale"],
        "reperes": [],
        "contexte": None,
        "jalons_metier": synthese_longue["jalons_metier"],
        "vigilances": synthese_longue["vigilances"],
        "contrastes": synthese_longue["contrastes"],
        "saisonnalite": synthese_longue["saisonnalite"],
        "niveau_confiance": "Niveau de confiance de cette lecture : " + synthese_longue["niveau_confiance"] + ".",
        "nb_observations_periode": len(profils_historique),
        "nb_observations_historique": len(profils_historique),
        "premiere_observation_sans_historique": False,
        "profils_periode": profils_historique,
    }


# Point d'entrée unique de l'onglet Tendances (Étape 4B.3). `profils_historique` doit être construit
# par l'appelant à partir des SEULS exports dont la date est antérieure ou égale à la fin de la
# période sélectionnée (aucune fuite du futur) ; `nb_observations_periode` est le nombre de ces
# observations qui appartiennent effectivement à la sélection de l'utilisatrice (les autres, en
# tête de liste, ne forment que l'historique de référence). Le mode est déduit uniquement de ces
# deux comptes -- jamais d'une date ou d'un mois codé en dur.
def construire_lecture_tendances(profils_historique, nb_observations_periode):
    if len(profils_historique) == 0 or nb_observations_periode == 0:
        return {
            "mode": MODE_OBSERVATION_UNIQUE,
            "titre": "Comment se situe cette période ?",
            "synthese": "Aucune observation disponible sur cette période.",
            "reperes": [],
            "contexte": None,
            "jalons_metier": [],
            "vigilances": [],
            "contrastes": [],
            "saisonnalite": None,
            "niveau_confiance": "aucune donnée",
            "nb_observations_periode": 0,
            "nb_observations_historique": len(profils_historique),
            "premiere_observation_sans_historique": True,
            "profils_periode": [],
        }

    n_historique = len(profils_historique)
    mode = determiner_mode_tendances(nb_observations_periode, n_historique)

    if mode == MODE_HISTORIQUE_COMPLET:
        return _lecture_mode_historique_complet(profils_historique)
    if mode == MODE_OBSERVATION_UNIQUE:
        return _lecture_mode_observation_unique(profils_historique)
    indice_debut_periode = n_historique - nb_observations_periode
    return _lecture_mode_periode_etendue(profils_historique, indice_debut_periode)


# ---------------------------------------------------------------------------
# Moteur Livraison (Étape 4C)
#
# Même philosophie que le moteur Produit (voie A) : volume élevé n'est pas un problème en soi
# (Black Friday/Noël génèrent naturellement du suivi de commande), l'éligibilité repose sur la
# CONVERGENCE de plusieurs familles de preuve indépendantes -- jamais un score arbitraire seul --
# et au moins une famille de "conséquence opérationnelle/client" (Expérience, Effort, Relances,
# Issue défavorable, Coût) est requise pour la "Priorité" ; Volume et Concentration transporteur
# restent des familles de CONTEXTE, jamais suffisantes seules.
#
# Grain unique : subject_cluster ("motif") -- transporteur/issue/coût sont des ÉLÉMENTS
# D'EXPLICATION d'un signal déjà identifié par motif, jamais des cartes séparées par combinaison
# (voir section 4 du prompt 4C).
# ---------------------------------------------------------------------------

SEUIL_MINIMUM_EVALUATION_LIVRAISON = SEUIL_MINIMUM_EVALUATION_PRODUIT  # même seuil, cohérence transversale

# Issues finales considérées "défavorables" : elles impliquent une compensation réelle
# (réexpédition, remboursement, geste commercial) ou un dossier resté sans résolution connue.
# "Livraison confirmée", "Colis retrouvé" et "Erreur d'adresse corrigée" se résolvent sans
# compensation -- classification métier explicite, pas une déduction automatique du CSAT.
ISSUES_LIVRAISON_DEFAVORABLES = ("Réexpédié", "Remboursé", "Geste commercial", "Sans réponse client")

# Taxonomie connue telle que communiquée (Étape 4C, section 2) -- sert uniquement au contrôle
# qualité (détecter une valeur hors taxonomie), jamais à filtrer silencieusement les tickets.
ISSUES_LIVRAISON_CONNUES = (
    "Colis retrouvé", "Réexpédié", "Remboursé", "Erreur d'adresse corrigée",
    "Livraison confirmée", "Geste commercial", "Sans réponse client",
)

POIDS_SCORE_LIVRAISON_CSAT = 0.25
POIDS_SCORE_LIVRAISON_EFFORT = 0.20
POIDS_SCORE_LIVRAISON_RELANCES = 0.20
POIDS_SCORE_LIVRAISON_ISSUE = 0.15
POIDS_SCORE_LIVRAISON_TEMPORALITE = 0.10
POIDS_SCORE_LIVRAISON_VOLUME = 0.10

PLAFOND_SCORE_LIVRAISON_ECART_RELATIF = 1.0  # un écart relatif au-delà de +100 % compte comme maximum


def part_issues_defavorables(tickets):
    if len(tickets) == 0:
        return None
    n_defavorable = 0
    for ticket in tickets:
        if ticket["issue_livraison_finale"] in ISSUES_LIVRAISON_DEFAVORABLES:
            n_defavorable = n_defavorable + 1
    return n_defavorable / len(tickets)


def _obtenir_n_distribution_issue(item):
    return item["n"]


# Distribution des issues finales d'un ensemble de tickets -- "qu'est-il finalement arrivé à ces
# dossiers ?" (Étape 4C, section 8). Retourne uniquement les issues réellement présentes, triées
# par volume décroissant.
def distribution_issues_livraison(tickets):
    total = len(tickets)
    if total == 0:
        return []

    comptes = {}
    for ticket in tickets:
        issue = ticket["issue_livraison_finale"]
        if issue in comptes:
            comptes[issue] = comptes[issue] + 1
        else:
            comptes[issue] = 1

    distribution = []
    for issue, n in comptes.items():
        distribution.append({"issue": issue, "n": n, "part_pct": n / total * 100})

    return sorted(distribution, key=_obtenir_n_distribution_issue, reverse=True)


# Niveau historique d'un sujet Livraison = sa part du volume Livraison total, observation par
# observation antérieure -- même principe que construire_niveaux_historiques (Produit), mais sur
# le grain unique subject_cluster (pas de tuple de grain à gérer).
def construire_niveaux_historiques_livraison(historique_livraison_par_fichier, sujet):
    niveaux = []
    for tickets_fichier in historique_livraison_par_fichier:
        if len(tickets_fichier) == 0:
            continue
        tickets_sujet_fichier = []
        for ticket in tickets_fichier:
            if ticket["subject_cluster"] == sujet:
                tickets_sujet_fichier.append(ticket)
        part = calculer_part(len(tickets_sujet_fichier), len(tickets_fichier))
        if part is not None:
            niveaux.append(part)
    return niveaux


# Le transporteur est une dimension d'INVESTIGATION, jamais une identité de signal (Étape 4C.2) :
# il ne décide plus de rien (l'éligibilité du motif est déjà tranchée ailleurs, sans lui). La
# question n'est plus "ce transporteur constitue-t-il un signal ?" mais "apporte-t-il une piste
# utile à l'intérieur d'un problème déjà identifié ?" -- il suffit donc qu'UN SEUL des 3
# indicateurs (expérience, relances, issues) montre un contraste lisible avec le reste du MÊME
# motif, sur un échantillon qui respecte le seuil minimum déjà utilisé partout ailleurs dans ce
# moteur (SEUIL_MINIMUM_EVALUATION_LIVRAISON) -- pas de nouveau seuil inventé. "Transporteur X
# représente 60 % des tickets" ne suffit toujours pas seul (ça peut simplement refléter la part
# des commandes expédiées avec X, que nous ne connaissons pas) -- jamais de "taux d'incident
# transporteur", seulement une "part des dossiers observés associés à X".
def _transporteur_merite_investigation(lecture_csat, lecture_relances, lecture_issues):
    return (
        lecture_csat == VOCABULAIRE_ECART_MARQUE
        or lecture_relances == VOCABULAIRE_ECART_MARQUE
        or lecture_issues == VOCABULAIRE_ECART_MARQUE
    )


def evaluer_concentration_transporteur_livraison(tickets_sujet):
    transporteurs = set()
    for ticket in tickets_sujet:
        transporteurs.add(ticket["transporteur"])

    meilleur_info = None
    meilleur_nb_marqueurs = -1

    for transporteur in transporteurs:
        tickets_transporteur = []
        tickets_reste = []
        for ticket in tickets_sujet:
            if ticket["transporteur"] == transporteur:
                tickets_transporteur.append(ticket)
            else:
                tickets_reste.append(ticket)

        if len(tickets_transporteur) < SEUIL_MINIMUM_EVALUATION_LIVRAISON:
            continue

        csat_transporteur = moyenne(tickets_transporteur, "csat")
        csat_reste = moyenne(tickets_reste, "csat")
        lecture_csat_transporteur = lire_ecart_csat(csat_transporteur, csat_reste)

        relances_transporteur = moyenne(tickets_transporteur, "nombre_relances")
        relances_reste = moyenne(tickets_reste, "nombre_relances")
        lecture_relances_transporteur = lire_ecart_effort(relances_transporteur, relances_reste)

        part_defavorable_transporteur = part_issues_defavorables(tickets_transporteur)
        part_defavorable_reste = part_issues_defavorables(tickets_reste)
        lecture_issue_transporteur = lire_ecart_effort(part_defavorable_transporteur, part_defavorable_reste)

        if not _transporteur_merite_investigation(
            lecture_csat_transporteur, lecture_relances_transporteur, lecture_issue_transporteur
        ):
            continue

        nb_marqueurs = 0
        if lecture_csat_transporteur == VOCABULAIRE_ECART_MARQUE:
            nb_marqueurs = nb_marqueurs + 1
        if lecture_relances_transporteur == VOCABULAIRE_ECART_MARQUE:
            nb_marqueurs = nb_marqueurs + 1
        if lecture_issue_transporteur == VOCABULAIRE_ECART_MARQUE:
            nb_marqueurs = nb_marqueurs + 1

        if nb_marqueurs > meilleur_nb_marqueurs:
            meilleur_nb_marqueurs = nb_marqueurs
            meilleur_info = {
                "transporteur": transporteur,
                "n": len(tickets_transporteur),
                "part_du_motif_pct": len(tickets_transporteur) / len(tickets_sujet) * 100,
                "csat": csat_transporteur, "csat_reste_motif": csat_reste, "lecture_csat": lecture_csat_transporteur,
                "relances_moyen": relances_transporteur, "relances_reste_motif": relances_reste,
                "lecture_relances": lecture_relances_transporteur,
                "part_defavorable_pct": (
                    part_defavorable_transporteur * 100 if part_defavorable_transporteur is not None else None
                ),
                "part_defavorable_reste_motif_pct": (
                    part_defavorable_reste * 100 if part_defavorable_reste is not None else None
                ),
                "lecture_issue": lecture_issue_transporteur,
                "prudence_echantillon": (
                    "n=" + str(len(tickets_transporteur)) + " dossiers observés pour ce transporteur sur ce "
                    "motif -- association observée sur l'échantillon disponible, pas une cause démontrée."
                ),
            }

    return meilleur_info


# Phrase "piste d'investigation", jamais présentée comme un signal ou un diagnostic (Étape 4C.2,
# section 10) -- ne mentionne que les indicateurs réellement marqués pour CE transporteur, jamais
# de "taux d'incident" ni de jugement sur la fiabilité du transporteur lui-même.
def texte_piste_transporteur_livraison(info_transporteur):
    elements = []
    if info_transporteur["lecture_csat"] == VOCABULAIRE_ECART_MARQUE:
        elements.append("une satisfaction plus basse dans cet échantillon")
    if info_transporteur["lecture_relances"] == VOCABULAIRE_ECART_MARQUE:
        elements.append("davantage de relances dans cet échantillon")
    if info_transporteur["lecture_issue"] == VOCABULAIRE_ECART_MARQUE:
        elements.append("davantage d'issues défavorables dans cet échantillon")

    texte = (
        formater_pourcentage(info_transporteur["part_du_motif_pct"]) + " des dossiers observés sur ce motif "
        "sont associés à " + info_transporteur["transporteur"]
    )
    if len(elements) > 0:
        texte = texte + " ; ils présentent " + _joindre_liste_fr(elements) + "."
    else:
        texte = texte + "."
    return texte


# Contrôle qualité (Étape 4C, section 24) : ne corrige rien, se contente de rapporter ce qui
# s'écarte de ce que les données sont censées contenir -- champs manquants, valeurs hors
# taxonomie connue, valeurs impossibles (négatives). Volontairement silencieux sur toute
# incohérence qui demanderait une heuristique fragile pour être détectée.
def controler_qualite_donnees_livraison(tickets_livraison):
    anomalies = []

    n_sans_transporteur = 0
    n_sans_relances = 0
    n_relances_negatives = 0
    n_sans_issue = 0
    n_issue_hors_taxonomie = 0
    n_resolution_negative = 0
    n_replies_negatifs = 0

    for ticket in tickets_livraison:
        if ticket["transporteur"] is None or ticket["transporteur"] == "":
            n_sans_transporteur = n_sans_transporteur + 1

        if ticket["nombre_relances"] is None:
            n_sans_relances = n_sans_relances + 1
        elif ticket["nombre_relances"] < 0:
            n_relances_negatives = n_relances_negatives + 1

        if ticket["issue_livraison_finale"] is None or ticket["issue_livraison_finale"] == "":
            n_sans_issue = n_sans_issue + 1
        elif ticket["issue_livraison_finale"] not in ISSUES_LIVRAISON_CONNUES:
            n_issue_hors_taxonomie = n_issue_hors_taxonomie + 1

        if ticket["full_resolution_time_hours"] is not None and ticket["full_resolution_time_hours"] < 0:
            n_resolution_negative = n_resolution_negative + 1
        if ticket["replies"] is not None and ticket["replies"] < 0:
            n_replies_negatifs = n_replies_negatifs + 1

    if n_sans_transporteur > 0:
        anomalies.append(str(n_sans_transporteur) + " ticket(s) Livraison sans transporteur renseigné.")
    if n_sans_relances > 0:
        anomalies.append(str(n_sans_relances) + " ticket(s) Livraison sans nombre_relances renseigné.")
    if n_relances_negatives > 0:
        anomalies.append(str(n_relances_negatives) + " ticket(s) Livraison avec un nombre_relances négatif.")
    if n_sans_issue > 0:
        anomalies.append(str(n_sans_issue) + " ticket(s) Livraison sans issue_livraison_finale renseignée.")
    if n_issue_hors_taxonomie > 0:
        anomalies.append(
            str(n_issue_hors_taxonomie) + " ticket(s) Livraison avec une issue_livraison_finale hors taxonomie connue."
        )
    if n_resolution_negative > 0:
        anomalies.append(str(n_resolution_negative) + " ticket(s) Livraison avec un temps de résolution négatif.")
    if n_replies_negatifs > 0:
        anomalies.append(str(n_replies_negatifs) + " ticket(s) Livraison avec un nombre d'échanges négatif.")

    return anomalies


# Piste d'investigation générique, construite à partir des familles réellement actives -- jamais
# une cause prétendue, une invitation à vérifier (Étape 4C, section 18).
def texte_action_investigation_livraison(familles_actives, concentration_info):
    pistes = []
    if "D" in familles_actives:
        pistes.append("la séquence de relances")
    if "E" in familles_actives:
        pistes.append("les dossiers avec une issue défavorable")
    if "C" in familles_actives:
        pistes.append("les dossiers à résolution longue")
    if concentration_info is not None:
        pistes.append("les dossiers associés au transporteur " + concentration_info["transporteur"])
    if len(pistes) == 0:
        return "À investiguer : vérifier un échantillon de dossiers récents sur ce motif."
    return "À investiguer : vérifier " + _joindre_liste_fr(pistes) + "."


# Signal multicritère pour UN motif Livraison (grain unique : subject_cluster). Familles de
# preuve (Étape 4C, affinées en 4C.1) :
#   A Demande       -- volume/part au-dessus du niveau habituel (jamais suffisante seule)
#   B Expérience    -- CSAT du motif vs reste de Livraison sur la période
#   C Effort        -- échanges / résolution / réouvertures vs reste de Livraison -- compte pour
#                      l'ÉLIGIBILITÉ uniquement si fortement étayé (voir c_fort_actif ci-dessous),
#                      jamais sur un seul indicateur qui bouge légèrement (Étape 4C.1, section 4)
#   D Relances      -- nombre_relances vs reste de Livraison (dimension propre à Livraison)
#   E Issue         -- part d'issues défavorables vs reste de Livraison
#   F Coût          -- non disponible actuellement (aucun order_id sur les tickets Livraison
#                      observés) -- toujours inactive, jamais simulée, jamais affichée comme "0"
#   H Persistance   -- part du volume Livraison inhabituelle sur plusieurs observations antérieures
# Le transporteur n'est plus une famille d'éligibilité (Étape 4C.1, section 5) : c'est une
# dimension d'INVESTIGATION, calculée et affichée uniquement à l'intérieur d'un signal déjà
# éligible au niveau du motif -- jamais un signal transporteur autonome, jamais un facteur qui
# fait basculer un motif en priorité.
def construire_signal_sujet_livraison_voie_a(sujet, tickets_sujet, tickets_livraison_periode, niveaux_historiques):
    n = len(tickets_sujet)
    if n < SEUIL_MINIMUM_EVALUATION_LIVRAISON:
        return None

    total_univers = len(tickets_livraison_periode)
    part_candidat = calculer_part(n, total_univers)
    ecart_temporel = ecart_relatif_temporel(niveaux_historiques, part_candidat)
    texte_temporalite = evaluer_temporalite(niveaux_historiques, part_candidat)
    confiance = confiance_historique(len(niveaux_historiques))
    _persistance_brute, part_elevees, nb_observations = evaluer_persistance_temporelle(
        niveaux_historiques, part_candidat
    )

    temporel_compte = (
        ecart_temporel is not None
        and ecart_temporel >= SEUIL_ECART_RELATIF_TEMPORALITE
        and confiance >= SEUIL_CONFIANCE_TEMPORELLE_MINIMALE
    )

    csat_candidat = moyenne(tickets_sujet, "csat")
    n_csat_candidat = _compte_valeurs_non_nulles(tickets_sujet, "csat")
    csat_reference = moyenne(tickets_livraison_periode, "csat")
    n_csat_reference = _compte_valeurs_non_nulles(tickets_livraison_periode, "csat")
    lecture_csat = lire_ecart_csat(csat_candidat, csat_reference)
    ecart_csat = None
    if csat_candidat is not None and csat_reference is not None:
        ecart_csat = csat_reference - csat_candidat

    replies_candidat = moyenne(tickets_sujet, "replies")
    replies_reference = moyenne(tickets_livraison_periode, "replies")
    lecture_replies = lire_ecart_effort(replies_candidat, replies_reference)

    resolution_candidat = moyenne(tickets_sujet, "full_resolution_time_hours")
    resolution_reference = moyenne(tickets_livraison_periode, "full_resolution_time_hours")
    lecture_resolution = lire_ecart_effort(resolution_candidat, resolution_reference)

    reopens_candidat = moyenne(tickets_sujet, "reopens")
    reopens_reference = moyenne(tickets_livraison_periode, "reopens")
    lecture_reopens = lire_ecart_effort(reopens_candidat, reopens_reference)

    ecart_effort_max = 0.0
    for valeur_candidat_effort, valeur_reference_effort in (
        (replies_candidat, replies_reference),
        (resolution_candidat, resolution_reference),
        (reopens_candidat, reopens_reference),
    ):
        if valeur_candidat_effort is not None and valeur_reference_effort is not None and valeur_reference_effort > 0:
            ecart = (valeur_candidat_effort - valeur_reference_effort) / valeur_reference_effort
            if ecart > ecart_effort_max:
                ecart_effort_max = ecart

    relances_candidat = moyenne(tickets_sujet, "nombre_relances")
    relances_reference = moyenne(tickets_livraison_periode, "nombre_relances")
    lecture_relances = lire_ecart_effort(relances_candidat, relances_reference)
    ecart_relances = None
    if relances_candidat is not None and relances_reference is not None and relances_reference > 0:
        ecart_relances = (relances_candidat - relances_reference) / relances_reference

    part_defavorable_candidat = part_issues_defavorables(tickets_sujet)
    part_defavorable_reference = part_issues_defavorables(tickets_livraison_periode)
    lecture_issue = lire_ecart_effort(part_defavorable_candidat, part_defavorable_reference)
    ecart_issue_defavorable = None
    if (
        part_defavorable_candidat is not None and part_defavorable_reference is not None
        and part_defavorable_reference > 0
    ):
        ecart_issue_defavorable = (part_defavorable_candidat - part_defavorable_reference) / part_defavorable_reference

    # Transporteur : dimension d'investigation, calculée systématiquement mais jamais utilisée
    # pour décider si CE motif est prioritaire (Étape 4C.1, section 5-6) -- voir plus bas, affichée
    # uniquement lorsque le motif est déjà éligible par ailleurs.
    concentration_info = evaluer_concentration_transporteur_livraison(tickets_sujet)

    # ---- Familles de preuve, indépendantes ----
    a_actif = (
        n >= SEUIL_VOLUME_ABSOLU_NOTABLE
        or (part_candidat is not None and part_candidat >= SEUIL_VOLUME_PART_NOTABLE)
        or temporel_compte
    )
    b_actif = lecture_csat == VOCABULAIRE_ECART_MARQUE

    nb_dimensions_effort_marquees = 0
    if lecture_replies == VOCABULAIRE_ECART_MARQUE:
        nb_dimensions_effort_marquees = nb_dimensions_effort_marquees + 1
    if lecture_resolution == VOCABULAIRE_ECART_MARQUE:
        nb_dimensions_effort_marquees = nb_dimensions_effort_marquees + 1
    if lecture_reopens == VOCABULAIRE_ECART_MARQUE:
        nb_dimensions_effort_marquees = nb_dimensions_effort_marquees + 1
    c_actif = nb_dimensions_effort_marquees >= 1  # informatif -- affiché dès qu'un indicateur bouge
    # "Fortement étayé" (Étape 4C.1, section 4) : soit au moins 2 des 3 indicateurs d'effort
    # convergent, soit un seul indicateur montre un écart réellement marqué -- au moins le double
    # du seuil déjà utilisé pour "écart marqué", pas un nouveau seuil inventé. Seul C compte cette
    # distinction : B/D/E sont chacune une comparaison unique et déjà bien définie, pas un OU de
    # plusieurs indicateurs bruités comme C.
    c_fort_actif = nb_dimensions_effort_marquees >= 2 or ecart_effort_max >= 2 * SEUIL_EFFORT_ECART_MARQUE_REL

    d_actif = lecture_relances == VOCABULAIRE_ECART_MARQUE
    e_actif = lecture_issue == VOCABULAIRE_ECART_MARQUE
    f_actif = False  # coût jamais calculable actuellement -- aucun order_id sur les tickets Livraison
    h_actif = (
        temporel_compte
        and part_elevees >= SEUIL_PART_ELEVEES_PERSISTANCE
        and nb_observations >= NB_OBSERVATIONS_MIN_PERSISTANCE
    )

    # Liste informative (affichage/contexte) -- C y figure dès qu'il est actif, même faiblement.
    familles_actives = []
    if a_actif:
        familles_actives.append("A")
    if b_actif:
        familles_actives.append("B")
    if c_actif:
        familles_actives.append("C")
    if d_actif:
        familles_actives.append("D")
    if e_actif:
        familles_actives.append("E")
    if f_actif:
        familles_actives.append("F")
    if h_actif:
        familles_actives.append("H")

    # ---- Éligibilité (Étape 4C.1) : C ne compte ici que fortement étayé (c_fort_actif) --
    # jamais le transporteur (G), retiré de l'éligibilité (section 5-6). B/D/E/F comptent dès leur
    # seuil "marqué" habituel, déjà une comparaison unique et bien définie.
    familles_eligibilite = []
    if a_actif:
        familles_eligibilite.append("A")
    if b_actif:
        familles_eligibilite.append("B")
    if c_fort_actif:
        familles_eligibilite.append("C")
    if d_actif:
        familles_eligibilite.append("D")
    if e_actif:
        familles_eligibilite.append("E")
    if f_actif:
        familles_eligibilite.append("F")
    if h_actif:
        familles_eligibilite.append("H")

    consequences_eligibilite = []
    for lettre_famille in familles_eligibilite:
        if lettre_famille in ("B", "C", "D", "E", "F"):
            consequences_eligibilite.append(lettre_famille)
    nb_consequences = len(consequences_eligibilite)
    nb_total_eligibilite = len(familles_eligibilite)

    tier = None
    regle_eligibilite = None
    if nb_total_eligibilite >= 3 and nb_consequences >= 2:
        tier = "priorite_principale"
        regle_eligibilite = "convergence d'au moins 3 familles de preuve, dont au moins 2 conséquences opérationnelles/client"
    elif nb_consequences >= 2:
        tier = "priorite_secondaire"
        regle_eligibilite = "convergence de 2 familles de conséquence opérationnelle/client"
    elif nb_consequences == 1 and (a_actif or h_actif):
        tier = "priorite_secondaire"
        regle_eligibilite = "une conséquence opérationnelle/client fortement étayée, avec demande ou persistance clairement étayée"
    elif a_actif and (b_actif or c_actif or d_actif or e_actif or f_actif):
        tier = "a_surveiller"
        regle_eligibilite = "volume au-dessus du niveau habituel avec un signal de conséquence encore isolé ou peu étayé"
    elif h_actif:
        tier = "a_surveiller"
        regle_eligibilite = "récurrence observée sur plusieurs périodes, sans conséquence opérationnelle/client démontrée"

    eligible = tier is not None

    if tier == "priorite_principale":
        niveau_priorite = "Priorité principale"
    elif tier == "priorite_secondaire":
        niveau_priorite = "Priorité secondaire"
    elif tier == "a_surveiller":
        niveau_priorite = "À surveiller"
    else:
        niveau_priorite = None

    elements_contributifs = []
    if a_actif:
        elements_contributifs.append("volume/demande au-dessus du niveau habituel")
    if b_actif:
        elements_contributifs.append("satisfaction nettement sous la référence Livraison")
    if c_actif:
        elements_contributifs.append("effort de traitement nettement au-dessus de la référence")
    if d_actif:
        elements_contributifs.append("nombre de relances nettement au-dessus de la référence")
    if e_actif:
        elements_contributifs.append("part d'issues défavorables nettement au-dessus de la référence")
    if h_actif:
        elements_contributifs.append("récurrence observée sur plusieurs périodes")

    niveau_historique_moyen_pct = None
    moyenne_hist = _moyenne_liste(niveaux_historiques)
    if moyenne_hist is not None:
        niveau_historique_moyen_pct = moyenne_hist * 100

    part_univers_pct = None
    if part_candidat is not None:
        part_univers_pct = part_candidat * 100

    ecart_pct = None
    if ecart_temporel is not None:
        ecart_pct = ecart_temporel * 100

    return {
        "sujet": sujet,
        "eligible": eligible,
        "tier": tier,
        "regle_eligibilite": regle_eligibilite,
        "niveau_priorite": niveau_priorite,
        "familles_actives": familles_actives,
        "observation_principale": _phrase_elements(elements_contributifs),
        "elements_contributifs": elements_contributifs,
        "action_investigation": texte_action_investigation_livraison(familles_actives, concentration_info),
        "volume": {"n": n, "part_univers_pct": part_univers_pct, "univers": total_univers},
        "reference": {
            "niveau_historique_moyen_pct": niveau_historique_moyen_pct,
            "nb_observations": nb_observations,
            "ecart_pct": ecart_pct,
            "confiance_historique": confiance,
        },
        "experience": {
            "csat": csat_candidat, "n_csat": n_csat_candidat,
            "csat_reference": csat_reference, "n_csat_reference": n_csat_reference,
            "lecture": lecture_csat,
        },
        "effort": {
            "replies_moyen": replies_candidat, "replies_reference": replies_reference,
            "lecture_replies": lecture_replies,
            "resolution_h_moyenne": resolution_candidat, "resolution_h_reference": resolution_reference,
            "lecture_resolution": lecture_resolution,
            "reopens_moyen": reopens_candidat, "reopens_reference": reopens_reference,
            "lecture_reopens": lecture_reopens,
        },
        "relances": {
            "moyen": relances_candidat, "reference": relances_reference, "lecture": lecture_relances,
        },
        "issues": {
            "distribution": distribution_issues_livraison(tickets_sujet),
            "part_defavorable_pct": part_defavorable_candidat * 100 if part_defavorable_candidat is not None else None,
            "part_defavorable_reference_pct": (
                part_defavorable_reference * 100 if part_defavorable_reference is not None else None
            ),
            "lecture": lecture_issue,
        },
        "cout": None,  # jamais calculable actuellement -- aucun order_id sur les tickets Livraison
        "temporalite": texte_temporalite,
        "concentration_transporteur": concentration_info,
        "prudence": "Association observée sur les données disponibles, pas une cause démontrée.",
        "score_interne": None,  # usage interne uniquement (classement) -- jamais affiché en UI
        "_ecart_temporel": ecart_temporel,
        "_ecart_csat": ecart_csat,
        "_ecart_effort_max": ecart_effort_max,
        "_ecart_relances": ecart_relances,
        "_ecart_issue_defavorable": ecart_issue_defavorable,
        "_confiance_historique": confiance,
    }


def calculer_score_interne_livraison(signal):
    composante_csat = 0.0
    if signal["_ecart_csat"] is not None and signal["_ecart_csat"] > 0:
        composante_csat = _normaliser(signal["_ecart_csat"], PLAFOND_SCORE_CSAT)

    composante_effort = _normaliser(signal["_ecart_effort_max"], PLAFOND_SCORE_EFFORT)

    composante_relances = 0.0
    if signal["_ecart_relances"] is not None and signal["_ecart_relances"] > 0:
        composante_relances = _normaliser(signal["_ecart_relances"], PLAFOND_SCORE_LIVRAISON_ECART_RELATIF)

    composante_issue = 0.0
    if signal["_ecart_issue_defavorable"] is not None and signal["_ecart_issue_defavorable"] > 0:
        composante_issue = _normaliser(signal["_ecart_issue_defavorable"], PLAFOND_SCORE_LIVRAISON_ECART_RELATIF)

    composante_temporalite = 0.0
    if signal["_ecart_temporel"] is not None and signal["_ecart_temporel"] > 0:
        composante_temporalite = (
            _normaliser(signal["_ecart_temporel"], PLAFOND_SCORE_TEMPORALITE) * signal["_confiance_historique"]
        )

    composante_volume = _normaliser(signal["volume"]["n"], PLAFOND_SCORE_VOLUME_ABSOLU)

    return (
        composante_csat * POIDS_SCORE_LIVRAISON_CSAT
        + composante_effort * POIDS_SCORE_LIVRAISON_EFFORT
        + composante_relances * POIDS_SCORE_LIVRAISON_RELANCES
        + composante_issue * POIDS_SCORE_LIVRAISON_ISSUE
        + composante_temporalite * POIDS_SCORE_LIVRAISON_TEMPORALITE
        + composante_volume * POIDS_SCORE_LIVRAISON_VOLUME
    )


def obtenir_score_interne_livraison(signal):
    return signal["score_interne"]


# Lecture d'activité (Étape 4C, section 15) : jamais une alerte -- décrit le poids de Livraison
# sur la période et le contexte métier qui le explique éventuellement (Black Friday, Noël...),
# sans jamais affirmer de causalité. Peut coexister avec "aucun problème logistique prioritaire".
def construire_lecture_activite_livraison(tickets_livraison_periode, tickets_totaux_periode, evenements_contexte):
    volume = len(tickets_livraison_periode)
    total = len(tickets_totaux_periode)
    part_pct = None
    if total > 0:
        part_pct = volume / total * 100

    if part_pct is None:
        observation = str(volume) + " ticket(s) Livraison sur cette période."
    else:
        observation = (
            "Livraison représente " + str(round(part_pct)) + " % des contacts sur cette période ("
            + str(volume) + " ticket(s))."
        )

    contexte_texte = texte_contexte(evenements_contexte)

    return {
        "observation": observation,
        "volume": volume,
        "part_pct": part_pct,
        "contexte": contexte_texte,
    }


# Orchestration voie A Livraison : un signal par motif (grain unique subject_cluster), triés par
# score interne, plafonnés à l'affichage -- jamais de fusion artificielle entre motifs distincts
# (Étape 4C, section 17 : un mauvais CSAT ou beaucoup de relances communs à deux motifs ne les
# rend pas identiques).
def moteur_livraison_voie_a(tickets_livraison_periode, historique_livraison_par_fichier, nombre_max_signaux):
    sujets_presents = set()
    for ticket in tickets_livraison_periode:
        sujets_presents.add(ticket["subject_cluster"])

    signaux = []
    for sujet in sujets_presents:
        tickets_sujet = []
        for ticket in tickets_livraison_periode:
            if ticket["subject_cluster"] == sujet:
                tickets_sujet.append(ticket)

        niveaux_historiques = construire_niveaux_historiques_livraison(historique_livraison_par_fichier, sujet)
        signal = construire_signal_sujet_livraison_voie_a(
            sujet, tickets_sujet, tickets_livraison_periode, niveaux_historiques
        )
        if signal is not None:
            signaux.append(signal)

    for signal in signaux:
        signal["score_interne"] = calculer_score_interne_livraison(signal)

    signaux_prioritaires_bruts = []
    signaux_a_surveiller_bruts = []
    for signal in signaux:
        if signal["tier"] in ("priorite_principale", "priorite_secondaire"):
            signaux_prioritaires_bruts.append(signal)
        elif signal["tier"] == "a_surveiller":
            signaux_a_surveiller_bruts.append(signal)

    signaux_prioritaires_tries = sorted(signaux_prioritaires_bruts, key=obtenir_score_interne_livraison, reverse=True)
    signaux_a_surveiller_tries = sorted(signaux_a_surveiller_bruts, key=obtenir_score_interne_livraison, reverse=True)

    prioritaires = []
    for i in range(min(nombre_max_signaux, len(signaux_prioritaires_tries))):
        prioritaires.append(signaux_prioritaires_tries[i])

    a_surveiller = []
    for i in range(min(nombre_max_signaux, len(signaux_a_surveiller_tries))):
        a_surveiller.append(signaux_a_surveiller_tries[i])

    sujets_silencieux = []
    for sujet in sujets_presents:
        deja_liste = False
        for signal in prioritaires + a_surveiller:
            if signal["sujet"] == sujet:
                deja_liste = True
        if not deja_liste:
            sujets_silencieux.append(sujet)

    return {
        "prioritaires": prioritaires,
        "a_surveiller": a_surveiller,
        "nb_prioritaires_avant_plafond": len(signaux_prioritaires_tries),
        "nb_a_surveiller_avant_plafond": len(signaux_a_surveiller_tries),
        "sujets_silencieux": sorted(sujets_silencieux),
    }


# ---------------------------------------------------------------------------
# Moteur Avant-vente & conversion (Étape 4D)
#
# Objectif : pas "combien de tickets Avant-vente" ni "le taux de conversion est Y %", mais quels
# types de contact semblent associés à des parcours d'achat différents, et où se situent les
# vraies pistes d'accompagnement. Vocabulaire imposé : "achat OBSERVÉ" dans la fenêtre, jamais
# "conversion causée" -- association observée, jamais une cause démontrée (Étape 4D, section 4).
# ---------------------------------------------------------------------------

FENETRE_CONVERSION_JOURS = 30

SEUIL_MINIMUM_EVALUATION_AVANT_VENTE = SEUIL_MINIMUM_EVALUATION_PRODUIT  # même seuil, cohérence transversale

TYPE_CONTACT_SPONTANE = "Contact spontané"
TYPE_CONTACT_RDV = "RDV conseil"

RDV_STATUT_HONORE = "Honoré"
RDV_STATUT_ANNULE = "Annulé"
RDV_STATUT_NO_SHOW = "No-show"

PARCOURS_RDV_HONORE = "rdv_honore"
PARCOURS_RDV_NON_HONORE = "rdv_non_honore"
PARCOURS_SPONTANE = "spontane"

# Motif structurellement identique au parcours RDV lui-même (une demande de RDV EST le ticket) --
# exclu de l'analyse par motif pour ne pas raconter deux fois la même histoire sous deux angles
# différents (section 11 : le motif sert à trouver des histoires distinctes du parcours RDV).
SUJET_DEMANDE_RDV = "Demande de rendez-vous conseil téléphonique"


# Verrou d'interprétation RDV (Étape 4D, section 3) : "Annulé" et "No-show" ne sont JAMAIS un
# conseil reçu -- seul "Honoré" l'est. Un ticket hors parcours RDV est un contact spontané.
def determiner_parcours_avant_vente(ticket):
    if ticket["type_contact_avant_vente"] == TYPE_CONTACT_RDV:
        if ticket["rdv_statut"] == RDV_STATUT_HONORE:
            return PARCOURS_RDV_HONORE
        return PARCOURS_RDV_NON_HONORE
    return PARCOURS_SPONTANE


# Attribution commande -> contact (Étape 4D.1, section 1-2) : raisonnée PAR COMMANDE, pas par
# ticket. Pour chaque commande admissible d'un client, le contact Avant-vente retenu est le PLUS
# RÉCENT parmi ceux qui la précèdent dans la fenêtre (contact.created_at < commande.order_date <=
# contact.created_at + fenetre_jours) -- jamais le plus ancien. "Dernier contact Avant-vente
# observé avant l'achat", jamais "le contact a causé l'achat". Une commande ne peut avoir qu'UN
# SEUL contact associé (par construction : on choisit LE plus récent). Un même contact PEUT en
# revanche se voir attribuer plusieurs commandes si aucun autre contact ne s'intercale entre elles
# -- pour "achat observé", le contact reste binaire (au moins une commande = OUI) ; pour
# panier/délai, seule la PREMIÈRE commande chronologique qui lui est attribuée est utilisée
# (section 2). Retourne une liste de (ticket, commande_retenue_ou_None, plusieurs_commandes_bool).
def _obtenir_order_date_commande(commande):
    return commande["order_date"]


def resoudre_achats_observes_avant_vente(tickets, index_commandes_email, fenetre_jours):
    tickets_par_email = {}
    for ticket in tickets:
        email = ticket["requester_email"]
        if email in tickets_par_email:
            tickets_par_email[email].append(ticket)
        else:
            tickets_par_email[email] = [ticket]

    commandes_attribuees_par_ticket_id = {}
    for email, tickets_client in tickets_par_email.items():
        commandes_client = index_commandes_email.get(email, [])
        for commande in commandes_client:
            candidats = []
            for ticket in tickets_client:
                date_debut = ticket["created_at"]
                date_fin = date_debut + datetime.timedelta(days=fenetre_jours)
                if date_debut < commande["order_date"] <= date_fin:
                    candidats.append(ticket)

            if len(candidats) == 0:
                continue

            contact_le_plus_recent = candidats[0]
            for candidat in candidats:
                if candidat["created_at"] > contact_le_plus_recent["created_at"]:
                    contact_le_plus_recent = candidat

            cle_ticket = contact_le_plus_recent["ticket_id"]
            if cle_ticket in commandes_attribuees_par_ticket_id:
                commandes_attribuees_par_ticket_id[cle_ticket].append(commande)
            else:
                commandes_attribuees_par_ticket_id[cle_ticket] = [commande]

    resultats = []
    for ticket in tickets:
        commandes_attribuees = commandes_attribuees_par_ticket_id.get(ticket["ticket_id"])
        if commandes_attribuees is None:
            resultats.append((ticket, None, False))
            continue
        commandes_triees = sorted(commandes_attribuees, key=_obtenir_order_date_commande)
        premiere_commande_attribuee = commandes_triees[0]
        resultats.append((ticket, premiere_commande_attribuee, len(commandes_attribuees) > 1))

    return resultats


# Statistiques d'un groupe de contacts (parcours ou motif) à partir des résultats déjà
# dédupliqués -- "part des contacts suivis d'un achat observé", jamais "taux de conversion
# client" (le dénominateur est le contact, pas le client -- Étape 4D, section 22).
def calculer_stats_achat_observe(resultats_groupe):
    n_contacts = len(resultats_groupe)
    montants = []
    delais = []
    n_achats = 0
    for ticket, commande, plusieurs_commandes in resultats_groupe:
        if commande is not None:
            n_achats = n_achats + 1
            montants.append(commande["montant_total"])
            delais.append((commande["order_date"] - ticket["created_at"]).days)

    taux_pct = None
    if n_contacts > 0:
        taux_pct = n_achats / n_contacts * 100

    panier_moyen = None
    panier_median = None
    if len(montants) > 0:
        panier_moyen = sum(montants) / len(montants)
        panier_median = mediane(montants)

    delai_moyen = None
    delai_median = None
    if len(delais) > 0:
        delai_moyen = sum(delais) / len(delais)
        delai_median = mediane(delais)

    return {
        "n_contacts": n_contacts,
        "n_achats": n_achats,
        "taux_pct": taux_pct,
        "panier_moyen": panier_moyen,
        "panier_median": panier_median,
        "n_panier": len(montants),
        "delai_moyen": delai_moyen,
        "delai_median": delai_median,
        "n_delai": len(delais),
    }


# Comparaison prudente entre parcours (Étape 4D, section 9) : ne fabrique jamais un avantage --
# ne cite un écart que s'il franchit le même seuil relatif "marqué" déjà utilisé partout ailleurs
# (SEUIL_EFFORT_ECART_MARQUE_REL), sinon conclut génériquement à des taux proches. N'affirme
# jamais de causalité, jamais de classement "meilleur parcours".
def texte_conclusion_parcours_rdv(stats_honore, stats_non_honore, stats_spontane):
    groupes = []
    if stats_honore["taux_pct"] is not None and stats_honore["n_contacts"] >= SEUIL_MINIMUM_EVALUATION_AVANT_VENTE:
        groupes.append(("RDV honoré", stats_honore["taux_pct"]))
    if (
        stats_non_honore["taux_pct"] is not None
        and stats_non_honore["n_contacts"] >= SEUIL_MINIMUM_EVALUATION_AVANT_VENTE
    ):
        groupes.append(("RDV annulé/no-show", stats_non_honore["taux_pct"]))
    if stats_spontane["taux_pct"] is not None and stats_spontane["n_contacts"] >= SEUIL_MINIMUM_EVALUATION_AVANT_VENTE:
        groupes.append(("contact spontané", stats_spontane["taux_pct"]))

    if len(groupes) < 2:
        return (
            "Échantillon insuffisant sur au moins un parcours pour comparer les taux d'achat observé "
            "cette période."
        )

    taux_valeurs = []
    for nom, taux in groupes:
        taux_valeurs.append(taux)
    ecart_max = max(taux_valeurs) - min(taux_valeurs)
    reference_moyenne = sum(taux_valeurs) / len(taux_valeurs)

    if reference_moyenne == 0 or ecart_max / reference_moyenne < SEUIL_EFFORT_ECART_MARQUE_REL:
        return (
            "Les parcours observés présentent des taux d'achat observé proches sur cette période. Les données "
            "disponibles ne permettent pas d'identifier un avantage net associé à l'un des parcours."
        )

    plus_haut = max(groupes, key=_obtenir_deuxieme_element)
    plus_bas = min(groupes, key=_obtenir_deuxieme_element)
    return (
        "Le parcours " + plus_haut[0] + " montre un taux d'achat observé plus élevé que le parcours "
        + plus_bas[0] + " sur cette période (association observée, pas une cause démontrée -- les parcours "
        "peuvent différer sur d'autres plans, comme le niveau d'intention initial)."
    )


def _obtenir_deuxieme_element(couple):
    return couple[1]


# "Contacts Avant-vente avant achat" (Étape 4D.1, section 4) : pour CHAQUE commande créditée,
# compte uniquement les contacts du même client dont created_at est STRICTEMENT antérieur à la
# date de la commande, dans la même fenêtre de fenetre_jours qui a servi à l'attribution -- jamais
# un total de tickets sur toute la période (l'ancienne métrique 4D était mal nommée : elle
# comptait le total de contacts du client, contacts POST-achat inclus). Une commande sans aucun
# contact éligible dans cette fenêtre n'est de toute façon jamais créditée par
# resoudre_achats_observes_avant_vente, donc chaque achat compté ici a au moins 1 contact.
# Retourne None si aucun achat n'a été crédité (rien à calculer).
def analyser_contacts_avant_achat(resultats, fenetre_jours):
    tickets_par_client = {}
    for ticket, commande, plusieurs_commandes in resultats:
        email = ticket["requester_email"]
        if email in tickets_par_client:
            tickets_par_client[email].append(ticket)
        else:
            tickets_par_client[email] = [ticket]

    comptes_par_achat = []
    for ticket, commande, plusieurs_commandes in resultats:
        if commande is None:
            continue
        email = ticket["requester_email"]
        date_limite_inf = commande["order_date"] - datetime.timedelta(days=fenetre_jours)
        n_contacts_avant = 0
        for autre_ticket in tickets_par_client[email]:
            if date_limite_inf <= autre_ticket["created_at"] < commande["order_date"]:
                n_contacts_avant = n_contacts_avant + 1
        comptes_par_achat.append(n_contacts_avant)

    if len(comptes_par_achat) == 0:
        return None

    n_plusieurs = 0
    distribution = {"1": 0, "2": 0, "3+": 0}
    for n in comptes_par_achat:
        if n > 1:
            n_plusieurs = n_plusieurs + 1
        if n <= 1:
            distribution["1"] = distribution["1"] + 1
        elif n == 2:
            distribution["2"] = distribution["2"] + 1
        else:
            distribution["3+"] = distribution["3+"] + 1

    return {
        "n_achats_credites": len(comptes_par_achat),
        "part_plusieurs_contacts_pct": n_plusieurs / len(comptes_par_achat) * 100,
        "nombre_moyen_contacts_avant_achat": sum(comptes_par_achat) / len(comptes_par_achat),
        "nombre_median_contacts_avant_achat": mediane(comptes_par_achat),
        "distribution_nb_contacts": distribution,
    }


# Analyse complète du parcours RDV : demandes/honorés/annulés/no-show, puis achat observé par
# parcours (spontané inclus pour comparaison), puis conclusion prudente générique.
def analyser_parcours_rdv(tickets_avant_vente, resultats_dedupliques, fenetre_jours=FENETRE_CONVERSION_JOURS):
    n_demandes = 0
    n_honore = 0
    n_annule = 0
    n_no_show = 0
    for ticket in tickets_avant_vente:
        if ticket["type_contact_avant_vente"] == TYPE_CONTACT_RDV:
            n_demandes = n_demandes + 1
            if ticket["rdv_statut"] == RDV_STATUT_HONORE:
                n_honore = n_honore + 1
            elif ticket["rdv_statut"] == RDV_STATUT_ANNULE:
                n_annule = n_annule + 1
            elif ticket["rdv_statut"] == RDV_STATUT_NO_SHOW:
                n_no_show = n_no_show + 1

    groupes = {PARCOURS_RDV_HONORE: [], PARCOURS_RDV_NON_HONORE: [], PARCOURS_SPONTANE: []}
    for resultat in resultats_dedupliques:
        ticket = resultat[0]
        parcours = determiner_parcours_avant_vente(ticket)
        groupes[parcours].append(resultat)

    stats_honore = calculer_stats_achat_observe(groupes[PARCOURS_RDV_HONORE])
    stats_non_honore = calculer_stats_achat_observe(groupes[PARCOURS_RDV_NON_HONORE])
    stats_spontane = calculer_stats_achat_observe(groupes[PARCOURS_SPONTANE])

    conclusion = texte_conclusion_parcours_rdv(stats_honore, stats_non_honore, stats_spontane)

    return {
        "rdv_demandes": n_demandes,
        "rdv_honore": n_honore,
        "rdv_annule": n_annule,
        "rdv_no_show": n_no_show,
        "stats_rdv_honore": stats_honore,
        "stats_rdv_non_honore": stats_non_honore,
        "stats_spontane": stats_spontane,
        "conclusion": conclusion,
        "contacts_avant_achat": analyser_contacts_avant_achat(resultats_dedupliques, fenetre_jours),
    }


# Lecture d'activité (jamais une alerte) -- même principe que Livraison (Étape 4C, section 15).
def construire_lecture_activite_avant_vente(tickets_avant_vente, tickets_totaux_periode, evenements_contexte):
    volume = len(tickets_avant_vente)
    total = len(tickets_totaux_periode)
    part_pct = None
    if total > 0:
        part_pct = volume / total * 100

    if part_pct is None:
        observation = str(volume) + " ticket(s) Avant-vente sur cette période."
    else:
        observation = (
            "Avant-vente représente " + str(round(part_pct)) + " % des contacts sur cette période ("
            + str(volume) + " ticket(s))."
        )

    return {
        "observation": observation,
        "volume": volume,
        "part_pct": part_pct,
        "contexte": texte_contexte(evenements_contexte),
    }


def _obtenir_n_distribution_canal(item):
    return item["n"]


# Canal = dimension purement DESCRIPTIVE (Étape 4D, section 7C et 19) : les canaux sont corrélés
# aux parcours (un RDV est par construction au téléphone dans ce jeu de données), donc aucune
# comparaison de taux d'achat par canal n'est calculée ici -- seulement des comptages/parts, pour
# éviter toute conclusion du type "le téléphone convertit mieux" qui ignorerait la composition des
# parcours.
def distribution_canal_avant_vente(tickets):
    total = len(tickets)
    if total == 0:
        return []

    comptes = {}
    for ticket in tickets:
        canal = ticket["via_channel"]
        if canal in comptes:
            comptes[canal] = comptes[canal] + 1
        else:
            comptes[canal] = 1

    distribution = []
    for canal, n in comptes.items():
        distribution.append({"canal": canal, "n": n, "part_pct": n / total * 100})

    return sorted(distribution, key=_obtenir_n_distribution_canal, reverse=True)


def controler_qualite_donnees_avant_vente(tickets_avant_vente, tickets_hors_avant_vente):
    anomalies = []

    n_sans_type_contact = 0
    n_type_contact_invalide = 0
    n_rdv_sans_statut = 0
    n_rdv_statut_invalide = 0
    n_rdv_non_telephone = 0
    for ticket in tickets_avant_vente:
        type_contact = ticket["type_contact_avant_vente"]
        if type_contact is None:
            n_sans_type_contact = n_sans_type_contact + 1
        elif type_contact not in (TYPE_CONTACT_SPONTANE, TYPE_CONTACT_RDV):
            n_type_contact_invalide = n_type_contact_invalide + 1

        if type_contact == TYPE_CONTACT_RDV:
            statut = ticket["rdv_statut"]
            if statut is None:
                n_rdv_sans_statut = n_rdv_sans_statut + 1
            elif statut not in (RDV_STATUT_HONORE, RDV_STATUT_ANNULE, RDV_STATUT_NO_SHOW):
                n_rdv_statut_invalide = n_rdv_statut_invalide + 1
            if ticket["via_channel"] != "Téléphone":
                n_rdv_non_telephone = n_rdv_non_telephone + 1

    n_type_contact_hors_perimetre = 0
    n_rdv_statut_hors_perimetre = 0
    for ticket in tickets_hors_avant_vente:
        if ticket["type_contact_avant_vente"] is not None:
            n_type_contact_hors_perimetre = n_type_contact_hors_perimetre + 1
        if ticket["rdv_statut"] is not None:
            n_rdv_statut_hors_perimetre = n_rdv_statut_hors_perimetre + 1

    if n_sans_type_contact > 0:
        anomalies.append(str(n_sans_type_contact) + " ticket(s) Avant-vente sans type_contact_avant_vente renseigné.")
    if n_type_contact_invalide > 0:
        anomalies.append(str(n_type_contact_invalide) + " ticket(s) Avant-vente avec un type_contact_avant_vente hors taxonomie.")
    if n_rdv_sans_statut > 0:
        anomalies.append(str(n_rdv_sans_statut) + " RDV conseil sans rdv_statut renseigné.")
    if n_rdv_statut_invalide > 0:
        anomalies.append(str(n_rdv_statut_invalide) + " RDV conseil avec un rdv_statut hors taxonomie.")
    if n_rdv_non_telephone > 0:
        anomalies.append(str(n_rdv_non_telephone) + " RDV conseil hors canal téléphone.")
    if n_type_contact_hors_perimetre > 0:
        anomalies.append(
            str(n_type_contact_hors_perimetre) + " ticket(s) hors Avant-vente avec type_contact_avant_vente renseigné."
        )
    if n_rdv_statut_hors_perimetre > 0:
        anomalies.append(str(n_rdv_statut_hors_perimetre) + " ticket(s) hors Avant-vente avec rdv_statut renseigné.")

    return anomalies


# Comparaison de taux d'achat observé, jamais un jugement de "bonne" ou "mauvaise" performance --
# symétrique de lire_ecart_effort mais dans l'autre sens (ici, c'est un taux plus BAS que la
# référence qui constitue le signal, pas un taux plus haut). Réutilise le même seuil relatif
# SEUIL_EFFORT_ECART_MARQUE_REL déjà validé ailleurs -- aucun nouveau seuil inventé.
def lire_ecart_taux_achat_observe(taux_candidat, taux_reference):
    if taux_candidat is None or taux_reference is None or taux_reference == 0:
        return VOCABULAIRE_ECART_INDISPONIBLE

    ecart_relatif = (taux_reference - taux_candidat) / taux_reference  # positif si le candidat achète MOINS
    if abs(ecart_relatif) <= 0.05:
        return VOCABULAIRE_ECART_PROCHE
    if ecart_relatif >= SEUIL_EFFORT_ECART_MARQUE_REL:
        return VOCABULAIRE_ECART_MARQUE
    if ecart_relatif > 0:
        return VOCABULAIRE_ECART_BAS
    return VOCABULAIRE_ECART_HAUT


# Niveau historique d'un motif Avant-vente = sa part du volume Avant-vente total, observation par
# observation antérieure -- même principe que Livraison/Produit, grain unique subject_cluster.
def construire_niveaux_historiques_avant_vente(historique_avant_vente_par_fichier, sujet):
    niveaux = []
    for tickets_fichier in historique_avant_vente_par_fichier:
        if len(tickets_fichier) == 0:
            continue
        tickets_sujet_fichier = []
        for ticket in tickets_fichier:
            if ticket["subject_cluster"] == sujet:
                tickets_sujet_fichier.append(ticket)
        part = calculer_part(len(tickets_sujet_fichier), len(tickets_fichier))
        if part is not None:
            niveaux.append(part)
    return niveaux


# Signal multicritère pour UN motif Avant-vente (grain unique : subject_cluster, la demande de RDV
# elle-même exclue -- voir SUJET_DEMANDE_RDV). Familles de preuve (Étape 4D, section 11) :
#   A Demande     -- volume/part au-dessus du niveau habituel, ou hausse récente vs historique
#   C Achat       -- part de contacts avec achat observé NETTEMENT sous la référence Avant-vente
#   F Expérience  -- CSAT du motif vs reste d'Avant-vente (contexte, jamais suffisante seule)
#   H Persistance -- part du volume Avant-vente inhabituelle sur plusieurs observations antérieures
# B (intention RDV) et G (contexte) restent des dimensions propres au parcours/à la lecture
# d'activité, pas des familles de ce candidat par motif. Aucun "score de conversion" : l'éligibilité
# est une convergence de familles, jamais un chiffre seul (Étape 4D, section 13) --
#   OPPORTUNITÉ = demande notable ET achat observé nettement sous la référence ;
#   À SURVEILLER = achat sous la référence mais volume insuffisant pour qualifier d'opportunité,
#                  OU demande en forte hausse récente sans assez de recul sur l'achat.
# Un motif à fort volume mais achat proche de la référence n'est ni l'un ni l'autre : c'est de
# l'ACTIVITÉ (section 15), pas un problème à signaler.
def construire_signal_motif_avant_vente(sujet, tickets_sujet, resultats_sujet, tickets_avant_vente_periode,
                                         resultats_avant_vente_periode, niveaux_historiques, evenements_contexte):
    n = len(tickets_sujet)
    if n < SEUIL_MINIMUM_EVALUATION_AVANT_VENTE:
        return None

    total_univers = len(tickets_avant_vente_periode)
    part_candidat = calculer_part(n, total_univers)
    ecart_temporel = ecart_relatif_temporel(niveaux_historiques, part_candidat)
    texte_temporalite = evaluer_temporalite(niveaux_historiques, part_candidat)
    confiance = confiance_historique(len(niveaux_historiques))
    _persistance_brute, part_elevees, nb_observations = evaluer_persistance_temporelle(
        niveaux_historiques, part_candidat
    )

    temporel_compte = (
        ecart_temporel is not None
        and ecart_temporel >= SEUIL_ECART_RELATIF_TEMPORALITE
        and confiance >= SEUIL_CONFIANCE_TEMPORELLE_MINIMALE
    )

    stats_candidat = calculer_stats_achat_observe(resultats_sujet)
    stats_reference = calculer_stats_achat_observe(resultats_avant_vente_periode)
    lecture_achat = lire_ecart_taux_achat_observe(stats_candidat["taux_pct"], stats_reference["taux_pct"])

    csat_candidat = moyenne(tickets_sujet, "csat")
    n_csat_candidat = _compte_valeurs_non_nulles(tickets_sujet, "csat")
    csat_reference = moyenne(tickets_avant_vente_periode, "csat")
    n_csat_reference = _compte_valeurs_non_nulles(tickets_avant_vente_periode, "csat")

    a_actif = (
        n >= SEUIL_VOLUME_ABSOLU_NOTABLE
        or (part_candidat is not None and part_candidat >= SEUIL_VOLUME_PART_NOTABLE)
        or temporel_compte
    )
    c_actif = lecture_achat == VOCABULAIRE_ECART_MARQUE
    h_actif = (
        temporel_compte
        and part_elevees >= SEUIL_PART_ELEVEES_PERSISTANCE
        and nb_observations >= NB_OBSERVATIONS_MIN_PERSISTANCE
    )

    familles_actives = []
    if a_actif:
        familles_actives.append("A")
    if c_actif:
        familles_actives.append("C")
    if h_actif:
        familles_actives.append("H")

    if a_actif and c_actif:
        tier = "opportunite"
        regle_eligibilite = "volume notable et part de contacts avec achat observé nettement sous la référence Avant-vente"
    elif c_actif:
        tier = "a_surveiller"
        regle_eligibilite = "achat observé sous la référence, mais volume insuffisant pour qualifier d'opportunité"
    elif a_actif and temporel_compte:
        tier = "a_surveiller"
        regle_eligibilite = "forte hausse récente du motif, pas encore assez de recul sur l'achat observé"
    else:
        tier = None
        regle_eligibilite = None

    eligible = tier is not None
    if tier == "opportunite":
        niveau = "Opportunité à investiguer"
    elif tier == "a_surveiller":
        niveau = "À surveiller"
    else:
        niveau = None

    elements_contributifs = []
    if a_actif:
        elements_contributifs.append("volume/demande au-dessus du niveau habituel")
    if c_actif:
        elements_contributifs.append("part de contacts avec achat observé nettement sous la référence Avant-vente")
    if h_actif:
        elements_contributifs.append("récurrence observée sur plusieurs périodes")

    contexte_texte = texte_contexte(evenements_contexte)

    return {
        "sujet": sujet,
        "eligible": eligible,
        "tier": tier,
        "regle_eligibilite": regle_eligibilite,
        "niveau": niveau,
        "familles_actives": familles_actives,
        "observation_principale": _phrase_elements(elements_contributifs),
        "piste_investigation": (
            "Vérifier si l'information produit répond suffisamment à cette demande avant contact, ou si "
            "elle correspond à une attente produit non couverte."
        ),
        "volume": {
            "n": n, "part_univers_pct": part_candidat * 100 if part_candidat is not None else None,
            "univers": total_univers,
        },
        "temporalite": texte_temporalite,
        "achat_observe": stats_candidat,
        "achat_observe_reference_pct": stats_reference["taux_pct"],
        "lecture_achat": lecture_achat,
        "experience": {
            "csat": csat_candidat, "n_csat": n_csat_candidat,
            "csat_reference": csat_reference, "n_csat_reference": n_csat_reference,
        },
        "contexte": contexte_texte,
        "prudence": "Association observée sur les données disponibles, pas une cause démontrée.",
    }


# Orchestration : un signal par motif (hors demande de RDV), pas de plafond artificiel imposé --
# 0 opportunité est une sortie valide (Étape 4D, section 14).
def moteur_avant_vente_motifs(tickets_avant_vente_periode, resultats_avant_vente_periode,
                               historique_avant_vente_par_fichier, evenements_contexte, nombre_max_signaux):
    resultats_par_ticket_id = {}
    for ticket, commande, plusieurs_commandes in resultats_avant_vente_periode:
        resultats_par_ticket_id[ticket["ticket_id"]] = (ticket, commande, plusieurs_commandes)

    sujets_presents = set()
    for ticket in tickets_avant_vente_periode:
        if ticket["subject_cluster"] != SUJET_DEMANDE_RDV:
            sujets_presents.add(ticket["subject_cluster"])

    signaux = []
    for sujet in sujets_presents:
        tickets_sujet = []
        resultats_sujet = []
        for ticket in tickets_avant_vente_periode:
            if ticket["subject_cluster"] == sujet:
                tickets_sujet.append(ticket)
                resultats_sujet.append(resultats_par_ticket_id[ticket["ticket_id"]])

        niveaux_historiques = construire_niveaux_historiques_avant_vente(historique_avant_vente_par_fichier, sujet)
        signal = construire_signal_motif_avant_vente(
            sujet, tickets_sujet, resultats_sujet, tickets_avant_vente_periode,
            resultats_avant_vente_periode, niveaux_historiques, evenements_contexte,
        )
        if signal is not None:
            signaux.append(signal)

    opportunites_brutes = []
    a_surveiller_bruts = []
    for signal in signaux:
        if signal["tier"] == "opportunite":
            opportunites_brutes.append(signal)
        elif signal["tier"] == "a_surveiller":
            a_surveiller_bruts.append(signal)

    def _obtenir_volume_signal_av(signal):
        return signal["volume"]["n"]

    opportunites_triees = sorted(opportunites_brutes, key=_obtenir_volume_signal_av, reverse=True)
    a_surveiller_triees = sorted(a_surveiller_bruts, key=_obtenir_volume_signal_av, reverse=True)

    opportunites = []
    for i in range(min(nombre_max_signaux, len(opportunites_triees))):
        opportunites.append(opportunites_triees[i])

    a_surveiller = []
    for i in range(min(nombre_max_signaux, len(a_surveiller_triees))):
        a_surveiller.append(a_surveiller_triees[i])

    sujets_silencieux = []
    for sujet in sujets_presents:
        deja_liste = False
        for signal in opportunites + a_surveiller:
            if signal["sujet"] == sujet:
                deja_liste = True
        if not deja_liste:
            sujets_silencieux.append(sujet)

    return {
        "opportunites": opportunites,
        "a_surveiller": a_surveiller,
        "nb_opportunites_avant_plafond": len(opportunites_triees),
        "nb_a_surveiller_avant_plafond": len(a_surveiller_triees),
        "sujets_silencieux": sorted(sujets_silencieux),
    }


# ---------------------------------------------------------------------------
# Moteur Impact & confiance / NPS (Étape 4E)
#
# Objectif : pas "quel est notre NPS" mais que disent les signaux de confiance disponibles, avec
# quelle solidité, et quel lien (association, jamais causalité) avec le Customer Care. NPS = %
# promoteurs (9-10) - % détracteurs (0-6), toujours affiché en entier (-100 à +100), jamais un
# benchmark industrie -- seulement une lecture de composition, de prudence d'échantillon et
# d'alignement/divergence avec les autres signaux Care disponibles.
# ---------------------------------------------------------------------------

FENETRE_NPS_EXPERIENCE_JOURS = 60
SEUIL_PRUDENCE_ECHANTILLON_NPS = 15  # en dessous : "lecture à prendre avec prudence" -- jamais "significatif"
SEUIL_RESOLUTION_RAPIDE_H = 24
# Déplacé depuis app.py (Étape 4E) : seuil CSAT partagé, réutilisé ici pour l'alignement CSAT-NPS
# et pour identifier les cas compatibles service recovery -- même seuil que le reste de l'app,
# jamais un nouveau nombre inventé pour la même notion de "CSAT insatisfaisant".
SEUIL_CSAT_INSATISFAISANT = 4

TYPE_EXPERIENCE_SAV_RECURRENT = "SAV récurrent"
TYPE_EXPERIENCE_SAV = "SAV"
TYPE_EXPERIENCE_LIVRAISON = "Problème livraison"
TYPE_EXPERIENCE_REMPLACEMENT = "Remplacement"
TYPE_EXPERIENCE_RESOLUTION_RAPIDE = "Résolution rapide"
TYPE_EXPERIENCE_RESOLUTION_LONGUE = "Résolution longue"
TYPE_EXPERIENCE_AUTRE = "Autre contact"
# Jamais "Contacted"/"Never contacted" -- le matching dépend de l'email, de la fenêtre et de la
# disponibilité des données : l'absence de correspondance ne prouve pas l'absence de contact
# (Étape 4E, section 12).
TYPE_EXPERIENCE_AUCUN = "Aucun contact Care identifié dans les données disponibles"

TEXTE_PRUDENCE_BIAIS_SELECTION = (
    "Cette association ne permet pas d'isoler l'effet du Customer Care : les clients avec un contact "
    "Care identifié l'ont souvent contacté à la suite d'un problème déjà rencontré (biais de sélection)."
)

# Étape 4E.1 -- la prudence d'échantillon n'est plus un seuil absolu opposant "fiable" à "pas
# fiable" (l'ancien n>=15 -> "étayée" était une fausse équivalence statistique). Elle situe le
# volume de réponses du mois PAR RAPPORT AUX AUTRES OBSERVATIONS DÉJÀ DISPONIBLES (jamais les
# futures) -- un rang, comme rang_relatif ailleurs dans le fichier, jamais un score opaque.
ETAT_PRUDENCE_PREMIERE_OBSERVATION = "premiere_observation"
ETAT_PRUDENCE_VOLUME_FAIBLE = "volume_faible"
ETAT_PRUDENCE_VOLUME_HABITUEL = "volume_habituel"
ETAT_PRUDENCE_VOLUME_ELEVE = "volume_eleve"


# NPS toujours affiché en entier (-100 à +100), jamais de décimale dans l'UI (Étape 4E, section 2)
# -- l'arrondi n'est appliqué qu'à l'affichage, jamais aux calculs internes (calculer_nps garde sa
# précision complète).
def formater_nps_entier(valeur):
    if valeur is None:
        return None
    valeur_arrondie = round(valeur)
    if valeur_arrondie > 0:
        return "+" + str(valeur_arrondie)
    return str(valeur_arrondie)


# Composition explicite (Étape 4E, section 4/23) : reconstruit promoteurs/passifs/détracteurs
# directement depuis les scores 0-10, jamais seulement le score agrégé -- le NPS recalculé ici doit
# toujours être cohérent avec calculer_nps (contrôle qualité, voir controler_qualite_donnees_nps).
def calculer_composition_nps(reponses):
    n = len(reponses)
    if n == 0:
        return None

    n_promoteurs = 0
    n_passifs = 0
    n_detracteurs = 0
    for reponse in reponses:
        score = reponse["score"]
        if score >= 9:
            n_promoteurs = n_promoteurs + 1
        elif score >= 7:
            n_passifs = n_passifs + 1
        else:
            n_detracteurs = n_detracteurs + 1

    return {
        "n": n,
        "n_promoteurs": n_promoteurs, "part_promoteurs_pct": n_promoteurs / n * 100,
        "n_passifs": n_passifs, "part_passifs_pct": n_passifs / n * 100,
        "n_detracteurs": n_detracteurs, "part_detracteurs_pct": n_detracteurs / n * 100,
        "nps": (n_promoteurs / n - n_detracteurs / n) * 100,
    }


# Position du n du mois DANS LA DISTRIBUTION DES n DÉJÀ DISPONIBLES (elle-même incluse), jamais
# un seuil absolu (Étape 4E.1, section 3) -- réutilise rang_relatif tel quel, comme pour le NPS et
# le CSAT dans evaluer_alignement_care_nps. Pas d'historique antérieur (premier mois disponible) :
# rang_relatif renvoie déjà None dans ce cas, jamais une comparaison à des données qui n'existent
# pas encore (no future leakage, section 6).
def evaluer_prudence_echantillon_nps(historique_n, index):
    rang_n = rang_relatif(historique_n[:index + 1], index)
    if rang_n is None:
        return ETAT_PRUDENCE_PREMIERE_OBSERVATION
    if rang_n <= SEUIL_RANG_BAS_STRICT:
        return ETAT_PRUDENCE_VOLUME_FAIBLE
    if rang_n >= SEUIL_RANG_HAUT_STRICT:
        return ETAT_PRUDENCE_VOLUME_ELEVE
    return ETAT_PRUDENCE_VOLUME_HABITUEL


# Vocabulaire jamais statistique ("fiable"/"significatif"/"étayé" pouvant suggérer une validation
# statistique) -- une lecture contextuelle du volume, rien de plus (Étape 4E.1, section 4). La
# prudence est TOUJOURS énoncée avant le chiffre NPS lui-même quand le volume est faible --
# l'appelant (UI) doit afficher ce texte en premier, jamais après une lecture positive/négative.
def texte_prudence_echantillon_nps(etat, n):
    if etat == ETAT_PRUDENCE_PREMIERE_OBSERVATION:
        return str(n) + " réponses -- pas encore assez d'historique pour situer ce volume."
    if etat == ETAT_PRUDENCE_VOLUME_FAIBLE:
        return (
            "Lecture à prendre avec prudence : " + str(n) + " réponses, un volume nettement "
            "inférieur aux observations précédentes."
        )
    if etat == ETAT_PRUDENCE_VOLUME_ELEVE:
        return str(n) + " réponses -- observation mieux documentée que la plupart des périodes précédentes."
    return str(n) + " réponses -- volume de réponses dans la plage habituelle des périodes précédentes."


def _obtenir_cle_mois_reponse(reponse):
    return reponse["date_reponse"].strftime("%Y-%m")


# Historique NPS mensuel, chronologique -- base commune pour l'affichage et pour les baselines
# (jamais de fuite du futur : l'appelant qui évalue le mois M ne doit utiliser que
# historique[:index_de_M + 1], jamais au-delà -- Étape 4E, section 21).
def construire_historique_nps_par_mois(reponses):
    par_mois = {}
    for reponse in reponses:
        cle = _obtenir_cle_mois_reponse(reponse)
        if cle in par_mois:
            par_mois[cle].append(reponse)
        else:
            par_mois[cle] = [reponse]

    historique = []
    for cle in sorted(par_mois.keys()):
        reponses_mois = par_mois[cle]
        composition = calculer_composition_nps(reponses_mois)
        historique.append({
            "cle_mois": cle,
            "reponses": reponses_mois,
            "composition": composition,
            "n": len(reponses_mois),
            "nps": composition["nps"],
        })
    return historique


# Profil Care agrégé pour un mois calendaire (tickets de ce mois, toutes catégories confondues) --
# sert de contrepoint au NPS du même mois pour la lecture d'alignement/divergence.
def construire_profil_care_mensuel(tickets_mois):
    n = len(tickets_mois)
    if n == 0:
        return None

    n_sav = 0
    for ticket in tickets_mois:
        if categoriser(ticket) == CATEGORIE_SAV_PRODUIT:
            n_sav = n_sav + 1

    return {
        "n_tickets": n,
        "csat": moyenne(tickets_mois, "csat"),
        "n_csat": _compte_valeurs_non_nulles(tickets_mois, "csat"),
        "part_sav_pct": n_sav / n * 100,
        "frt_moyen": moyenne(tickets_mois, "first_reply_time_min"),
        "resolution_moyenne": moyenne(tickets_mois, "full_resolution_time_hours"),
        "reopens_moyen": moyenne(tickets_mois, "reopens"),
        "replies_moyen": moyenne(tickets_mois, "replies"),
    }


# Étape 4E.1 -- un rang extrême (rang_relatif) ne suffit pas : dans une série resserrée (ex.
# +7,+9,+7,+6), le minimum de la série est mécaniquement au rang le plus bas alors que l'écart
# réel ne représente que quelques points sur une étendue tout aussi faible -- ça ne mérite pas une
# histoire "NPS bas". AMPLITUDE mesure si l'écart au centre du reste de la série est une part
# suffisante de l'étendue (max-min) déjà observée -- pas un nombre de points NPS fixé d'avance
# (jamais "5 points" ou "10 points" comme seuil magique), mais une proportion, donc automatiquement
# adaptative à l'échelle réelle de la série. Réutilise mediane (déjà existante, gelée) comme centre
# robuste, pas une moyenne sensible à un seul point extrême. Ce n'est pas un test statistique
# (pas de p-value) -- seulement "l'écart est-il assez grand pour mériter une histoire manager".
SEUIL_AMPLITUDE_PART_ETENDUE_NPS = 0.5


def amplitude_relative_etendue(valeurs, index):
    valeur_cible = valeurs[index]
    if valeur_cible is None:
        return None
    valeurs_valides = []
    for valeur in valeurs:
        if valeur is not None:
            valeurs_valides.append(valeur)
    if len(valeurs_valides) <= 1:
        return None
    etendue = max(valeurs_valides) - min(valeurs_valides)
    if etendue == 0:
        return None
    reste = []
    for i in range(len(valeurs)):
        if i != index and valeurs[i] is not None:
            reste.append(valeurs[i])
    centre_reste = mediane(reste)
    if centre_reste is None:
        return None
    return abs(valeur_cible - centre_reste) / etendue


# Alignement/divergence NPS <-> signaux Care (Étape 4E, section 11/19/29 ; amplitude ajoutée en
# 4E.1) : réutilise rang_relatif (position dans la distribution disponible, déjà validé et non
# modifié en Tendances) sur les séries NPS et CSAT mensuelles, jamais une moyenne fragile ni un
# score composite opaque. Bornée à historique[:index_mois_cible+1] des deux côtés -- aucune fuite
# du futur. Retourne None si l'historique est trop court pour situer le mois (jamais forcé). Le
# rang seul ne suffit plus à qualifier un type : l'amplitude du NPS (voir
# amplitude_relative_etendue) doit aussi être suffisante -- une position extrême dans une série
# compacte ne déclenche plus rien.
def evaluer_alignement_care_nps(historique_nps_mensuel, historique_care_mensuel, index_mois_cible):
    if index_mois_cible >= len(historique_nps_mensuel) or index_mois_cible >= len(historique_care_mensuel):
        return None

    nps_valeurs = []
    for item in historique_nps_mensuel[:index_mois_cible + 1]:
        nps_valeurs.append(item["nps"])

    csat_valeurs = []
    reopens_valeurs = []
    resolution_valeurs = []
    for item in historique_care_mensuel[:index_mois_cible + 1]:
        if item is None:
            csat_valeurs.append(None)
            reopens_valeurs.append(None)
            resolution_valeurs.append(None)
        else:
            csat_valeurs.append(item["csat"])
            reopens_valeurs.append(item["reopens_moyen"])
            resolution_valeurs.append(item["resolution_moyenne"])

    if len(nps_valeurs) < 2:
        return None

    rang_nps = rang_relatif(nps_valeurs, index_mois_cible)
    rang_csat = rang_relatif(csat_valeurs, index_mois_cible)
    rang_reopens = rang_relatif(reopens_valeurs, index_mois_cible)
    rang_resolution = rang_relatif(resolution_valeurs, index_mois_cible)

    if rang_nps is None:
        return None

    amplitude_nps = amplitude_relative_etendue(nps_valeurs, index_mois_cible)
    amplitude_suffisante = amplitude_nps is not None and amplitude_nps >= SEUIL_AMPLITUDE_PART_ETENDUE_NPS

    nps_bas = rang_nps <= SEUIL_RANG_BAS_STRICT
    nps_haut = rang_nps >= SEUIL_RANG_HAUT_STRICT
    csat_bas = rang_csat is not None and rang_csat <= SEUIL_RANG_BAS_STRICT
    csat_haut = rang_csat is not None and rang_csat >= SEUIL_RANG_HAUT_STRICT
    effort_degrade = (
        (rang_reopens is not None and rang_reopens >= SEUIL_RANG_HAUT_STRICT)
        or (rang_resolution is not None and rang_resolution >= SEUIL_RANG_HAUT_STRICT)
    )

    type_alignement = None
    if amplitude_suffisante:
        if nps_bas and (csat_bas or effort_degrade):
            type_alignement = "alignement_negatif"
        elif nps_haut and csat_haut and not effort_degrade:
            type_alignement = "alignement_positif"
        elif nps_bas and csat_haut and not effort_degrade:
            type_alignement = "divergence"
        elif nps_haut and (csat_bas or effort_degrade):
            type_alignement = "divergence"

    return {
        "type": type_alignement,
        "rang_nps": rang_nps,
        "rang_csat": rang_csat,
        "effort_degrade": effort_degrade,
        "amplitude_nps": amplitude_nps,
        "amplitude_suffisante": amplitude_suffisante,
    }


# Texte associé à l'alignement -- jamais de causalité, uniquement une coïncidence temporelle
# observée (Étape 4E, section 1/19).
def texte_alignement_care_nps(resultat_alignement, profil_care_mois, mois_nom):
    if resultat_alignement is None or resultat_alignement["type"] is None:
        return None

    type_alignement = resultat_alignement["type"]
    if type_alignement == "alignement_negatif":
        return (
            "Le NPS de " + mois_nom + " recule dans la série disponible, période où plusieurs indicateurs "
            "d'effort et/ou de satisfaction Care se dégradent également -- signal de confiance à surveiller."
        )
    if type_alignement == "alignement_positif":
        return (
            "Le NPS de " + mois_nom + " progresse dans la série disponible, période où l'expérience Care "
            "reste tenue -- contexte cohérent, sans signal majeur."
        )
    if type_alignement == "divergence":
        if profil_care_mois is not None and profil_care_mois["csat"] is not None and profil_care_mois["csat"] >= SEUIL_CSAT_INSATISFAISANT:
            return (
                "Le NPS de " + mois_nom + " se situe en retrait dans la série disponible, alors que le CSAT "
                "de la période reste au-dessus de " + str(SEUIL_CSAT_INSATISFAISANT) + " -- ces deux mesures "
                "de la voix du client ne racontent pas nécessairement la même chose."
            )
        return (
            "Le NPS et les signaux Care de " + mois_nom + " ne vont pas dans le même sens sur cette période "
            "-- à explorer plutôt qu'à expliquer automatiquement."
        )
    return None


# Segmentation "contact Care identifié" (Étape 4E, section 12/30) : fondée sur le matching
# ticket<->NPS (dernier_ticket_avant, inchangé), PAS sur le champ auto-déclaré a_contacte_support --
# deux dimensions différentes du dataset, jamais confondues.
def segmenter_nps_par_contact_care(reponses, index_tickets_email, fenetre_jours):
    avec_contact = []
    sans_contact = []
    for reponse in reponses:
        ticket = dernier_ticket_avant(reponse, index_tickets_email, fenetre_jours)
        if ticket is not None:
            avec_contact.append(reponse)
        else:
            sans_contact.append(reponse)

    return {
        "contact_identifie": {"reponses": avec_contact, "composition": calculer_composition_nps(avec_contact)},
        "aucun_contact_identifie": {"reponses": sans_contact, "composition": calculer_composition_nps(sans_contact)},
    }


# Type d'expérience associé au dernier contact Care avant la réponse NPS (déplacé depuis app.py,
# logique inchangée -- seul le libellé "aucun contact" est corrigé, voir TYPE_EXPERIENCE_AUCUN).
def determiner_type_experience_nps(reponse, index_tickets_email, fenetre_jours):
    ticket = dernier_ticket_avant(reponse, index_tickets_email, fenetre_jours)
    if ticket is None:
        return TYPE_EXPERIENCE_AUCUN

    categorie = categoriser(ticket)
    if categorie == CATEGORIE_SAV_PRODUIT:
        if ticket["prior_sav_count"] is not None and ticket["prior_sav_count"] >= 1:
            return TYPE_EXPERIENCE_SAV_RECURRENT
        return TYPE_EXPERIENCE_SAV
    if categorie == "Livraison":
        return TYPE_EXPERIENCE_LIVRAISON

    resolution = ticket["resolution_type"]
    if resolution is not None and "Remplacement" in resolution:
        return TYPE_EXPERIENCE_REMPLACEMENT

    resolution_heures = ticket["full_resolution_time_hours"]
    if resolution_heures is not None:
        if resolution_heures < SEUIL_RESOLUTION_RAPIDE_H:
            return TYPE_EXPERIENCE_RESOLUTION_RAPIDE
        return TYPE_EXPERIENCE_RESOLUTION_LONGUE

    return TYPE_EXPERIENCE_AUTRE


# Confiance par type d'expérience -- affiché seulement pour les groupes au moins étayés (seuil de
# prudence), jamais un classement ("SAV = mauvais NPS") : ces populations sont structurellement
# différentes (Étape 4E, section 17).
def analyser_nps_par_type_experience(reponses, index_tickets_email, fenetre_jours, seuil_min):
    par_type = {}
    for reponse in reponses:
        type_exp = determiner_type_experience_nps(reponse, index_tickets_email, fenetre_jours)
        if type_exp in par_type:
            par_type[type_exp].append(reponse)
        else:
            par_type[type_exp] = [reponse]

    resultats = []
    for type_exp, reponses_type in par_type.items():
        if len(reponses_type) < seuil_min:
            continue
        resultats.append({
            "type_experience": type_exp,
            "composition": calculer_composition_nps(reponses_type),
        })
    return resultats


# Service recovery (Étape 4E, section 16/38) : identifie seulement les cas COMPATIBLES avec une
# récupération positive (contact Care identifié + CSAT élevé sur ce ticket + réponse NPS
# promoteur) -- jamais une transformation causale démontrée. Aucun NPS avant/après individuel
# n'existe dans les données (une seule réponse par client dans l'immense majorité des cas), donc
# aucune comparaison "avant/après" n'est tentée ici, uniquement un état croisé.
def evaluer_cas_compatibles_service_recovery(reponses, index_tickets_email, fenetre_jours, seuil_csat_eleve):
    cas = []
    for reponse in reponses:
        if reponse["score"] < 9:  # doit être promoteur
            continue
        ticket = dernier_ticket_avant(reponse, index_tickets_email, fenetre_jours)
        if ticket is None:
            continue
        if ticket["csat"] is None or ticket["csat"] < seuil_csat_eleve:
            continue
        cas.append({"reponse": reponse, "ticket": ticket})
    return cas


def controler_qualite_donnees_nps(reponses, index_tickets_email, fenetre_jours):
    anomalies = []

    n_score_hors_bornes = 0
    n_email_manquant = 0
    n_date_manquante = 0
    for reponse in reponses:
        score = reponse["score"]
        if score is None or score < 0 or score > 10:
            n_score_hors_bornes = n_score_hors_bornes + 1
        if not reponse["email_client"]:
            n_email_manquant = n_email_manquant + 1
        if not reponse["date_reponse"]:
            n_date_manquante = n_date_manquante + 1

    compte_email = {}
    for reponse in reponses:
        email = reponse["email_client"]
        compte_email[email] = compte_email.get(email, 0) + 1
    n_emails_avec_plusieurs_reponses = 0
    for email, compte in compte_email.items():
        if compte > 1:
            n_emails_avec_plusieurs_reponses = n_emails_avec_plusieurs_reponses + 1

    n_ticket_posterieur = 0
    for reponse in reponses:
        ticket = dernier_ticket_avant(reponse, index_tickets_email, fenetre_jours)
        if ticket is not None and ticket["created_at"] > reponse["date_reponse"]:
            n_ticket_posterieur = n_ticket_posterieur + 1

    if n_score_hors_bornes > 0:
        anomalies.append(str(n_score_hors_bornes) + " réponse(s) NPS avec un score hors de l'échelle 0-10.")
    if n_email_manquant > 0:
        anomalies.append(str(n_email_manquant) + " réponse(s) NPS sans email client.")
    if n_date_manquante > 0:
        anomalies.append(str(n_date_manquante) + " réponse(s) NPS sans date de réponse.")
    if n_emails_avec_plusieurs_reponses > 0:
        anomalies.append(
            str(n_emails_avec_plusieurs_reponses) + " client(s) avec plusieurs réponses NPS sur l'historique "
            "disponible (information, pas une erreur)."
        )
    if n_ticket_posterieur > 0:
        anomalies.append(
            str(n_ticket_posterieur) + " cas où le ticket associé est postérieur à la réponse NPS -- ne "
            "devrait jamais arriver (anomalie de matching à investiguer)."
        )

    return anomalies


# ---------------------------------------------------------------------------
# Composition Vue d'ensemble (Étape 5A, éditorialisée en 5A.1)
#
# Aucun moteur métier recalculé ici : ces fonctions consomment UNIQUEMENT les sorties déjà
# produites par les moteurs validés (Produit/Livraison/Avant-vente/Tendances/Impact & confiance).
# Elles sélectionnent, dédupliquent/fusionnent par DIMENSIONS STRUCTURÉES (jamais par
# correspondance de texte ni matching sémantique) et composent -- jamais de réévaluation
# d'éligibilité des moteurs sources, jamais de nouveau score global opaque
# ("Health score 72/100" et assimilés explicitement interdits).
#
# 5A.1 distingue deux types de signal, une classification qui n'existe qu'à ce niveau de
# composition (n'affecte aucun moteur source) :
#   - TRANSVERSAL : Tendances (vigilance) et/ou Impact & confiance (alignement NPS) -- décrivent
#     une dégradation globale de l'expérience/l'effort, pas un sujet précis.
#   - DIAGNOSTIQUE/CATÉGORIEL : Produit/Livraison/Avant-vente -- disent OÙ regarder.
# Un signal transversal (Tendances + NPS alignement négatif) peut être FUSIONNÉ en une seule
# histoire quand des dimensions structurées se recoupent (CSAT/effort/part de catégorie, lues sur
# les profils déjà chargés pour Tendances -- jamais un nouveau champ métier). La divergence NPS
# reste distincte par construction (elle raconte autre chose qu'une dégradation alignée).
# ---------------------------------------------------------------------------

SEUIL_MAX_SIGNAUX_ATTENTION_VUE_ENSEMBLE = 3
FENETRE_ANTICIPATION_VUE_ENSEMBLE_JOURS = 30
MAX_ANTICIPATIONS_VUE_ENSEMBLE = 2

CATEGORIE_LIVRAISON_VUE_ENSEMBLE = "Livraison"
CATEGORIE_AVANT_VENTE_VUE_ENSEMBLE = "Avant-vente / conseil"

_PHRASES_FAMILLE_COMMUNE_VUE_ENSEMBLE = {
    "B": "une satisfaction en retrait",
    "C": "un effort de traitement plus soutenu",
    "D": "un coût par dossier plus élevé",
    "E": "une récurrence observée sur plusieurs périodes",
}


# Ne garde que le tier le plus haut de chaque moteur catégoriel ("Priorité principale" pour
# Produit/Livraison, la totalité de la liste "opportunites" déjà pré-filtrée par le moteur pour
# Avant-vente) -- les tiers secondaires/à surveiller restent la propriété des onglets spécialisés,
# jamais remontés ici (évite que la Vue d'ensemble affiche systématiquement les mêmes signaux de
# fond à chaque période, quel que soit le contexte réel).
def extraire_candidats_categoriels_vue_ensemble(
    signaux_produit_prioritaires, signaux_livraison_prioritaires, signaux_av_opportunites,
):
    candidats = []
    for signal in signaux_produit_prioritaires:
        if signal["niveau_priorite"] == "Priorité principale":
            candidats.append({
                "categorie": CATEGORIE_SAV_PRODUIT, "onglet_cible": "Produit",
                "sujet": signal["sujet"], "observation_principale": signal["observation_principale"],
                "familles_actives": signal["familles_actives"], "volume_n": signal["volume"]["n"],
                "part_univers_pct": signal["volume"]["part_univers_pct"],
            })
    for signal in signaux_livraison_prioritaires:
        if signal["niveau_priorite"] == "Priorité principale":
            candidats.append({
                "categorie": CATEGORIE_LIVRAISON_VUE_ENSEMBLE, "onglet_cible": "Livraison",
                "sujet": signal["sujet"], "observation_principale": signal["observation_principale"],
                "familles_actives": signal["familles_actives"], "volume_n": signal["volume"]["n"],
                "part_univers_pct": signal["volume"]["part_univers_pct"],
            })
    for signal in signaux_av_opportunites:
        candidats.append({
            "categorie": CATEGORIE_AVANT_VENTE_VUE_ENSEMBLE, "onglet_cible": "Avant-vente & conversion",
            "sujet": signal["sujet"], "observation_principale": signal["observation_principale"],
            "familles_actives": signal["familles_actives"], "volume_n": signal["volume"]["n"],
            "part_univers_pct": signal["volume"]["part_univers_pct"],
        })
    return candidats


# Matérialité pour la Vue d'ensemble (Étape 5A.1, section 6-7) : un signal catégoriel déjà
# "Priorité principale"/"Opportunité" dans son propre moteur n'est pas automatiquement matériel
# pour une synthèse manager de 30 secondes. Règle explicable, jamais un score :
#   - Produit/Livraison (familles B=expérience/C=effort communes aux deux moteurs) : matériel
#     seulement si B (expérience client réellement dégradée) est active -- le critère le plus
#     proche d'un "vrai impact client cette période" parmi les familles déjà calculées. E
#     (récurrence) est délibérément EXCLU de ce test : un problème chronique déjà connu chaque
#     période n'est pas ce qui distingue CETTE période-ci.
#   - Avant-vente (pas de famille B dans son schéma -- A/C/H seulement, C étant déjà une condition
#     d'éligibilité systématique de "opportunité") : matériel seulement si son volume représente
#     une part réellement notable de la catégorie (réutilise SEUIL_VOLUME_PART_NOTABLE, déjà
#     utilisé par Produit/Livraison pour leur propre famille A -- pas un nouveau seuil inventé).
def signal_categoriel_est_materiel_vue_ensemble(candidat):
    if candidat["categorie"] == CATEGORIE_AVANT_VENTE_VUE_ENSEMBLE:
        part = candidat["part_univers_pct"]
        return part is not None and (part / 100) >= SEUIL_VOLUME_PART_NOTABLE
    return "B" in candidat["familles_actives"]


def filtrer_candidats_materiels_vue_ensemble(candidats):
    materiels = []
    for candidat in candidats:
        if signal_categoriel_est_materiel_vue_ensemble(candidat):
            materiels.append(candidat)
    return materiels


def _texte_familles_communes_vue_ensemble(groupe):
    familles_communes = None
    for candidat in groupe:
        familles_candidat = set(candidat["familles_actives"])
        if familles_communes is None:
            familles_communes = familles_candidat
        else:
            familles_communes = familles_communes & familles_candidat

    if familles_communes is None:
        return None

    phrases = []
    for lettre in ("B", "C", "D", "E"):
        if lettre in familles_communes:
            phrases.append(_PHRASES_FAMILLE_COMMUNE_VUE_ENSEMBLE[lettre])

    if len(phrases) == 0:
        return None

    texte = phrases[0]
    for i in range(1, len(phrases)):
        if i == len(phrases) - 1:
            texte = texte + " et " + phrases[i]
        else:
            texte = texte + ", " + phrases[i]
    return texte


# Fusionne, par CATÉGORIE STRUCTURÉE uniquement (jamais par texte ni par preuve partagée -- voir
# le principe déjà verrouillé pour Produit en 4A.3), les candidats qui partagent la même catégorie
# source. Le point commun affiché (familles B/C/D/E partagées) est lu sur les familles_actives
# déjà calculées par chaque moteur, jamais reparsé depuis un texte généré.
def regrouper_candidats_par_categorie_vue_ensemble(candidats):
    par_categorie = {}
    for candidat in candidats:
        categorie = candidat["categorie"]
        if categorie in par_categorie:
            par_categorie[categorie].append(candidat)
        else:
            par_categorie[categorie] = [candidat]

    regroupes = []
    for categorie, groupe in par_categorie.items():
        volume_total = 0
        for candidat in groupe:
            volume_total = volume_total + candidat["volume_n"]

        if len(groupe) == 1:
            unique = groupe[0]
            regroupes.append({
                "categorie": categorie, "onglet_cible": unique["onglet_cible"],
                "titre": categorie + " — " + unique["sujet"],
                "texte": unique["observation_principale"],
                "volume_n": unique["volume_n"],
            })
        else:
            sujets = []
            for candidat in groupe:
                sujets.append(candidat["sujet"])
            texte_sujets = sujets[0]
            for i in range(1, len(sujets)):
                if i == len(sujets) - 1:
                    texte_sujets = texte_sujets + " et " + sujets[i]
                else:
                    texte_sujets = texte_sujets + ", " + sujets[i]

            texte_communes = _texte_familles_communes_vue_ensemble(groupe)
            if texte_communes is not None:
                texte = (
                    categorie + " concentre plusieurs signaux de la période (" + texte_sujets
                    + "), avec sur chacun " + texte_communes + "."
                )
            else:
                texte = (
                    categorie + " concentre plusieurs signaux distincts de la période (" + texte_sujets + ")."
                )

            regroupes.append({
                "categorie": categorie, "onglet_cible": groupe[0]["onglet_cible"],
                "titre": categorie + " — plusieurs signaux",
                "texte": texte,
                "volume_n": volume_total,
            })

    return regroupes


# Diagnostics structurés pour la fusion transversale (Étape 5A.1, section 3) : lit le rang de
# CSAT/effort/part-de-catégorie sur les profils déjà chargés pour Tendances (aucune donnée ni
# calcul métier nouveau -- rang_relatif est la même primitive déjà utilisée partout ailleurs).
# Jamais de NLP sur le texte de la vigilance : uniquement des champs déjà structurés.
def evaluer_diagnostics_structures_transversal_vue_ensemble(profils_historique, index_dernier, categories_a_verifier):
    if index_dernier < 0 or index_dernier >= len(profils_historique):
        return {"csat_bas": False, "effort_haut": False, "categories_part_haute": []}

    csat_valeurs = []
    reopens_valeurs = []
    resolution_valeurs = []
    replies_valeurs = []
    for profil in profils_historique:
        csat_valeurs.append(profil["csat"])
        reopens_valeurs.append(profil["reopens"])
        resolution_valeurs.append(profil["resolution_h"])
        replies_valeurs.append(profil["replies"])

    rang_csat = rang_relatif(csat_valeurs, index_dernier)
    rang_reopens = rang_relatif(reopens_valeurs, index_dernier)
    rang_resolution = rang_relatif(resolution_valeurs, index_dernier)
    rang_replies = rang_relatif(replies_valeurs, index_dernier)

    csat_bas = rang_csat is not None and rang_csat <= SEUIL_RANG_BAS_STRICT
    effort_haut = (
        (rang_reopens is not None and rang_reopens >= SEUIL_RANG_HAUT_STRICT)
        or (rang_resolution is not None and rang_resolution >= SEUIL_RANG_HAUT_STRICT)
        or (rang_replies is not None and rang_replies >= SEUIL_RANG_HAUT_STRICT)
    )

    categories_part_haute = []
    for categorie in categories_a_verifier:
        parts_categorie = []
        for profil in profils_historique:
            volume_profil = profil["volume"]
            if volume_profil is not None and volume_profil > 0:
                parts_categorie.append(profil["mix_categories"].get(categorie, 0) / volume_profil)
            else:
                parts_categorie.append(None)
        rang_part = rang_relatif(parts_categorie, index_dernier)
        if rang_part is not None and rang_part >= SEUIL_RANG_HAUT_STRICT:
            categories_part_haute.append(categorie)

    return {"csat_bas": csat_bas, "effort_haut": effort_haut, "categories_part_haute": categories_part_haute}


# Texte du signal transversal fusionné -- composé UNIQUEMENT à partir de champs structurés
# (diagnostics ci-dessus), jamais un parsing du texte de la vigilance. Le texte NPS, lui, est
# repris tel quel (déjà généré et verrouillé par le moteur Impact & confiance), en corroboration.
def texte_signal_transversal_vue_ensemble(diagnostics, texte_alignement_nps_fusionne):
    fragments = []
    for categorie in diagnostics["categories_part_haute"]:
        fragments.append(categorie + " qui prend davantage de place")
    if diagnostics["effort_haut"]:
        fragments.append("un effort de traitement plus soutenu (résolution, réouvertures ou échanges)")
    if diagnostics["csat_bas"]:
        fragments.append("une satisfaction en retrait")

    if len(fragments) == 0:
        texte = "L'expérience client se dégrade sur cette période, sans dimension structurée dominante identifiée."
    else:
        texte_fragments = fragments[0]
        for i in range(1, len(fragments)):
            if i == len(fragments) - 1:
                texte_fragments = texte_fragments + " et " + fragments[i]
            else:
                texte_fragments = texte_fragments + ", " + fragments[i]
        texte = "Expérience client sous tension : " + texte_fragments + "."

    if texte_alignement_nps_fusionne is not None:
        texte = texte + " Le NPS évolue également dans le même sens."

    return texte


# Orchestration "Ce qui mérite votre attention" (Étape 5A.1) :
# 1. Signal transversal : Tendances (vigilance) et NPS "alignement_negatif" FUSIONNÉS en une seule
#    histoire structurée quand les deux sont présents ; NPS seul ou Tendances seul restent une
#    histoire transversale à eux seuls. "divergence" NPS reste TOUJOURS distincte (elle raconte
#    autre chose qu'une dégradation alignée -- jamais fusionnée).
# 2. Signaux catégoriels : filtrés par matérialité AVANT regroupement/tri (jamais après) --
#    un signal non matériel reste dans son onglet spécialisé, jamais remonté ici.
# 3. Plafond appliqué EN DERNIER, après fusion et matérialité -- jamais le mécanisme de sélection
#    principal (section 15).
def construire_signaux_attention_vue_ensemble(
    candidats_categoriels, vigilance_tendances, alignement_nps, texte_alignement_nps, diagnostics_transversaux, nombre_max,
):
    signaux_bruts = []

    nps_fusionnable = (
        alignement_nps is not None and alignement_nps["type"] == "alignement_negatif" and texte_alignement_nps is not None
    )

    if vigilance_tendances is not None or nps_fusionnable:
        texte_nps_pour_fusion = None
        if nps_fusionnable:
            texte_nps_pour_fusion = texte_alignement_nps

        if vigilance_tendances is not None and nps_fusionnable:
            onglet_transversal = "Tendances, Impact & confiance"
        elif nps_fusionnable:
            onglet_transversal = "Impact & confiance"
        else:
            onglet_transversal = "Tendances"

        texte_transversal = texte_signal_transversal_vue_ensemble(diagnostics_transversaux, texte_nps_pour_fusion)
        signaux_bruts.append({
            "categorie": None, "onglet_cible": onglet_transversal, "titre": "Expérience client sous tension",
            "texte": texte_transversal, "volume_n": None, "priorite_tri": 0,
        })

    if alignement_nps is not None and alignement_nps["type"] == "divergence" and texte_alignement_nps is not None:
        signaux_bruts.append({
            "categorie": None, "onglet_cible": "Impact & confiance", "titre": "Confiance (NPS)",
            "texte": texte_alignement_nps, "volume_n": None, "priorite_tri": 0,
        })

    candidats_materiels = filtrer_candidats_materiels_vue_ensemble(candidats_categoriels)
    candidats_regroupes = regrouper_candidats_par_categorie_vue_ensemble(candidats_materiels)
    for candidat in candidats_regroupes:
        candidat["priorite_tri"] = 1
        signaux_bruts.append(candidat)

    def _cle_tri_signal_vue_ensemble(signal):
        volume = signal["volume_n"]
        if volume is None:
            volume = 0
        return (signal["priorite_tri"], -volume)

    signaux_tries = sorted(signaux_bruts, key=_cle_tri_signal_vue_ensemble)

    retenus = []
    for i in range(min(nombre_max, len(signaux_tries))):
        retenus.append(signaux_tries[i])

    non_retenus = []
    for i in range(min(nombre_max, len(signaux_tries)), len(signaux_tries)):
        non_retenus.append(signaux_tries[i])

    return {"retenus": retenus, "non_retenus": non_retenus, "bruts_avant_composition": signaux_bruts}


# Contrôle de cohérence Lecture <-> Attention (Étape 5A.1, section 13) : un pur prédicat
# d'observation (audit/test), jamais utilisé pour réécrire la Lecture ni pour supprimer
# silencieusement des cartes -- "Réévaluer d'abord la matérialité des cartes" (déjà le mécanisme
# ci-dessus), pas une seconde suppression. Incohérent seulement si la Lecture ne signale aucune
# vigilance ET que la sélection Attention est malgré tout pleine (au plafond) -- le signe que la
# matérialité n'a pas suffisamment filtré pour cette période.
def verifier_coherence_lecture_attention_vue_ensemble(vigilance_absente, nb_signaux_retenus, plafond):
    if vigilance_absente and nb_signaux_retenus >= plafond:
        return False
    return True


# "Ce qui tient" -- jamais une carte positive forcée (0 est valide). Deux sources possibles,
# toutes deux structurelles : (1) capacité prévue ET volume simultanément bas dans la distribution
# des observations disponibles (rang_relatif, réutilisé tel quel) -- "capacité réduite mais demande
# faible, pas de tension" ; (2) un alignement NPS positif déjà validé par le moteur Impact.
def construire_signal_positif_vue_ensemble(rang_capacite, rang_volume, alignement_nps, texte_alignement_nps):
    signaux = []
    if rang_capacite is not None and rang_volume is not None:
        if rang_capacite <= SEUIL_RANG_BAS_STRICT and rang_volume <= SEUIL_RANG_BAS_STRICT:
            signaux.append(
                "Capacité prévue réduite sur la période, mais demande également faible à l'échelle des "
                "observations disponibles -- pas de tension observée malgré la capacité restreinte."
            )
    if alignement_nps is not None and alignement_nps["type"] == "alignement_positif" and texte_alignement_nps is not None:
        signaux.append(texte_alignement_nps)
    return signaux[:2]


# Points d'anticipation (Étape 5A.1, section 16-20) : un événement compte comme anticipation
# seulement si sa date de DÉBUT est strictement postérieure à la fin de la période observée --
# jamais via le chevauchement de contexte_periode (date_fin_evenement >= date_fin_periode), qui
# laisserait passer un événement déjà commencé et encore en cours (section 18-19 : un tel
# événement reste uniquement "contexte actuel", jamais dupliqué en anticipation). Sélection
# plafonnée (MAX_ANTICIPATIONS_VUE_ENSEMBLE) et priorisée sur les transitions Staffing (capacité)
# avant les événements Commercial/Produit -- pas un mini-calendrier des 30 prochains jours.
def construire_points_anticipation_vue_ensemble(evenements, date_fin_periode, nombre_max=MAX_ANTICIPATIONS_VUE_ENSEMBLE):
    date_fin_fenetre = date_fin_periode + datetime.timedelta(days=FENETRE_ANTICIPATION_VUE_ENSEMBLE_JOURS)

    candidats = []
    for evenement in evenements:
        if evenement["date_debut"] is None:
            continue
        if evenement["date_debut"] > date_fin_periode and evenement["date_debut"] <= date_fin_fenetre:
            candidats.append(evenement)

    def _cle_tri_anticipation(evenement):
        if evenement["type"] == "Staffing":
            priorite_type = 0
        else:
            priorite_type = 1
        return (priorite_type, evenement["date_debut"])

    candidats_tries = sorted(candidats, key=_cle_tri_anticipation)

    retenus = []
    for i in range(min(nombre_max, len(candidats_tries))):
        retenus.append(candidats_tries[i])
    return retenus


# Onglets vers lesquels approfondir -- dérivés des signaux réellement affichés (retenus), jamais
# une liste fixe de 9 boutons permanents (section 20).
def construire_navigation_vue_ensemble(signaux_retenus):
    onglets = []
    for signal in signaux_retenus:
        onglet = signal["onglet_cible"]
        if onglet not in onglets:
            onglets.append(onglet)
    return onglets


# ---------------------------------------------------------------------------
# Composition Tendances UI (Étape 5B)
#
# Le moteur Tendances (Étape 4B, verrouillé) n'est pas modifié : ces fonctions consomment
# uniquement ses sorties (construire_lecture_tendances, profils_historique déjà chargés via
# construire_profil_observation) pour composer l'affichage -- aucune règle de vigilance/jalon/
# contraste/saisonnalité recalculée ici.
# ---------------------------------------------------------------------------

# Lecture pure d'un dict déjà produit par le moteur (profil["mix_categories"]), jamais une
# nouvelle règle métier : la catégorie qui compte le plus de tickets sur l'observation.
def categorie_dominante_mix_tendances(mix_categories):
    dominante = None
    plus_grand = -1
    for categorie, n in mix_categories.items():
        if n > plus_grand:
            plus_grand = n
            dominante = categorie
    return dominante


# "Période analysée" vs "Historique de référence" (Étape 5B, section 8) : une ligne discrète,
# dérivée des profils déjà chargés (aucune fuite du futur par construction -- profils_historique
# ne contient jamais une observation postérieure à la période sélectionnée, voir onglet_tendances).
# Mode observation unique -> la période analysée est l'observation elle-même ; sinon -> la fenêtre
# des nb_observations_periode dernières observations du profils_historique fourni.
def construire_texte_periode_reference_tendances(profils_historique, mode, nb_observations_periode):
    if len(profils_historique) == 0:
        return None

    profils_periode = profils_historique[len(profils_historique) - nb_observations_periode:]

    if mode == MODE_OBSERVATION_UNIQUE:
        texte_periode_analysee = (
            "Période analysée : " + str(profils_periode[0]["date_debut"]) + " – " + str(profils_periode[0]["date_fin"])
        )
    else:
        texte_periode_analysee = (
            "Période analysée : " + str(profils_periode[0]["date_debut"]) + " → " + str(profils_periode[-1]["date_fin"])
        )

    texte_historique_reference = "Historique de référence : jusqu'au " + str(profils_historique[-1]["date_fin"])
    return texte_periode_analysee + "  ·  " + texte_historique_reference


# ---------------------------------------------------------------------------
# Composition Agents (Étape 5C.1) -- "Portrait factuel des contributions"
#
# Aucun score, aucun classement, aucun jugement ("meilleur"/"pire"/"performance") : uniquement des
# constats descriptifs (charge relative aux heures planifiées, mix de catégories, CSAT+n) que le
# manager interprète lui-même. Pas de moteur "Force à valoriser"/"Point d'accompagnement"/
# "À investiguer" à ce stade (reporté à une étape ultérieure sur décision explicite).
# ---------------------------------------------------------------------------

STATUT_AGENT_PLANIFIE_ACTIF = "planifie_actif"
STATUT_AGENT_RENFORT_NON_PLANIFIE = "renfort_non_planifie"
STATUT_AGENT_PLANIFIE_SANS_ACTIVITE = "planifie_sans_activite"
STATUT_AGENT_ABSENT = "absent"


# Somme exacte des créneaux Planning de l'agent sur la période (plusieurs créneaux/jour déjà gérés
# par construction -- planning[agent][jour] est une LISTE de (début, fin)). Jamais déduit des
# tickets (verrou Étape 1 : planning prévu ≠ activité observée).
def heures_planifiees_agent(planning, agent):
    total = 0
    for jour, plages in planning.get(agent, {}).items():
        for debut, fin in plages:
            total = total + (fin - debut)
    return total


# Répartition en % par catégorie -- dict creux (une catégorie absente n'a simplement pas de clé,
# jamais une KeyError côté appelant si celui-ci utilise .get(categorie, 0)).
def mix_pct_agent(tickets_agent):
    n = len(tickets_agent)
    if n == 0:
        return {}
    compte = {}
    for ticket in tickets_agent:
        categorie = categoriser(ticket)
        if categorie in compte:
            compte[categorie] = compte[categorie] + 1
        else:
            compte[categorie] = 1
    mix = {}
    for categorie, nombre in compte.items():
        mix[categorie] = nombre / n * 100
    return mix


# Ratio purement descriptif (jamais "productivité") -- None si l'agent n'a aucune heure planifiée
# cette période (évite toute division par zéro, jamais un 0 ou un ∞ trompeur).
def charge_relative_agent(nb_tickets, heures_planifiees):
    if heures_planifiees is None or heures_planifiees == 0:
        return None
    return nb_tickets / heures_planifiees


# Événement Staffing structurellement lié à cet agent (champ "perimetre", jamais un texte parsé/
# deviné) dont la fenêtre chevauche la période observée -- même sémantique de chevauchement que
# contexte_periode, restreinte à un seul agent nommé.
def _evenement_absence_agent(evenements, agent, date_debut, date_fin):
    for evenement in evenements:
        if evenement["type"] != "Staffing":
            continue
        if evenement["perimetre"] != agent:
            continue
        if evenement["date_fin"] >= date_debut and evenement["date_debut"] <= date_fin:
            return evenement
    return None


def _cle_tri_agent_alphabetique(ligne_roster):
    return ligne_roster["agent"]


# Population affichée = union structurée de trois sources, jamais un roster historique fabriqué
# (Étape 5C.1, section 11) :
#   - agents avec activité observée (tickets) cette période ;
#   - agents planifiés (Planning) cette période, même à 0 ticket ;
#   - agents nommés par un événement Staffing (perimetre) dont la fenêtre chevauche la période --
#     seul moyen structuré de savoir qu'un agent existe mais est absent (congé), sans planning ni
#     ticket cette semaine-là.
# Un agent qui n'apparaît dans AUCUNE des trois sources n'existe simplement pas pour cette période
# (ex. Sam hors décembre, Sofia avant le 15/09, Lucie hors janvier-juin) -- jamais une ligne à 0
# fabriquée. Tri alphabétique (section 9) : source la plus stable et neutre disponible, aucun
# classement par volume/CSAT.
#
# plannings_periode : même format que construire_plannings_periode -- une liste de
# (date_debut, date_fin, planning_dict), une entrée par export couvert par la période (une période
# étendue peut en contenir plusieurs) -- jamais un seul dict de planning brut, pour rester cohérent
# avec le format déjà utilisé partout ailleurs dans l'application (planning_s2).
def construire_roster_agents(tickets_periode, plannings_periode, evenements, date_debut, date_fin):
    tickets_par_agent = grouper_par(tickets_periode, "assignee")

    agents_connus = set()
    for agent in tickets_par_agent.keys():
        if agent is not None:
            agents_connus.add(agent)
    for date_debut_semaine, date_fin_semaine, planning_semaine in plannings_periode:
        for agent in planning_semaine.keys():
            if agent != NOM_AGENT_DEFAUT:
                agents_connus.add(agent)
    for evenement in evenements:
        if evenement["type"] == "Staffing" and evenement["perimetre"] is not None:
            if evenement["date_fin"] >= date_debut and evenement["date_debut"] <= date_fin:
                agents_connus.add(evenement["perimetre"])

    roster = []
    for agent in agents_connus:
        tickets_agent = tickets_par_agent.get(agent, [])
        heures = 0
        for date_debut_semaine, date_fin_semaine, planning_semaine in plannings_periode:
            heures = heures + heures_planifiees_agent(planning_semaine, agent)
        nb_tickets = len(tickets_agent)

        if nb_tickets > 0 and heures > 0:
            statut = STATUT_AGENT_PLANIFIE_ACTIF
            evenement_absence = None
        elif nb_tickets > 0 and heures == 0:
            statut = STATUT_AGENT_RENFORT_NON_PLANIFIE
            evenement_absence = None
        elif nb_tickets == 0 and heures > 0:
            statut = STATUT_AGENT_PLANIFIE_SANS_ACTIVITE
            evenement_absence = None
        else:
            statut = STATUT_AGENT_ABSENT
            evenement_absence = _evenement_absence_agent(evenements, agent, date_debut, date_fin)

        roster.append({
            "agent": agent,
            "tickets": tickets_agent,
            "heures_planifiees": heures,
            "statut": statut,
            "evenement_absence": evenement_absence,
        })

    return sorted(roster, key=_cle_tri_agent_alphabetique)


# Lecture d'équipe purement factuelle (Étape 5C.1, section 20-22) : qui contribue cette période,
# jamais qui performe le mieux. Aucun texte codé par agent -- générique, fonctionne pour n'importe
# quelle composition d'équipe. Le mix détaillé (qui traite plus de quoi) reste dans la table/le
# détail, jamais transformé ici en comparaison "plus marquée" (aucun seuil défendable identifié,
# voir compte-rendu Étape 5C.1 section 23).
def construire_lecture_equipe_agents(roster):
    agents_actifs = []
    for ligne in roster:
        if ligne["statut"] == STATUT_AGENT_PLANIFIE_ACTIF or ligne["statut"] == STATUT_AGENT_RENFORT_NON_PLANIFIE:
            agents_actifs.append(ligne["agent"])

    if len(agents_actifs) == 0:
        return "Aucune activité observée pour cette période."

    texte_agents = agents_actifs[0]
    for i in range(1, len(agents_actifs)):
        if i == len(agents_actifs) - 1:
            texte_agents = texte_agents + " et " + agents_actifs[i]
        else:
            texte_agents = texte_agents + ", " + agents_actifs[i]

    return (
        "Cette période, l'activité est portée par " + texte_agents + ". Les portefeuilles traités "
        "diffèrent d'un agent à l'autre (voir la répartition par catégorie ci-dessous) -- les écarts "
        "de résolution, d'échanges ou d'utilisation des macros doivent être lus avec ce contexte."
    )


# Historique d'un agent -- UNIQUEMENT les observations où il a une activité réelle (tickets > 0),
# jamais un point à 0 fabriqué pendant une absence ou avant son arrivée (Étape 5C.1, section 27) :
# une semaine où l'agent n'apparaît pas dans les tickets n'entre simplement pas dans la liste
# retournée. Aucune fuite du futur : à l'appelant de ne fournir que des exports <= date_a_fin
# (même discipline que Tendances 4B/5B).
def construire_historique_agent(exports_avec_tickets_et_planning, agent):
    historique = []
    for date_debut_export, date_fin_export, tickets_export, planning_export in exports_avec_tickets_et_planning:
        tickets_agent = []
        for ticket in tickets_export:
            if ticket["assignee"] == agent:
                tickets_agent.append(ticket)

        if len(tickets_agent) == 0:
            continue

        heures = heures_planifiees_agent(planning_export, agent)
        historique.append({
            "date_debut": date_debut_export,
            "date_fin": date_fin_export,
            "tickets": len(tickets_agent),
            "heures_planifiees": heures,
            "charge_relative": charge_relative_agent(len(tickets_agent), heures),
            "csat": moyenne(tickets_agent, "csat"),
            "n_csat": _compte_valeurs_non_nulles(tickets_agent, "csat"),
            "mix_pct": mix_pct_agent(tickets_agent),
        })
    return historique


# ------------------------------------------------------------------
# Composition Actions & améliorations (Étape 5D.1)
#
# Remplace l'ancien onglet "Alertes & suggestions" (Étape 5D, audit) : ce n'est plus une page
# d'alertes triées par un score inter-familles opaque, mais une petite boucle d'amélioration
# continue CX -- pistes par famille explicable, actions déjà menées, ce qu'on observe après
# (jamais présenté comme une preuve de causalité). Aucun moteur 4A-4E n'est touché ni consommé
# ici : les signaux transverses restent la propriété de Vue d'ensemble et des onglets spécialisés
# (voir compte-rendu Étape 5D, section 41 -- ce module ne redevient pas un "centre des signaux").
# ------------------------------------------------------------------

FAMILLE_STANDARDISATION_ACTIONS = "Standardisation"
FAMILLE_SELF_SERVICE_ACTIONS = "Self-service"
FAMILLE_RETOURS_CLIENTS_ACTIONS = "Retours clients"

SEUIL_MACRO_BASSE_ACTIONS = 20
SEUIL_MACRO_HAUTE_ACTIONS = 50
SEUIL_REPLIES_FAQ_ACTIONS = 3
SEUIL_CSAT_VERBATIM_ACTIONS = 2
SEUIL_VERBATIMS_GROUPE_ACTIONS = 10

TEXTE_PRUDENCE_AVANT_APRES_ACTIONS = (
    "Cette évolution ne permet pas, à elle seule, d'attribuer le changement à l'action -- "
    "d'autres facteurs de la période ont pu jouer."
)


def sujet_deja_traite_actions(sujet, suivi_suggestions):
    entree = suivi_suggestions.get(sujet)
    if entree is None:
        return False
    return entree["statut"] == "Fait" and entree["date_action"] is not None


def _cle_tri_volume_piste_actions(piste):
    return piste["volume"]


# Piste de standardisation : sujet avec volume suffisant, satisfaction insuffisante, et une macro
# absente ou insuffisamment adoptée. Wording jamais prescriptif ("Piste : évaluer...", jamais
# "Créer une macro") -- voir compte-rendu Étape 5D.1 section 7. Triée par volume décroissant
# (critère affiché, pas de score composite caché).
def identifier_pistes_standardisation(tickets_periode, suivi_suggestions, seuil_minimum_sujet, seuil_csat_insatisfaisant):
    sujets = grouper_par(tickets_periode, "subject_cluster")
    pistes = []
    for sujet, tickets_sujet in sujets.items():
        volume = len(tickets_sujet)
        if volume < seuil_minimum_sujet:
            continue

        csat_sujet = moyenne(tickets_sujet, "csat")
        if csat_sujet is None or csat_sujet >= seuil_csat_insatisfaisant:
            continue

        if sujet_deja_traite_actions(sujet, suivi_suggestions):
            continue

        macro_sujet = taux_rempli(tickets_sujet, "macro_applied")
        if macro_sujet < SEUIL_MACRO_BASSE_ACTIONS:
            sous_type = "macro_absente"
            piste_texte = "Piste : évaluer la création d'une macro pour standardiser la réponse sur ce sujet."
        elif macro_sujet >= SEUIL_MACRO_HAUTE_ACTIONS:
            sous_type = "macro_insuffisante"
            piste_texte = (
                "Piste : la macro existe et est déjà bien utilisée -- explorer si son contenu ou la "
                "nature du problème sous-jacent explique la satisfaction encore insuffisante."
            )
        else:
            sous_type = "macro_partielle"
            piste_texte = "Piste : évaluer un renforcement de l'adoption de la macro déjà existante sur ce sujet."

        pistes.append({
            "famille": FAMILLE_STANDARDISATION_ACTIONS,
            "sujet": sujet,
            "sous_type": sous_type,
            "volume": volume,
            "csat": csat_sujet,
            "usage_macro_pct": macro_sujet,
            "piste": piste_texte,
        })
    return sorted(pistes, key=_cle_tri_volume_piste_actions, reverse=True)


# Piste de self-service : sujet avec volume suffisant et un nombre d'échanges moyen élevé --
# signe qu'une information manque au client dès le premier contact. Même discipline de wording
# et de tri que la standardisation.
def identifier_pistes_self_service(tickets_periode, suivi_suggestions, seuil_minimum_sujet, seuil_replies_faq):
    sujets = grouper_par(tickets_periode, "subject_cluster")
    pistes = []
    for sujet, tickets_sujet in sujets.items():
        volume = len(tickets_sujet)
        if volume < seuil_minimum_sujet:
            continue

        if sujet_deja_traite_actions(sujet, suivi_suggestions):
            continue

        replies_moyen = moyenne(tickets_sujet, "replies")
        if replies_moyen is None or replies_moyen < seuil_replies_faq:
            continue

        pistes.append({
            "famille": FAMILLE_SELF_SERVICE_ACTIONS,
            "sujet": sujet,
            "volume": volume,
            "echanges_moyens": replies_moyen,
            "csat": moyenne(tickets_sujet, "csat"),
            "piste": "Piste : évaluer l'ajout d'un contenu d'aide (FAQ) dédié à ce sujet.",
        })
    return sorted(pistes, key=_cle_tri_volume_piste_actions, reverse=True)


# Retours clients à explorer : regroupement de verbatims à CSAT très bas par sujet, à partir d'un
# volume de commentaires jugé significatif. Ce n'est jamais une "piste" au même titre que les deux
# familles ci-dessus (pas d'action suggérée) -- une matière d'investigation qualitative. Volontai-
# rement PAS filtré par suivi_suggestions (comportement identique à l'ancien onglet Alertes) : un
# sujet déjà marqué "Fait" peut légitimement continuer à faire remonter des verbatims négatifs
# distincts, ce n'est pas une "nouvelle piste" au sens de la section 12 du compte-rendu Étape 5D.1.
def identifier_retours_clients_a_explorer(tickets_periode, seuil_csat_verbatim, seuil_verbatims_groupe):
    tickets_verbatims = []
    for ticket in tickets_periode:
        csat_ticket = ticket["csat"]
        commentaire = ticket["csat_comment"]
        if csat_ticket is not None and csat_ticket <= seuil_csat_verbatim and commentaire:
            tickets_verbatims.append(ticket)

    sujets_verbatims = grouper_par(tickets_verbatims, "subject_cluster")
    groupes = []
    for sujet, tickets_sujet in sujets_verbatims.items():
        if len(tickets_sujet) >= seuil_verbatims_groupe:
            groupes.append({
                "famille": FAMILLE_RETOURS_CLIENTS_ACTIONS,
                "sujet": sujet,
                "volume": len(tickets_sujet),
                "csat": moyenne(tickets_sujet, "csat"),
                "tickets": tickets_sujet,
            })
    return sorted(groupes, key=_cle_tri_volume_piste_actions, reverse=True)


def _cle_tri_date_action_actions(action):
    return action["date_action"]


# Actions déjà menées, à partir de suivi_suggestions.xlsx -- SEULES les valeurs de statut réel-
# lement présentes dans le fichier sont utilisées (aujourd'hui "Fait" uniquement) ; aucun faux
# workflow (Nouveau/En cours/Bloqué) n'est inventé (Étape 5D.1 section 14). Discipline no-future-
# leakage stricte (section 16/36) : une action dont date_action est postérieure à date_fin_periode
# n'apparaît pas du tout dans le scope consulté, et l'impact "après" ne compte jamais un ticket créé
# après date_fin_periode -- même quand tickets_historique_business (source de impact_avant_apres)
# contient des données bien plus récentes que la période affichée. Une action sans date_action est
# exclue par prudence (section 17) : impossible de savoir si elle avait déjà eu lieu au moment de la
# période consultée. "Fait" ne colore jamais "succès" -- un impact neutre ou négatif reste affiché
# tel quel (section 18), rien n'est filtré sur la base du résultat.
def construire_actions_menees_actions(suivi_suggestions, tickets_historique_business, date_fin_periode):
    actions = []
    for sujet, entree in suivi_suggestions.items():
        if entree["statut"] != "Fait":
            continue

        date_action = entree["date_action"]
        if date_action is None:
            continue
        if date_action > date_fin_periode:
            continue

        tickets_sujet_bornes = []
        for ticket in tickets_historique_business:
            if ticket["subject_cluster"] != sujet:
                continue
            date_ticket = ticket["created_at"]
            if isinstance(date_ticket, datetime.datetime):
                date_ticket = date_ticket.date()
            if date_ticket <= date_fin_periode:
                tickets_sujet_bornes.append(ticket)

        impact = impact_avant_apres(tickets_sujet_bornes, date_action)

        actions.append({
            "sujet": sujet,
            "statut": entree["statut"],
            "date_action": date_action,
            "notes": entree["notes"],
            "impact": impact,
            "mesure_avant_disponible": impact["volume_avant"] > 0,
            "mesure_apres_disponible": impact["volume_apres"] > 0,
        })
    return sorted(actions, key=_cle_tri_date_action_actions, reverse=True)


# ------------------------------------------------------------------
# Composition Couverture -- pression / tension (Étape 5E.1)
#
# L'audit 5E a établi une confusion conceptuelle dans l'ancien mécanisme "hotspot" : le ratio
# demandes/agents planifiés mesure une PRESSION DE CHARGE ("combien de demandes par rapport à la
# capacité planifiée ?"), jamais une TENSION DE COUVERTURE au sens où l'entend un manager (une
# dégradation réelle de la réactivité). Les deux notions sont ici explicitement séparées :
#   - PRESSION = position du ratio demandes/capacité du créneau par rapport à l'historique
#     disponible (jamais un seuil absolu arbitraire) ;
#   - TENSION = convergence PRESSION MATÉRIELLE + RÉACTIVITÉ (FRT local) DÉGRADÉE.
# Step 1 (agents_en_poste, construire_activite_par_jour_heure, activite_observee,
# renfort_non_planifie) reste totalement inchangé et continue d'alimenter ce module -- aucune
# redéfinition locale de la capacité/activité observée. Discipline no-future-leakage identique à
# 4B/5A.1/5B/5C.1 : toute référence historique est bornée aux exports dont la date < date_limite
# (jamais la période affichée elle-même, jamais un export postérieur).
# ------------------------------------------------------------------

JOURS_ORDRE_COUVERTURE = [
    ("Lundi", 0), ("Mardi", 1), ("Mercredi", 2), ("Jeudi", 3),
    ("Vendredi", 4), ("Samedi", 5), ("Dimanche", 6),
]

HEURE_DEBUT_PRESSION_COUVERTURE = 7
HEURE_FIN_PRESSION_COUVERTURE = 21

# Garde-fou de volume (Étape 5E.1, section 9) : 1-4 demandes ne suffisent jamais à qualifier une
# pression ou une tension, quel que soit le ratio -- la heatmap continue d'afficher le nombre brut
# de demandes, seule la QUALIFICATION (pression/tension) est gelée sous ce seuil.
SEUIL_VOLUME_MIN_PRESSION_COUVERTURE = 5
SEUIL_VOLUME_MIN_FRT_LOCAL_COUVERTURE = 5

# Sous ce nombre de points de référence historique (cellules ou FRT locaux poolés depuis les
# exports strictement antérieurs), aucune classification relative n'est produite -- affichage de
# la valeur brute uniquement (Étape 5E.1, section 8 : "première observation").
SEUIL_MIN_REFERENCE_HISTORIQUE_COUVERTURE = 20

# Calibrés sur la distribution réelle des 14 exports disponibles (audit 5E.1, section 1) : Q75≈7,3,
# Q90≈9,5 sur les cellules à capacité>0 et volume≥5 -- exprimés ici en rang relatif (fraction,
# jamais une valeur de ratio en dur) pour rester valables quelle que soit l'échelle de l'historique
# disponible à la période consultée.
SEUIL_RANG_PRESSION_MARQUEE = 0.75
SEUIL_RANG_PRESSION_FORTE = 0.90
SEUIL_RANG_FRT_LOCAL_DEGRADE = 0.75

NIVEAU_PRESSION_HABITUELLE = "Pression habituelle"
NIVEAU_PRESSION_MARQUEE = "Pression marquée"
NIVEAU_PRESSION_FORTE = "Pression parmi les plus fortes observées"
NIVEAU_PRESSION_NON_QUALIFIABLE = "Historique insuffisant pour situer la pression"
NIVEAU_PRESSION_FAIBLE_VOLUME = "Volume trop faible pour être qualifié"

NIVEAU_ACTIVITE_HORS_CAPACITE_FAIBLE = "Activité hors capacité planifiée (volume faible)"
NIVEAU_ACTIVITE_HORS_CAPACITE_MATERIELLE = "Activité observée sans capacité planifiée"

NIVEAU_FRT_LOCAL_NORMAL = "Réactivité dans la norme habituelle"
NIVEAU_FRT_LOCAL_DEGRADE = "Réactivité dégradée localement"


# Même formule que rang_relatif (nb_inferieurs / n), adaptée à une cible EXTÉRIEURE à la liste de
# référence (rang_relatif suppose la cible incluse dans la série, via l'indexation self-inclusive
# valeurs[:index+1] -- ici la référence est un pool historique poolé, la cible n'y figure jamais).
# Même principe de lecture (0 = plus bas jamais observé, 1 = plus haut jamais observé).
def rang_relatif_vs_reference(valeur_cible, valeurs_reference):
    if valeur_cible is None:
        return None

    valeurs_valides = []
    for valeur in valeurs_reference:
        if valeur is not None:
            valeurs_valides.append(valeur)

    if len(valeurs_valides) == 0:
        return None

    nb_inferieurs = 0
    for valeur in valeurs_valides:
        if valeur < valeur_cible:
            nb_inferieurs = nb_inferieurs + 1
    return nb_inferieurs / len(valeurs_valides)


# Agents à afficher/compter dans la grille de pression : union des assignees de tickets ET des
# agents planifiés sur TOUTES les semaines sélectionnées (Étape 5E.1) -- jamais seulement la
# dernière semaine du planning, sinon un agent présent uniquement sur une semaine antérieure d'une
# sélection multi-semaines disparaîtrait de la capacité cumulée (même défaut que le bug ratio
# corrigé par ailleurs, ici côté "qui compter" plutôt que "combien compter"). plannings_periode :
# liste (date_debut, date_fin, planning), même format que construire_plannings_periode/planning_s2.
def construire_agents_grille_couverture(tickets, plannings_periode):
    agents_a_afficher = []
    for agent in grouper_par(tickets, "assignee"):
        if agent not in agents_a_afficher:
            agents_a_afficher.append(agent)
    for date_debut_semaine, date_fin_semaine, planning_semaine in plannings_periode:
        for agent in planning_semaine:
            if agent not in agents_a_afficher:
                agents_a_afficher.append(agent)

    agents_grille = []
    for agent in agents_a_afficher:
        if agent != NOM_AGENT_DEFAUT:
            agents_grille.append(agent)
    return agents_grille


def statut_creneau_standard(horaires_standard, jour, heure):
    plages = horaires_standard.get(jour, [])
    for debut, fin in plages:
        if debut <= heure < fin:
            return "Couverture requise"

    if len(plages) > 0:
        premiere_debut = plages[0][0]
        derniere_fin = plages[-1][1]
        if premiere_debut <= heure < derniere_fin:
            return "Pause déjeuner"

    return "Hors standard"


def _canal_dominant_couverture(compteur, total):
    if total == 0 or len(compteur) == 0:
        return None, 0
    canal_max = None
    compte_max = 0
    for cle, compte in compteur.items():
        if compte > compte_max:
            compte_max = compte
            canal_max = cle
    return canal_max, compte_max / total * 100


# Grille (jour, heure) -- version corrigée multi-semaines (Étape 5E.1, section 35-36) :
# plannings_periode est la liste de (date_debut, date_fin, planning) déjà utilisée partout ailleurs
# (construire_plannings_periode / planning_s2) -- jamais un seul planning "de la dernière semaine".
# Le ratio de pression est désormais : somme des demandes sur les semaines sélectionnées / SOMME
# des capacités par semaine (jamais la capacité d'une seule semaine appliquée à un volume cumulé
# sur plusieurs semaines). Pour 1 seule semaine sélectionnée (cas normal, non étendu), ce calcul
# est strictement identique au comportement précédent : capacite_cumulee == nb_agents affichés.
# "agents" (prénoms affichés/tooltip) reste l'UNION des présents sur les semaines sélectionnées
# (jamais un doublon "Amine, Amine"), distincte de capacite_cumulee (somme, utilisée uniquement
# comme dénominateur du ratio).
def construire_grille_pression_couverture(tickets_periode, plannings_periode, agents_grille, horaires_standard):
    demandes_par_jour_heure = {}
    canaux_par_jour_heure = {}
    categories_par_jour_heure = {}
    for nom_jour, numero_jour in JOURS_ORDRE_COUVERTURE:
        demandes_par_jour_heure[numero_jour] = {}
        canaux_par_jour_heure[numero_jour] = {}
        categories_par_jour_heure[numero_jour] = {}
        for heure in range(HEURE_DEBUT_PRESSION_COUVERTURE, HEURE_FIN_PRESSION_COUVERTURE):
            demandes_par_jour_heure[numero_jour][heure] = 0
            canaux_par_jour_heure[numero_jour][heure] = {}
            categories_par_jour_heure[numero_jour][heure] = {}

    for ticket in tickets_periode:
        moment = ticket["created_at"]
        jour_ticket = moment.weekday()
        heure_ticket = moment.hour
        if HEURE_DEBUT_PRESSION_COUVERTURE <= heure_ticket < HEURE_FIN_PRESSION_COUVERTURE:
            demandes_par_jour_heure[jour_ticket][heure_ticket] = demandes_par_jour_heure[jour_ticket][heure_ticket] + 1

            compteur_canal = canaux_par_jour_heure[jour_ticket][heure_ticket]
            canal = ticket["via_channel"]
            if canal in compteur_canal:
                compteur_canal[canal] = compteur_canal[canal] + 1
            else:
                compteur_canal[canal] = 1

            compteur_categorie = categories_par_jour_heure[jour_ticket][heure_ticket]
            categorie = categoriser(ticket)
            if categorie in compteur_categorie:
                compteur_categorie[categorie] = compteur_categorie[categorie] + 1
            else:
                compteur_categorie[categorie] = 1

    activite_par_jour_heure = construire_activite_par_jour_heure(
        tickets_periode, HEURE_DEBUT_PRESSION_COUVERTURE, HEURE_FIN_PRESSION_COUVERTURE
    )

    grille = []
    for heure in range(HEURE_DEBUT_PRESSION_COUVERTURE, HEURE_FIN_PRESSION_COUVERTURE):
        for nom_jour, numero_jour in JOURS_ORDRE_COUVERTURE:
            presents_uniques = set()
            capacite_cumulee = 0
            for date_debut_semaine, date_fin_semaine, planning_semaine in plannings_periode:
                presents_semaine = agents_en_poste(planning_semaine, agents_grille, numero_jour, heure)
                capacite_cumulee = capacite_cumulee + len(presents_semaine)
                presents_uniques.update(presents_semaine)
            presents = sorted(presents_uniques)

            demandes = demandes_par_jour_heure[numero_jour][heure]
            statut = statut_creneau_standard(horaires_standard, numero_jour, heure)

            if capacite_cumulee > 0:
                ratio = demandes / capacite_cumulee
            else:
                ratio = None

            if statut == "Couverture requise":
                actifs = activite_observee(activite_par_jour_heure, numero_jour, heure)
                renfort = renfort_non_planifie(presents, actifs)
            else:
                renfort = []

            canal_dominant, part_canal_dominant = _canal_dominant_couverture(
                canaux_par_jour_heure[numero_jour][heure], demandes
            )
            categorie_dominante, part_categorie_dominante = _canal_dominant_couverture(
                categories_par_jour_heure[numero_jour][heure], demandes
            )

            grille.append({
                "jour": nom_jour, "numero_jour": numero_jour, "heure": heure,
                "nb_agents": len(presents), "agents": presents, "capacite_cumulee": capacite_cumulee,
                "demandes": demandes, "ratio": ratio, "statut": statut,
                "canal_dominant": canal_dominant, "part_canal_dominant": part_canal_dominant,
                "categorie_dominante": categorie_dominante, "part_categorie_dominante": part_categorie_dominante,
                "renfort_non_planifie": renfort,
            })
    return grille


# Pool de référence historique (Étape 5E.1, section 5-7) : ratios de pression et médianes de FRT
# local, cellule par cellule, recalculés sur chaque export STRICTEMENT ANTÉRIEUR à date_limite
# (jamais la période affichée, jamais un export futur -- même borne que Tendances 4B/5B). Chaque
# export antérieur est traité comme une semaine isolée (capacite_cumulee == nb_agents de cette
# semaine), pas de double-agrégation avec la logique multi-semaines ci-dessus.
def construire_reference_historique_couverture(exports_disponibles, date_limite):
    ratios_reference = []
    frt_medians_reference = []

    for date_export, chemin in exports_disponibles:
        if date_export >= date_limite:
            continue

        tickets_semaine = charger_tickets(chemin)
        planning_semaine = charger_planning(chemin)
        horaires_standard_semaine = planning_semaine.get(NOM_AGENT_DEFAUT, {})
        date_fin_semaine = date_export + datetime.timedelta(days=6)
        plannings_semaine_liste = [(date_export, date_fin_semaine, planning_semaine)]
        agents_grille_semaine = construire_agents_grille_couverture(tickets_semaine, plannings_semaine_liste)

        grille_semaine = construire_grille_pression_couverture(
            tickets_semaine, plannings_semaine_liste, agents_grille_semaine, horaires_standard_semaine
        )

        for entree in grille_semaine:
            if entree["statut"] != "Couverture requise":
                continue
            if entree["capacite_cumulee"] > 0 and entree["demandes"] >= SEUIL_VOLUME_MIN_PRESSION_COUVERTURE:
                ratios_reference.append(entree["ratio"])

        en_creneau_semaine, _, _ = separer_creneau(tickets_semaine, plannings_semaine_liste)
        tickets_par_cellule_semaine = {}
        for ticket in en_creneau_semaine:
            moment = ticket["created_at"]
            cle = (moment.weekday(), moment.hour)
            if cle in tickets_par_cellule_semaine:
                tickets_par_cellule_semaine[cle].append(ticket)
            else:
                tickets_par_cellule_semaine[cle] = [ticket]

        for cle, tickets_cellule in tickets_par_cellule_semaine.items():
            valeurs_frt = []
            for ticket in tickets_cellule:
                if ticket["first_reply_time_min"] is not None:
                    valeurs_frt.append(ticket["first_reply_time_min"])
            if len(valeurs_frt) >= SEUIL_VOLUME_MIN_FRT_LOCAL_COUVERTURE:
                frt_medians_reference.append(mediane(valeurs_frt))

    return ratios_reference, frt_medians_reference


# Capacité 0 traité comme une catégorie à part (jamais un ratio "infini") -- distingue signal
# faible (1-4 demandes, Étape 5E.1 section 10) de signal matériel (≥5), les deux restant EN DEHORS
# de la taxonomie pression/tension (aucun ratio n'existe pour les qualifier).
def niveau_pression_couverture(demandes, capacite_cumulee, rang_pression):
    if capacite_cumulee == 0:
        if demandes == 0:
            return None
        if demandes < SEUIL_VOLUME_MIN_PRESSION_COUVERTURE:
            return NIVEAU_ACTIVITE_HORS_CAPACITE_FAIBLE
        return NIVEAU_ACTIVITE_HORS_CAPACITE_MATERIELLE

    if demandes < SEUIL_VOLUME_MIN_PRESSION_COUVERTURE:
        return NIVEAU_PRESSION_FAIBLE_VOLUME

    if rang_pression is None:
        return NIVEAU_PRESSION_NON_QUALIFIABLE

    if rang_pression >= SEUIL_RANG_PRESSION_FORTE:
        return NIVEAU_PRESSION_FORTE
    if rang_pression >= SEUIL_RANG_PRESSION_MARQUEE:
        return NIVEAU_PRESSION_MARQUEE
    return NIVEAU_PRESSION_HABITUELLE


# None = non mesurable localement (jamais assimilé à "normal" par défaut -- section 14, un
# échantillon insuffisant ne doit jamais produire une fausse assurance de réactivité correcte).
def niveau_frt_local_couverture(frt_local_n, rang_frt_local):
    if frt_local_n < SEUIL_VOLUME_MIN_FRT_LOCAL_COUVERTURE:
        return None
    if rang_frt_local is None:
        return None
    if rang_frt_local >= SEUIL_RANG_FRT_LOCAL_DEGRADE:
        return NIVEAU_FRT_LOCAL_DEGRADE
    return NIVEAU_FRT_LOCAL_NORMAL


# Tension = convergence stricte, jamais la pression seule (Étape 5E.1, section 12) : catégorie et
# canal n'entrent JAMAIS dans ce calcul (ils expliquent, ne déclenchent pas -- sections 17-18).
def creneau_est_tension_couverture(niveau_pression, niveau_frt_local):
    pression_materielle = niveau_pression in (NIVEAU_PRESSION_MARQUEE, NIVEAU_PRESSION_FORTE)
    frt_degrade = niveau_frt_local == NIVEAU_FRT_LOCAL_DEGRADE
    return pression_materielle and frt_degrade


# Enrichit chaque cellule de la grille avec pression relative, FRT local (médiane, jamais la
# moyenne -- Étape 5E.1 section 15 : robuste aux valeurs extrêmes, démontré sur Juillet Lundi 11h
# où la moyenne locale est tirée vers le haut par quelques tickets isolés alors que la médiane
# reste basse) et statut de tension. tickets_en_creneau_periode doit être borné à la période
# affichée (pas l'historique) -- seule construire_reference_historique_couverture regarde le passé.
def enrichir_grille_pression_tension_couverture(grille, tickets_en_creneau_periode, ratios_reference, frt_medians_reference):
    tickets_par_cellule = {}
    for ticket in tickets_en_creneau_periode:
        moment = ticket["created_at"]
        cle = (moment.weekday(), moment.hour)
        if cle in tickets_par_cellule:
            tickets_par_cellule[cle].append(ticket)
        else:
            tickets_par_cellule[cle] = [ticket]

    reference_pression_suffisante = len(ratios_reference) >= SEUIL_MIN_REFERENCE_HISTORIQUE_COUVERTURE
    reference_frt_suffisante = len(frt_medians_reference) >= SEUIL_MIN_REFERENCE_HISTORIQUE_COUVERTURE

    resultat = []
    for entree in grille:
        nouvelle_entree = dict(entree)

        # La pression/tension n'a de sens que sur un créneau "Couverture requise" -- un horaire
        # fermé par conception (avant ouverture, après fermeture, pause, week-end) n'est jamais
        # une anomalie de capacité, quel que soit son volume : ce volume reste uniquement suivi
        # dans "Demande hors couverture" (Étape 5E.1 -- bug réel trouvé en vérification navigateur,
        # où des créneaux "Hors standard" à 7h/18h+ remontaient à tort comme "activité sans
        # capacité planifiée").
        if entree["statut"] != "Couverture requise":
            nouvelle_entree["rang_pression"] = None
            nouvelle_entree["niveau_pression"] = None
            nouvelle_entree["frt_local_n"] = 0
            nouvelle_entree["frt_local_median"] = None
            nouvelle_entree["rang_frt_local"] = None
            nouvelle_entree["niveau_frt_local"] = None
            nouvelle_entree["est_tension"] = False
            resultat.append(nouvelle_entree)
            continue

        rang_pression = None
        if reference_pression_suffisante and entree["ratio"] is not None:
            rang_pression = rang_relatif_vs_reference(entree["ratio"], ratios_reference)
        nouvelle_entree["rang_pression"] = rang_pression
        nouvelle_entree["niveau_pression"] = niveau_pression_couverture(
            entree["demandes"], entree["capacite_cumulee"], rang_pression
        )

        tickets_cellule = tickets_par_cellule.get((entree["numero_jour"], entree["heure"]), [])
        valeurs_frt = []
        for ticket in tickets_cellule:
            if ticket["first_reply_time_min"] is not None:
                valeurs_frt.append(ticket["first_reply_time_min"])

        frt_local_n = len(valeurs_frt)
        frt_local_median = None
        if frt_local_n > 0:
            frt_local_median = mediane(valeurs_frt)

        rang_frt_local = None
        if reference_frt_suffisante and frt_local_n >= SEUIL_VOLUME_MIN_FRT_LOCAL_COUVERTURE:
            rang_frt_local = rang_relatif_vs_reference(frt_local_median, frt_medians_reference)

        nouvelle_entree["frt_local_n"] = frt_local_n
        nouvelle_entree["frt_local_median"] = frt_local_median
        nouvelle_entree["rang_frt_local"] = rang_frt_local
        nouvelle_entree["niveau_frt_local"] = niveau_frt_local_couverture(frt_local_n, rang_frt_local)
        nouvelle_entree["est_tension"] = creneau_est_tension_couverture(
            nouvelle_entree["niveau_pression"], nouvelle_entree["niveau_frt_local"]
        )
        resultat.append(nouvelle_entree)
    return resultat


# Synthèse "Lecture de couverture" (Étape 5E.1, section 23) -- 2-4 phrases maximum, uniquement
# dérivées des chiffres déjà calculés (aucun nouveau calcul, même principe que
# construire_conclusion_onglet). Jamais de texte générique : le compte de tensions/pressions
# pilote directement la phrase retenue.
def construire_lecture_couverture(nb_tensions, nb_pressions_marquees_absorbees, taux_sla_global, sla_objectif, hors_couverture_significatif):
    phrases = []

    if nb_tensions == 0 and nb_pressions_marquees_absorbees == 0:
        phrases.append("La pression de charge reste dans la norme habituelle sur cette observation.")
    elif nb_tensions == 0:
        phrases.append(
            str(nb_pressions_marquees_absorbees) + " créneau(x) présentent une pression parmi les plus "
            "marquées observées, sans dégradation nette de la première réponse locale."
        )
    else:
        phrase_tension = (
            str(nb_tensions) + " créneau(x) cumulent une pression de charge marquée et une réactivité "
            "dégradée localement -- à examiner en priorité."
        )
        phrases.append(phrase_tension)
        if nb_pressions_marquees_absorbees > 0:
            phrases.append(
                str(nb_pressions_marquees_absorbees) + " autre(s) créneau(x) à pression marquée restent "
                "absorbés, sans dégradation de la réactivité locale."
            )

    if taux_sla_global is not None:
        if taux_sla_global >= sla_objectif:
            phrases.append("La réactivité globale dépasse l'objectif SLA sur la période.")
        else:
            phrases.append("La réactivité globale reste sous l'objectif SLA sur la période.")

    if hors_couverture_significatif:
        phrases.append(
            "Le volume reçu hors couverture est par ailleurs en hausse marquée par rapport à l'historique récent."
        )

    return " ".join(phrases)


# ------------------------------------------------------------------
# Composition Produit -- investigation (Étape 5F.1)
#
# L'audit 5F a établi que 4A est sain et devient l'unique propriétaire de la priorisation
# analytique Produit -- ce module ne recalcule RIEN : il reformate les sorties déjà produites par
# moteur_produit_voie_a/voie_b (jamais un nouveau score, jamais un nouveau seuil analytique), et
# corrige un vrai défaut d'affichage trouvé en audit (le titre de carte produit x composant ne
# montrait que le produit, rendant deux signaux distincts indiscernables -- ex. les deux "Clarté"
# de septembre 2026). Le matching "dossiers associés" réutilise tickets_correspondant_candidat,
# la fonction structurelle que 4A utilise lui-même pour construire ses candidats -- jamais une
# reconstruction parallèle, jamais de correspondance par texte libre sur subject_cluster.
# ------------------------------------------------------------------

TEXTE_PRUDENCE_CAUSALE_PRODUIT = (
    "Chaque signal reste une association observée sur les données disponibles, jamais une cause démontrée."
)


# Décompose le titre en (préfixe produit, texte principal) -- Étape 6F, section 10 : permet à
# l'affichage de styler le préfixe produit différemment du composant (hiérarchie typographique),
# sans dupliquer la logique de grain déjà tranchée ici. prefixe=None pour un grain déjà
# auto-suffisant (composant consolidé ou produit x nature du problème).
def titre_signal_produit_parties(signal):
    if signal["grain"] == GRAIN_PRODUIT_COMPOSANT:
        return signal["_produit"], signal["_composant"]
    return None, signal["sujet"]


# Corrige le seul cas où sujet (4A, verrouillé) ne suffit pas à distinguer deux signaux à l'écran :
# un grain produit x composant partage son "sujet" (juste le nom produit) avec d'éventuels autres
# signaux produit x composant du même produit sur un composant différent. Les grains composant et
# produit x nature du problème restent inchangés (déjà distinguables via sujet seul).
def titre_signal_produit(signal):
    prefixe, principal = titre_signal_produit_parties(signal)
    if prefixe is None:
        return principal
    return prefixe + " — " + principal


# Reconstruit la clé structurelle (produit, composant, issue_type) d'un signal déjà construit par
# 4A, pour la seule identité indexable par la voie A -- jamais reformée à partir d'un
# texte affiché, toujours à partir des champs privés déjà calculés par construire_signal_produit_
# voie_a (_produit/_composant/_issue_type). Retourne (cle, grain) réutilisables tels quels par
# tickets_correspondant_candidat.
def cle_signal_produit(signal):
    grain = signal["grain"]
    if grain == GRAIN_COMPOSANT:
        return signal["sujet"], grain
    if grain == GRAIN_PRODUIT_COMPOSANT:
        return (signal["_produit"], signal["_composant"]), grain
    if grain == GRAIN_PRODUIT_ISSUE:
        return (signal["_produit"], signal["_issue_type"]), grain
    return None, grain


# "Dossiers associés" -- même principe de non-duplication qu'ailleurs (Étape 5F.1, section 29-30) :
# réutilise tickets_correspondant_candidat, la fonction structurelle DÉJÀ utilisée par 4A pour
# construire ses propres candidats. Un signal consolidé (ex. "Batterie / charge" au grain composant
# ayant absorbé des enfants produit x composant) retrouve correctement TOUS les tickets concernés :
# le grain composant matche déjà, par construction, tous les produits partageant ce composant --
# aucune reconstruction de la consolidation elle-même n'est nécessaire ici.
def construire_dossiers_associes_produit(signal, tickets_sav_produit_periode):
    cle, grain = cle_signal_produit(signal)
    if cle is None:
        return []
    return tickets_correspondant_candidat(tickets_sav_produit_periode, cle, grain)


# Lecture Produit -- 2-4 phrases, dérivées UNIQUEMENT des compteurs déjà produits par 4A
# (nb_prioritaires_avant_plafond / nb_a_surveiller_avant_plafond, déjà exposés par
# moteur_produit_voie_a) et du contexte SAV descriptif (part du volume). Aucun nouveau score,
# aucun nouveau seuil analytique -- même principe que construire_lecture_couverture (5E.1) et
# construire_conclusion_onglet historique. Nomme le signal le mieux prouvé (Priorité principale,
# premier de la liste déjà triée par 4A) uniquement quand un tel signal existe -- jamais "produit
# le plus problématique"/"cause principale", toujours "le niveau de preuve le plus complet".
def construire_lecture_produit(prioritaires_affiches, nb_prioritaires_avant_plafond,
                                a_surveiller_affiches, nb_a_surveiller_avant_plafond,
                                nb_dossiers_individuels, part_sav_pct):
    phrases = []
    nb_prioritaires_affiches = len(prioritaires_affiches)

    if nb_prioritaires_avant_plafond == 0:
        if part_sav_pct is not None:
            phrases.append(
                "Le SAV Produit représente " + formater_pourcentage(part_sav_pct) + " des demandes de la "
                "période, mais aucun signal n'atteint actuellement le niveau de priorité."
            )
        else:
            phrases.append("Aucun signal Produit n'atteint actuellement le niveau de priorité.")
    else:
        signal_principal = None
        for signal in prioritaires_affiches:
            if signal["niveau_priorite"] == "Priorité principale":
                signal_principal = signal
                break

        texte = (
            str(nb_prioritaires_avant_plafond) + " signal(aux) Produit présente(nt) une convergence "
            "suffisante pour être investigué(s)"
        )
        if signal_principal is not None:
            texte = texte + ", dont " + titre_signal_produit(signal_principal) + " avec le niveau de preuve le plus complet"
        if nb_prioritaires_avant_plafond > nb_prioritaires_affiches:
            texte = texte + " (" + str(nb_prioritaires_affiches) + " affiché(s) ici)"
        phrases.append(texte + ".")

    if nb_a_surveiller_avant_plafond > 0:
        phrases.append(
            str(nb_a_surveiller_avant_plafond) + " sujet(s) supplémentaire(s) reste(nt) à surveiller, "
            "avec une preuve encore partielle."
        )

    if nb_dossiers_individuels > 0:
        phrases.append(
            str(nb_dossiers_individuels) + " dossier(s) individuel(s) présente(nt) des caractéristiques "
            "justifiant une lecture humaine, indépendamment des signaux ci-dessus."
        )

    return " ".join(phrases)


# Remplace construire_insight_resolution (Étape 5F.1, section 5) : strictement descriptif, jamais
# "signal de défaut matériel réel à corriger". Le type de résolution dominant reste une donnée
# utile à transmettre (ce que les dossiers ont nécessité), sans verdict sur la cause.
def construire_texte_resolution_produit(lignes_resolution_triees, total_sav):
    if len(lignes_resolution_triees) == 0 or total_sav == 0:
        return None

    plus_frequente = lignes_resolution_triees[0]
    part = plus_frequente["Tickets"] / total_sav * 100
    type_resolution = plus_frequente["Type de résolution"]

    return (
        str(round(part)) + " % des dossiers SAV Produit observés ont donné lieu à : " + str(type_resolution) + "."
    )


# Remplace le texte "SAV récurrents" (Étape 5F.1, section 6) : retire "signal de défaut structurel
# plutôt qu'un cas isolé" -- reste factuel (combien de clients, concentration produit/composant),
# sans conclusion Produit automatique à partir du seul comptage prior_sav_count >= 1.
def construire_texte_sav_recurrents_produit(nb_recurrents, part_recurrents_pct, produit_principal, composant_principal):
    return (
        str(nb_recurrents) + " tickets (" + str(round(part_recurrents_pct)) + " % du SAV produit) concernent "
        "un client ayant déjà eu au moins un autre dossier SAV avant celui-ci. Concentré sur **"
        + produit_principal["Produit"] + "** (" + str(produit_principal["SAV récurrents"]) + " cas) et le "
        "composant **" + composant_principal["Composant"] + "** (" + str(composant_principal["SAV récurrents"]) + " cas)."
    )


# ------------------------------------------------------------------
# Composition Livraison — investigation (Étape 5G.1)
# ------------------------------------------------------------------
# Conclusion de l'audit 5G : le moteur 4C (moteur_livraison_voie_a) est sain et devient l'unique
# propriétaire de la priorisation analytique Livraison. Ce module ne recalcule RIEN : il reformate
# les sorties déjà produites par 4C (jamais un nouveau score, jamais un nouveau seuil analytique
# concurrent), et ajoute une capacité "dossiers associés" symétrique à celle de Produit (5F.1). Le
# grain Livraison étant unique (subject_cluster, pas de tuple), le matching est un filtre direct sur
# ce champ -- exactement le filtre déjà utilisé en interne par 4C lui-même (voir
# construire_niveaux_historiques_livraison et moteur_livraison_voie_a), jamais une correspondance
# par texte libre ni une reconstruction de moteur.
# ------------------------------------------------------------------

TEXTE_COUT_INDISPONIBLE_LIVRAISON = (
    "Impact financier non mesurable avec les données disponibles : les dossiers Livraison ne "
    "disposent pas actuellement d'un identifiant commande exploitable pour ce calcul."
)


# Lecture Livraison : juxtapose explicitement ACTIVITÉ (poids de Livraison sur la période, texte
# déjà produit par construire_lecture_activite_livraison, verrouillé 4C, section 15 de l'audit) et
# SIGNAL (compteurs déjà produits par moteur_livraison_voie_a) -- les deux dimensions restent
# distinctes, jamais fusionnées en une seule affirmation ("cette hausse cause ce signal").
def construire_lecture_livraison(observation_activite, nb_prioritaires_avant_plafond, nb_a_surveiller_avant_plafond):
    phrases = [observation_activite]

    if nb_prioritaires_avant_plafond == 0 and nb_a_surveiller_avant_plafond == 0:
        phrases.append("Aucun motif ne présente actuellement une convergence suffisante pour être investigué.")
    else:
        if nb_prioritaires_avant_plafond == 1:
            phrases.append("Un motif présente une convergence suffisante pour être investigué.")
        elif nb_prioritaires_avant_plafond > 1:
            phrases.append(
                str(nb_prioritaires_avant_plafond)
                + " motifs présentent une convergence suffisante pour être investigués."
            )
        if nb_a_surveiller_avant_plafond > 0:
            phrases.append(
                str(nb_a_surveiller_avant_plafond) + " motif(s) supplémentaire(s) reste(nt) à surveiller, "
                "avec une preuve encore partielle."
            )

    return " ".join(phrases)


# "Dossiers associés" Livraison : grain unique subject_cluster, identique au filtre déjà utilisé en
# interne par 4C (ticket["subject_cluster"] == sujet) -- aucune reconstruction de moteur, aucun
# matching par texte libre.
def construire_dossiers_associes_livraison(signal, tickets_livraison_periode):
    sujet = signal["sujet"]
    dossiers = []
    for ticket in tickets_livraison_periode:
        if ticket["subject_cluster"] == sujet:
            dossiers.append(ticket)
    return dossiers


# Croisement motif x issue finale, pour la section "Explorer les conséquences" (item 31 de l'audit) :
# réutilise le grain motif et distribution_issues_livraison (verrouillé 4C) -- une simple agrégation
# d'exploration, jamais utilisée pour trancher une priorité (4C reste seul décisionnaire).
def construire_croisement_motif_issue_livraison(tickets_livraison_periode):
    sujets = grouper_par(tickets_livraison_periode, "subject_cluster")
    lignes = []
    for sujet, tickets_sujet in sujets.items():
        distribution = distribution_issues_livraison(tickets_sujet)
        relances_moyennes = moyenne(tickets_sujet, "nombre_relances")
        issue_principale = None
        part_issue_principale_pct = None
        if len(distribution) > 0:
            issue_principale = distribution[0]["issue"]
            part_issue_principale_pct = distribution[0]["part_pct"]
        lignes.append({
            "sujet": sujet,
            "n": len(tickets_sujet),
            "relances_moyennes": relances_moyennes,
            "issue_principale": issue_principale,
            "part_issue_principale_pct": part_issue_principale_pct,
        })
    return lignes


# ------------------------------------------------------------------
# Composition Avant-vente — parcours & achats observés (Étape 5H.1)
# ------------------------------------------------------------------
# Conclusion de l'audit 5H : le moteur 4D (resoudre_achats_observes_avant_vente / analyser_parcours_
# rdv / moteur_avant_vente_motifs) est sain et devient l'unique propriétaire de l'attribution
# Avant-vente. Ce module ne recalcule RIEN : il reformate les sorties déjà produites par 4D (jamais
# une nouvelle recherche Shopify locale, jamais premiere_commande_apres). "Contacts associés" et
# "Achats associés" filtrent exclusivement resultats_achats_av (déjà déduplicué par 4D, règle du
# contact le plus récent déjà appliquée) sur le grain motif (subject_cluster) -- même principe
# structurel que construire_dossiers_associes_livraison (5G.1).
# ------------------------------------------------------------------


# Lecture Avant-vente : juxtapose ACTIVITÉ (poids Avant-vente sur la période, texte déjà produit par
# construire_lecture_activite_avant_vente, verrouillé 4D) et SIGNAL (compteurs déjà produits par
# moteur_avant_vente_motifs) -- jamais fusionnés en une seule affirmation causale.
def construire_lecture_avant_vente(observation_activite, nb_opportunites_avant_plafond, nb_a_surveiller_avant_plafond):
    phrases = [observation_activite]

    if nb_opportunites_avant_plafond == 0 and nb_a_surveiller_avant_plafond == 0:
        phrases.append("Aucune opportunité ne se détache actuellement sur cette période.")
    else:
        if nb_opportunites_avant_plafond == 1:
            phrases.append("Une opportunité présente une convergence suffisante pour être investiguée.")
        elif nb_opportunites_avant_plafond > 1:
            phrases.append(
                str(nb_opportunites_avant_plafond)
                + " opportunités présentent une convergence suffisante pour être investiguées."
            )
        if nb_a_surveiller_avant_plafond > 0:
            phrases.append(
                str(nb_a_surveiller_avant_plafond) + " motif(s) supplémentaire(s) reste(nt) à surveiller, "
                "avec une preuve encore partielle."
            )

    return " ".join(phrases)


# "Contacts associés" d'une opportunité : filtre resultats_achats_av (liste de (ticket, commande,
# plusieurs_commandes) déjà produite par resoudre_achats_observes_avant_vente) sur le grain motif
# exact -- aucune reconstruction de moteur, aucun matching par texte libre.
def construire_contacts_associes_avant_vente(signal, resultats_achats_av):
    sujet = signal["sujet"]
    contacts = []
    for ticket, commande, plusieurs_commandes in resultats_achats_av:
        if ticket["subject_cluster"] == sujet:
            contacts.append((ticket, commande, plusieurs_commandes))
    return contacts


# "Achats associés" d'une opportunité : sous-ensemble des contacts associés pour lesquels 4D a
# effectivement attribué une commande -- jamais une nouvelle recherche Shopify, uniquement un filtre
# sur une structure déjà produite par 4D (Étape 5H, section 26).
def construire_achats_associes_avant_vente(contacts_associes):
    achats = []
    for ticket, commande, plusieurs_commandes in contacts_associes:
        if commande is not None:
            achats.append((ticket, commande))
    return achats


# Table descriptive par motif (vue descriptive secondaire, jamais utilisée pour l'éligibilité 4D --
# celle-ci reste dans moteur_avant_vente_motifs). La demande de RDV elle-même est exclue, comme dans
# 4D (SUJET_DEMANDE_RDV), pour ne pas raconter deux fois la même histoire que la section Parcours.
def construire_table_sujets_avant_vente(tickets_avant_vente_periode, resultats_achats_av):
    sujets_presents = set()
    for ticket in tickets_avant_vente_periode:
        if ticket["subject_cluster"] != SUJET_DEMANDE_RDV:
            sujets_presents.add(ticket["subject_cluster"])

    lignes = []
    for sujet in sujets_presents:
        resultats_sujet = []
        for ticket, commande, plusieurs_commandes in resultats_achats_av:
            if ticket["subject_cluster"] == sujet:
                resultats_sujet.append((ticket, commande, plusieurs_commandes))
        stats = calculer_stats_achat_observe(resultats_sujet)
        lignes.append({
            "sujet": sujet,
            "n": stats["n_contacts"],
            "achat_observe_pct": stats["taux_pct"],
            "n_achats": stats["n_achats"],
            "panier_moyen": stats["panier_moyen"],
        })
    return lignes


# Table descriptive par pays (vue descriptive secondaire) -- jamais croisée avec l'agent (Étape 5H,
# section 34 : le croisement agent x pays est celui qui posait un risque d'usage RH, pas le pays
# seul).
def construire_table_pays_avant_vente(tickets_avant_vente_periode, resultats_achats_av):
    pays_presents = set()
    for ticket in tickets_avant_vente_periode:
        pays_presents.add(ticket["country"])

    lignes = []
    for pays in pays_presents:
        resultats_pays = []
        for ticket, commande, plusieurs_commandes in resultats_achats_av:
            if ticket["country"] == pays:
                resultats_pays.append((ticket, commande, plusieurs_commandes))
        stats = calculer_stats_achat_observe(resultats_pays)
        lignes.append({
            "pays": pays,
            "n": stats["n_contacts"],
            "achat_observe_pct": stats["taux_pct"],
            "n_achats": stats["n_achats"],
        })
    return lignes


# ------------------------------------------------------------------
# Composition Impact & confiance (Étape 5I.1)
# ------------------------------------------------------------------
# Conclusion de l'audit 5I : le moteur 4E (NPS, alignement, prudence d'échantillon) est sain -- le
# défaut identifié était un défaut de COMPOSITION UI : la "Lecture de confiance" utilisait toujours
# le DERNIER mois du fichier NPS (index_dernier_mois = len(historique)-1), indépendamment de la
# période sélectionnée dans la barre latérale, alors que Vue d'ensemble (5A, verrouillé, non modifié
# ici) calait déjà correctement son propre NPS sur la période sélectionnée. Ce module réutilise
# EXACTEMENT le même principe pour Impact & confiance -- jamais un mois "le plus proche", jamais un
# repli sur le dernier mois disponible : si le mois de la période sélectionnée n'a pas d'observation
# NPS, l'absence est honnête (aucune donnée), jamais masquée par une valeur d'un autre mois.
# ------------------------------------------------------------------


# Identifie l'observation NPS mensuelle correspondant EXACTEMENT au mois de la période sélectionnée
# (même principe que le calage déjà fait en Vue d'ensemble, Étape 5A -- jamais un mois approximatif,
# jamais de forward/backward-fill). Retourne None si aucune observation ne correspond.
def identifier_observation_nps_periode(historique_nps_mensuel, date_debut):
    cle_mois_periode = date_debut.strftime("%Y-%m")
    for index in range(len(historique_nps_mensuel)):
        if historique_nps_mensuel[index]["cle_mois"] == cle_mois_periode:
            return index
    return None


TEXTE_SENSIBILITE_PETIT_ECHANTILLON_NPS = (
    "Sur un petit échantillon, quelques réponses reclassées suffiraient à déplacer sensiblement le score."
)


# Rappel méthodologique (jamais un faux intervalle de confiance) affiché uniquement quand le volume
# du mois est déjà signalé comme faible ou encore trop récent pour être situé (Étape 5I, section 14).
def texte_sensibilite_echantillon_nps(etat_prudence):
    if etat_prudence == ETAT_PRUDENCE_VOLUME_FAIBLE or etat_prudence == ETAT_PRUDENCE_PREMIERE_OBSERVATION:
        return TEXTE_SENSIBILITE_PETIT_ECHANTILLON_NPS
    return None


TEXTE_CAVEAT_RECOUVREMENT_COUT = (
    "« Coût direct » et « Coût garantie » se recoupent partiellement (un remplacement sous garantie "
    "compte dans les deux) -- ne jamais additionner ces deux montants."
)


# Lecture Impact & confiance : combine confiance client (NPS + alignement, si disponible pour la
# période) et impact financier (coût direct de la période), en phrases distinctes -- jamais fusionnés
# en une seule affirmation, jamais un score composite (Étape 5I, section 9/28).
def construire_lecture_impact_confiance(item_nps, texte_alignement, montant_total_pertes, montant_evitable):
    phrases = []

    if item_nps is None:
        phrases.append("Aucune donnée NPS exploitable pour cette période.")
    else:
        phrases.append(
            "Le NPS de la période est de " + formater_nps_entier(item_nps["nps"]) + " sur "
            + str(item_nps["n"]) + " réponse(s)."
        )
        if texte_alignement is not None:
            phrases.append(texte_alignement)

    if montant_total_pertes > 0:
        phrase_cout = (
            "Le coût direct observé/estimé sur cette période s'élève à " + formater_montant(montant_total_pertes) + "."
        )
        if montant_evitable > 0:
            phrase_cout = phrase_cout + (
                " " + formater_pourcentage(montant_evitable / montant_total_pertes * 100)
                + " de ce montant est classé potentiellement évitable."
            )
        phrases.append(phrase_cout)
    else:
        phrases.append("Aucun coût direct identifié sur cette période.")

    return " ".join(phrases)
